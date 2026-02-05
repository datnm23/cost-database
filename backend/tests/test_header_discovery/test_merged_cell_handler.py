"""
Tests for MergedCellHandler
"""
import pytest
import os
import tempfile
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from app.services.header_discovery.merged_cell_handler import (
    MergedCellHandler,
    MergedHeaderResult,
    get_merged_cell_handler
)


class TestMergedCellHandler:
    """Test cases for MergedCellHandler."""

    @pytest.fixture
    def handler(self):
        return MergedCellHandler()

    @pytest.fixture
    def simple_excel(self):
        """Create a simple Excel file without merged cells."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"

            # Simple header row
            headers = ["STT", "Mô tả", "Đơn vị", "Khối lượng", "Đơn giá", "Thành tiền"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Data rows
            ws.cell(row=2, column=1, value=1)
            ws.cell(row=2, column=2, value="Item 1")
            ws.cell(row=2, column=3, value="m3")
            ws.cell(row=2, column=4, value=10)
            ws.cell(row=2, column=5, value=1000000)
            ws.cell(row=2, column=6, value=10000000)

            wb.save(f.name)
            yield f.name

        os.unlink(f.name)

    @pytest.fixture
    def merged_excel(self):
        """Create an Excel file with merged header cells."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"

            # First header row with merged cells
            ws.cell(row=1, column=1, value="STT")
            ws.cell(row=1, column=2, value="Mô tả")
            ws.cell(row=1, column=3, value="Vật liệu")
            ws.merge_cells('C1:D1')  # Merge "Vật liệu" across 2 columns
            ws.cell(row=1, column=5, value="Nhân công")
            ws.merge_cells('E1:F1')  # Merge "Nhân công" across 2 columns

            # Second header row
            ws.cell(row=2, column=1, value="")
            ws.cell(row=2, column=2, value="")
            ws.cell(row=2, column=3, value="Đơn giá")
            ws.cell(row=2, column=4, value="Thành tiền")
            ws.cell(row=2, column=5, value="Đơn giá")
            ws.cell(row=2, column=6, value="Thành tiền")

            # Data row
            ws.cell(row=3, column=1, value=1)
            ws.cell(row=3, column=2, value="Bê tông M200")
            ws.cell(row=3, column=3, value=500000)
            ws.cell(row=3, column=4, value=5000000)
            ws.cell(row=3, column=5, value=200000)
            ws.cell(row=3, column=6, value=2000000)

            wb.save(f.name)
            yield f.name

        os.unlink(f.name)

    # Test simple headers (no merge)
    def test_process_simple_headers(self, handler, simple_excel):
        """Test processing simple non-merged headers."""
        result = handler.process_headers(
            file_path=simple_excel,
            sheet_name="Data",
            header_row=0,
            num_columns=6
        )

        assert result.is_merged is False
        assert result.header_depth == 1
        assert len(result.column_names) == 6
        assert "STT" in result.column_names
        assert "Mô tả" in result.column_names

    # Test merged headers
    def test_process_merged_headers(self, handler, merged_excel):
        """Test processing merged header cells."""
        result = handler.process_headers(
            file_path=merged_excel,
            sheet_name="Data",
            header_row=0,
            num_columns=6
        )

        assert result.is_merged is True
        assert result.header_depth >= 2
        assert len(result.column_names) == 6

        # Check flattened names
        # Column 3 should be "Vật liệu - Đơn giá"
        assert "Vật liệu" in result.column_names[2]
        assert "Đơn giá" in result.column_names[2]

        # Column 4 should be "Vật liệu - Thành tiền"
        assert "Vật liệu" in result.column_names[3]
        assert "Thành tiền" in result.column_names[3]

    # Test error handling
    def test_process_nonexistent_file(self, handler):
        """Test processing non-existent file."""
        result = handler.process_headers(
            file_path="/nonexistent/file.xlsx",
            sheet_name="Data",
            header_row=0,
            num_columns=5
        )

        # Should return fallback column names
        assert result.is_merged is False
        assert len(result.column_names) == 5
        assert all("Column_" in name for name in result.column_names)

    def test_process_nonexistent_sheet(self, handler, simple_excel):
        """Test processing non-existent sheet."""
        result = handler.process_headers(
            file_path=simple_excel,
            sheet_name="NonExistent",
            header_row=0,
            num_columns=5
        )

        # Should return fallback column names
        assert result.is_merged is False
        assert len(result.column_names) == 5


class TestFlattenHeaders:
    """Test header flattening logic."""

    @pytest.fixture
    def handler(self):
        return MergedCellHandler()

    def test_flatten_single_row(self, handler):
        """Test flattening single row header."""
        matrix = [["A", "B", "C"]]
        result = handler._flatten_headers(matrix)

        assert result == ["A", "B", "C"]

    def test_flatten_two_rows(self, handler):
        """Test flattening two-row header."""
        matrix = [
            ["Group1", "Group1", "Group2"],
            ["SubA", "SubB", "SubC"]
        ]
        result = handler._flatten_headers(matrix)

        assert result == ["Group1 - SubA", "Group1 - SubB", "Group2 - SubC"]

    def test_flatten_with_empty_cells(self, handler):
        """Test flattening with empty cells."""
        matrix = [
            ["", "Header2", ""],
            ["SubA", "", "SubC"]
        ]
        result = handler._flatten_headers(matrix)

        assert result[0] == "SubA"
        assert result[1] == "Header2"
        assert result[2] == "SubC"

    def test_flatten_removes_duplicates(self, handler):
        """Test that duplicate values in same column are removed."""
        matrix = [
            ["Same", "Different"],
            ["Same", "Value"]  # First column has same value
        ]
        result = handler._flatten_headers(matrix)

        # Should not repeat "Same"
        assert result[0] == "Same"
        assert result[1] == "Different - Value"

    def test_flatten_empty_matrix(self, handler):
        """Test flattening empty matrix."""
        result = handler._flatten_headers([])
        assert result == []


class TestThreeLevelMergedHeaders:
    """Test three-level (deep) merged headers."""

    @pytest.fixture
    def handler(self):
        return MergedCellHandler()

    @pytest.fixture
    def three_level_excel(self):
        """Create an Excel file with 3-level merged headers."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"

            # Level 1: Top level category
            ws.cell(row=1, column=1, value="STT")
            ws.merge_cells('A1:A3')  # STT spans all 3 rows vertically
            ws.cell(row=1, column=2, value="Chi phí xây dựng")
            ws.merge_cells('B1:E1')  # Spans 4 columns

            # Level 2: Sub-categories
            ws.cell(row=2, column=2, value="Vật liệu")
            ws.merge_cells('B2:C2')  # Spans 2 columns
            ws.cell(row=2, column=4, value="Nhân công")
            ws.merge_cells('D2:E2')  # Spans 2 columns

            # Level 3: Detail columns
            ws.cell(row=3, column=2, value="Đơn giá")
            ws.cell(row=3, column=3, value="Thành tiền")
            ws.cell(row=3, column=4, value="Đơn giá")
            ws.cell(row=3, column=5, value="Thành tiền")

            # Data row
            ws.cell(row=4, column=1, value=1)
            ws.cell(row=4, column=2, value=500000)
            ws.cell(row=4, column=3, value=5000000)
            ws.cell(row=4, column=4, value=200000)
            ws.cell(row=4, column=5, value=2000000)

            wb.save(f.name)
            yield f.name

        os.unlink(f.name)

    def test_process_three_level_headers(self, handler, three_level_excel):
        """Test processing 3-level merged headers."""
        result = handler.process_headers(
            file_path=three_level_excel,
            sheet_name="Data",
            header_row=0,
            num_columns=5
        )

        assert result.is_merged is True
        assert result.header_depth >= 2  # Should detect multi-level
        assert len(result.column_names) == 5

        # First column should just be STT (vertically merged)
        assert "STT" in result.column_names[0]


class TestVerticalMergedCells:
    """Test vertically merged cells (row-spanning)."""

    @pytest.fixture
    def handler(self):
        return MergedCellHandler()

    @pytest.fixture
    def vertical_merge_excel(self):
        """Create Excel with vertical merges."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"

            # Header with vertical merge
            ws.cell(row=1, column=1, value="STT")
            ws.merge_cells('A1:A2')  # Vertical merge
            ws.cell(row=1, column=2, value="Mô tả")
            ws.merge_cells('B1:B2')  # Vertical merge
            ws.cell(row=1, column=3, value="Chi phí")
            ws.merge_cells('C1:D1')  # Horizontal merge
            ws.cell(row=2, column=3, value="Vật liệu")
            ws.cell(row=2, column=4, value="Nhân công")

            # Data
            ws.cell(row=3, column=1, value=1)
            ws.cell(row=3, column=2, value="Item 1")
            ws.cell(row=3, column=3, value=1000000)
            ws.cell(row=3, column=4, value=500000)

            wb.save(f.name)
            yield f.name

        os.unlink(f.name)

    def test_vertical_merge_headers(self, handler, vertical_merge_excel):
        """Test processing headers with vertical merges."""
        result = handler.process_headers(
            file_path=vertical_merge_excel,
            sheet_name="Data",
            header_row=0,
            num_columns=4
        )

        assert result.is_merged is True
        assert len(result.column_names) == 4

        # Vertically merged cells should appear once
        assert "STT" in result.column_names[0]
        assert "Mô tả" in result.column_names[1]
        # Chi phí - Vật liệu and Chi phí - Nhân công
        assert "Chi phí" in result.column_names[2]
        assert "Chi phí" in result.column_names[3]


class TestMixedHeaders:
    """Test mixed merged and simple headers."""

    @pytest.fixture
    def handler(self):
        return MergedCellHandler()

    @pytest.fixture
    def mixed_excel(self):
        """Create Excel with mixed merged and simple headers."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"

            # Mixed header - some merged, some simple
            ws.cell(row=1, column=1, value="STT")
            ws.cell(row=1, column=2, value="Mô tả công việc")
            ws.cell(row=1, column=3, value="Đơn vị")
            ws.cell(row=1, column=4, value="Chi phí")
            ws.merge_cells('D1:F1')  # Only this is merged

            ws.cell(row=2, column=1, value="")
            ws.cell(row=2, column=2, value="")
            ws.cell(row=2, column=3, value="")
            ws.cell(row=2, column=4, value="Vật liệu")
            ws.cell(row=2, column=5, value="Nhân công")
            ws.cell(row=2, column=6, value="Máy")

            # Data
            ws.cell(row=3, column=1, value=1)
            ws.cell(row=3, column=2, value="Đào đất")
            ws.cell(row=3, column=3, value="m3")
            ws.cell(row=3, column=4, value=100000)
            ws.cell(row=3, column=5, value=50000)
            ws.cell(row=3, column=6, value=30000)

            wb.save(f.name)
            yield f.name

        os.unlink(f.name)

    def test_mixed_headers(self, handler, mixed_excel):
        """Test processing mixed merged and simple headers."""
        result = handler.process_headers(
            file_path=mixed_excel,
            sheet_name="Data",
            header_row=0,
            num_columns=6
        )

        assert len(result.column_names) == 6
        # Simple columns should remain simple
        assert result.column_names[0] == "STT"
        assert "Mô tả" in result.column_names[1]
        assert "Đơn vị" in result.column_names[2]
        # Merged columns should be flattened
        assert "Chi phí" in result.column_names[3] or "Vật liệu" in result.column_names[3]


class TestLargeHeaders:
    """Test headers with many columns."""

    @pytest.fixture
    def handler(self):
        return MergedCellHandler()

    @pytest.fixture
    def wide_excel(self):
        """Create Excel with many columns."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"

            # Create 20 columns
            for col in range(1, 21):
                ws.cell(row=1, column=col, value=f"Header_{col}")
                ws.cell(row=2, column=col, value=col * 100)

            wb.save(f.name)
            yield f.name

        os.unlink(f.name)

    def test_many_columns(self, handler, wide_excel):
        """Test processing headers with many columns."""
        result = handler.process_headers(
            file_path=wide_excel,
            sheet_name="Data",
            header_row=0,
            num_columns=20
        )

        assert result.is_merged is False
        assert len(result.column_names) == 20
        assert result.column_names[0] == "Header_1"
        assert result.column_names[19] == "Header_20"


class TestSpecialCharacterHeaders:
    """Test headers with special characters."""

    @pytest.fixture
    def handler(self):
        return MergedCellHandler()

    @pytest.fixture
    def special_char_excel(self):
        """Create Excel with special characters in headers."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"

            headers = [
                "STT",
                "Mô tả (chi tiết)",
                "Đơn vị\ntính",  # Newline in header
                "Khối lượng [m³]",
                "Đơn giá / VNĐ",
            ]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                ws.cell(row=2, column=col, value=col)

            wb.save(f.name)
            yield f.name

        os.unlink(f.name)

    def test_special_characters(self, handler, special_char_excel):
        """Test processing headers with special characters."""
        result = handler.process_headers(
            file_path=special_char_excel,
            sheet_name="Data",
            header_row=0,
            num_columns=5
        )

        assert len(result.column_names) == 5
        assert "STT" in result.column_names[0]
        # Headers with parentheses should be preserved
        assert "chi tiết" in result.column_names[1]


class TestGetMergedCellHandler:
    """Test singleton factory function."""

    def test_singleton(self):
        """Test that factory returns singleton."""
        handler1 = get_merged_cell_handler()
        handler2 = get_merged_cell_handler()
        assert handler1 is handler2
