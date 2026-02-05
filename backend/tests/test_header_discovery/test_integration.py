"""
Integration tests for HeaderDiscoveryService
"""
import pytest
import os
import tempfile
from openpyxl import Workbook
from app.services.header_discovery import (
    get_header_discovery_service,
    HeaderDiscoveryResult
)


class TestHeaderDiscoveryIntegration:
    """Integration tests for the full header discovery workflow."""

    @pytest.fixture
    def discovery_service(self):
        return get_header_discovery_service()

    @pytest.fixture
    def standard_boq_excel(self):
        """Create a standard BOQ Excel file."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()

            # Create Summary sheet (should be skipped)
            ws_summary = wb.active
            ws_summary.title = "Summary"
            ws_summary.cell(row=1, column=1, value="Project Summary")
            ws_summary.cell(row=2, column=1, value="Total: 1,000,000,000 VND")

            # Create BOQ sheet (should be selected)
            ws_boq = wb.create_sheet("BOQ")
            headers = ["STT", "Mô tả công việc", "Đơn vị", "Khối lượng", "Đơn giá", "Thành tiền"]
            for col, header in enumerate(headers, 1):
                ws_boq.cell(row=1, column=col, value=header)

            # Add data rows
            data = [
                [1, "Đào đất hố móng", "m3", 100, 50000, 5000000],
                [2, "Bê tông lót M100", "m3", 20, 1200000, 24000000],
                [3, "Bê tông móng M250", "m3", 50, 1500000, 75000000],
            ]
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws_boq.cell(row=row_idx, column=col_idx, value=value)

            # Create Notes sheet (should be skipped)
            ws_notes = wb.create_sheet("Notes")
            ws_notes.cell(row=1, column=1, value="Important Notes")

            wb.save(f.name)
            yield f.name

        os.unlink(f.name)

    @pytest.fixture
    def header_at_row_5_excel(self):
        """Create an Excel file with header at row 5."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"

            # Project info rows (0-4)
            ws.cell(row=1, column=1, value="DỰ ÁN: Tòa nhà ABC")
            ws.cell(row=2, column=1, value="CHỦ ĐẦU TƯ: Công ty XYZ")
            ws.cell(row=3, column=1, value="ĐỊA ĐIỂM: Hà Nội")
            ws.cell(row=4, column=1, value="")
            ws.cell(row=5, column=1, value="BẢNG DỰ TOÁN CHI TIẾT")

            # Header row (index 5, Excel row 6)
            headers = ["STT", "Hạng mục", "ĐVT", "KL", "Đơn giá", "Thành tiền"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=6, column=col, value=header)

            # Data rows
            ws.cell(row=7, column=1, value=1)
            ws.cell(row=7, column=2, value="Item 1")
            ws.cell(row=7, column=3, value="m3")
            ws.cell(row=7, column=4, value=10)
            ws.cell(row=7, column=5, value=1000000)
            ws.cell(row=7, column=6, value=10000000)

            wb.save(f.name)
            yield f.name

        os.unlink(f.name)

    @pytest.fixture
    def english_headers_excel(self):
        """Create an Excel file with English headers."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "BOQ"

            headers = ["No.", "Description", "Unit", "Quantity", "Unit Price", "Amount"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Data
            ws.cell(row=2, column=1, value=1)
            ws.cell(row=2, column=2, value="Excavation")
            ws.cell(row=2, column=3, value="m3")
            ws.cell(row=2, column=4, value=100)
            ws.cell(row=2, column=5, value=50000)
            ws.cell(row=2, column=6, value=5000000)

            wb.save(f.name)
            yield f.name

        os.unlink(f.name)

    @pytest.fixture
    def merged_header_excel(self):
        """Create an Excel file with merged headers."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "BOQ"

            # First row with merged cells
            ws.cell(row=1, column=1, value="STT")
            ws.merge_cells('A1:A2')
            ws.cell(row=1, column=2, value="Mô tả")
            ws.merge_cells('B1:B2')
            ws.cell(row=1, column=3, value="Vật liệu")
            ws.merge_cells('C1:D1')
            ws.cell(row=1, column=5, value="Nhân công")
            ws.merge_cells('E1:F1')

            # Second row
            ws.cell(row=2, column=3, value="ĐG")
            ws.cell(row=2, column=4, value="TT")
            ws.cell(row=2, column=5, value="ĐG")
            ws.cell(row=2, column=6, value="TT")

            # Data row
            ws.cell(row=3, column=1, value=1)
            ws.cell(row=3, column=2, value="Bê tông")
            ws.cell(row=3, column=3, value=500000)
            ws.cell(row=3, column=4, value=5000000)
            ws.cell(row=3, column=5, value=200000)
            ws.cell(row=3, column=6, value=2000000)

            wb.save(f.name)
            yield f.name

        os.unlink(f.name)

    # Test 1: Header at row 0 (standard case)
    def test_standard_boq_discovery(self, discovery_service, standard_boq_excel):
        """Test discovery with standard BOQ file."""
        result = discovery_service.discover(standard_boq_excel)

        assert result.sheet_name == "BOQ"  # Should select BOQ, not Summary
        assert result.header_row == 0
        assert result.data_start_row == 1
        assert result.confidence_score > 40.0
        assert len(result.column_names) == 6

    # Test 2: Header at row 5+ (after project info)
    def test_header_after_project_info(self, discovery_service, header_at_row_5_excel):
        """Test detection when header is after project info rows."""
        result = discovery_service.discover(header_at_row_5_excel)

        assert result.header_row == 5  # 0-indexed, Excel row 6
        assert result.data_start_row >= 6

    # Test 3: Abbreviated Vietnamese headers
    def test_abbreviated_vietnamese_headers(self, discovery_service, standard_boq_excel):
        """Test that abbreviated headers are detected."""
        result = discovery_service.discover(standard_boq_excel)

        # Column type hints should be populated
        assert len(result.column_type_hints) > 0

    # Test 4: English headers
    def test_english_headers(self, discovery_service, english_headers_excel):
        """Test discovery with English headers."""
        result = discovery_service.discover(english_headers_excel)

        assert result.header_row == 0
        assert result.confidence_score > 40.0
        assert "Description" in result.column_names

    # Test 5: Merged multi-level headers
    def test_merged_headers(self, discovery_service, merged_header_excel):
        """Test discovery with merged headers."""
        result = discovery_service.discover(merged_header_excel)

        assert result.is_merged_header is True
        # Header could be at row 0 or 1 depending on detection
        assert result.header_row in [0, 1]
        # Data should start after all header rows
        assert result.data_start_row >= 2

    # Test 6: Sheet filtering
    def test_sheet_filtering(self, discovery_service, standard_boq_excel):
        """Test that Summary/Notes sheets are skipped."""
        result = discovery_service.discover(standard_boq_excel)

        # Should select BOQ, not Summary or Notes
        assert result.sheet_name == "BOQ"

    # Test 7: Specific sheet selection
    def test_specific_sheet_selection(self, discovery_service, standard_boq_excel):
        """Test specifying a specific sheet."""
        result = discovery_service.discover(standard_boq_excel, sheet_name="Summary")

        # Should use specified sheet even if it would be filtered
        assert result.sheet_name == "Summary"

    # Test 8: File not found
    def test_file_not_found(self, discovery_service):
        """Test error handling for non-existent file."""
        with pytest.raises(FileNotFoundError):
            discovery_service.discover("/nonexistent/file.xlsx")

    # Test 9: Invalid sheet name
    def test_invalid_sheet_name(self, discovery_service, standard_boq_excel):
        """Test error handling for invalid sheet name."""
        with pytest.raises(ValueError):
            discovery_service.discover(standard_boq_excel, sheet_name="NonExistent")

    # Test 10: Result serialization
    def test_result_to_dict(self, discovery_service, standard_boq_excel):
        """Test that result can be serialized to dict."""
        result = discovery_service.discover(standard_boq_excel)
        result_dict = result.to_dict()

        assert 'sheet_name' in result_dict
        assert 'header_row' in result_dict
        assert 'data_start_row' in result_dict
        assert 'column_names' in result_dict
        assert 'confidence_score' in result_dict

    # Test 11: Discover all sheets
    def test_discover_all_sheets(self, discovery_service, standard_boq_excel):
        """Test discovering headers in all valid sheets."""
        results = discovery_service.discover_all_sheets(standard_boq_excel)

        # Should only process BOQ sheet (Summary and Notes are filtered)
        assert len(results) == 1
        assert results[0].sheet_name == "BOQ"


class TestGetHeaderDiscoveryService:
    """Test singleton factory function."""

    def test_singleton(self):
        """Test that factory returns singleton."""
        service1 = get_header_discovery_service()
        service2 = get_header_discovery_service()
        assert service1 is service2
