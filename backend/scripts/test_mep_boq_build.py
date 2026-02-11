#!/usr/bin/env python3
"""
Test build: Process a single BOQ MEP file and export master items to Excel.

Usage:
    cd backend && python3 scripts/test_mep_boq_build.py
"""
import hashlib
import logging
import os
import sys
import time
from pathlib import Path

# Disable AI normalizer (rule-based is sufficient and faster)
os.environ["AI_NORMALIZATION_ENABLED"] = "false"
os.environ["AI_CONTEXT_ANALYSIS_ENABLED"] = "false"
os.environ["AI_DOMAIN_VALIDATION_ENABLED"] = "false"

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from app.core.database import SessionLocal, engine
from app.models.project import Project, ProjectType, ProjectStatus
from app.models.boq_file import BOQFile
from app.models.line_item import LineItem
from app.models.master_work_item import MasterWorkItem
from app.models.master_synonym import MasterSynonym
from app.services.file_service import FileService
from app.services.master_database_builder import (
    get_master_database_builder,
    BuildConfig,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("test_mep_build")

# Suppress noisy loggers
for noisy in [
    "sqlalchemy.engine",
    "app.services.normalization_orchestrator",
    "app.services.description_normalizer",
    "app.services.mep_equipment_normalizer",
    "app.utils.excel_processor",
    "app.services.ai_normalizer",
    "app.services.file_context_analyzer",
    "app.services.domain_validator",
]:
    logging.getLogger(noisy).setLevel(logging.WARNING)

# Configuration
BOQ_FILE = Path.home() / "Downloads" / "5. BOQ mời thầu MEP (final).xlsx"
PROJECT_CODE = "TEST-MEP-BUILD"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"


def create_project(db) -> Project:
    existing = db.query(Project).filter(Project.project_code == PROJECT_CODE).first()
    if existing:
        logger.info(f"Project exists: {existing.project_code} (ID: {existing.project_id})")
        return existing

    project = Project(
        project_code=PROJECT_CODE,
        project_name="Test MEP Build",
        project_type=ProjectType.infrastructure,
        location="Vietnam",
        status=ProjectStatus.active,
    )
    db.add(project)
    db.commit()
    db.refresh(project)
    logger.info(f"Created project: {project.project_code} (ID: {project.project_id})")
    return project


def upload_and_process(db, project_id: int, file_path: Path) -> int:
    filename = file_path.name
    with open(file_path, "rb") as f:
        file_hash = hashlib.sha256(f.read()).hexdigest()

    existing = db.query(BOQFile).filter(BOQFile.file_hash == file_hash).first()
    if existing:
        # Check if it actually has line items; if not, delete stale record
        li_count = db.query(LineItem).filter(LineItem.file_id == existing.file_id).count()
        if li_count > 0:
            logger.info(f"Already uploaded: {filename} (file_id={existing.file_id}, {li_count} items)")
            return existing.file_id
        else:
            logger.info(f"Stale upload found (0 items), re-processing: {filename}")
            db.query(BOQFile).filter(BOQFile.file_id == existing.file_id).delete()
            db.commit()

    service = FileService(db)
    with open(file_path, "rb") as f:
        saved_path = service.save_uploaded_file(f, filename, project_id)

    boq_file = BOQFile(
        project_id=project_id,
        file_name=filename,
        file_hash=file_hash,
        file_path=saved_path,
        total_rows=0,
        uploaded_by=1,
    )
    db.add(boq_file)
    db.commit()
    db.refresh(boq_file)

    # Try standard processing first
    result = service.process_file(
        file_id=boq_file.file_id,
        file_path=saved_path,
        column_mapping=None,
        user_id=1,
    )
    items_count = result.get("processed_items", 0)

    # If too few items extracted, try multi-sheet extraction
    if items_count < 50:
        logger.info(f"Only {items_count} items from auto-detect. Trying multi-sheet extraction...")
        items_count = _extract_all_sheets(db, boq_file.file_id, project_id, saved_path)

    logger.info(f"Processed {filename}: {items_count} line items total")
    return boq_file.file_id


def _extract_all_sheets(db, file_id: int, project_id: int, file_path: str) -> int:
    """
    Extract line items from all detail sheets of a MEP BOQ file.

    This BOQ has multi-sheet structure where each sheet is a subsystem:
    - Header at row 6-7: Stt | Nội dung công việc | ... | Đơn vị | Khối lượng | Đơn giá
    - Data starts from row 8 (0-indexed)
    - Description in column B (index 1)
    - Unit in column F (index 5)
    - Quantity in column G (index 6)
    - Unit price columns at index 7+ (material, labor, etc.)
    """
    import pandas as pd
    from app.services.normalization_orchestrator import get_normalization_orchestrator

    orchestrator = get_normalization_orchestrator()
    xl = pd.ExcelFile(file_path)

    # Skip summary and material list sheets
    skip_sheets = {'TH', '4. DMVL', '5. Prelim'}
    detail_sheets = [s for s in xl.sheet_names if s not in skip_sheets]

    total_added = 0
    row_counter = db.query(LineItem).filter(LineItem.file_id == file_id).count()

    for sheet_name in detail_sheets:
        try:
            df = pd.read_excel(xl, sheet_name=sheet_name, header=None)
            if df.shape[0] < 10 or df.shape[1] < 7:
                continue

            # Find header row (look for "Nội dung" or "Stt" in first 10 rows)
            header_row = None
            for i in range(min(10, df.shape[0])):
                row_text = ' '.join(str(v) for v in df.iloc[i] if pd.notna(v)).lower()
                if 'nội dung' in row_text or ('stt' in row_text and 'đơn vị' in row_text):
                    header_row = i
                    break

            if header_row is None:
                header_row = 6  # Default for this file format

            # Data starts after header (usually 2 rows: main header + sub-header)
            data_start = header_row + 2

            sheet_count = 0
            for idx in range(data_start, df.shape[0]):
                row = df.iloc[idx]

                # Get description (column B, index 1)
                desc = row.iloc[1] if pd.notna(row.iloc[1]) else None
                if not desc or not str(desc).strip():
                    continue

                desc = str(desc).strip()

                # Skip section headers (typically short, all-caps, or Roman numerals)
                if len(desc) < 3:
                    continue
                if desc.upper() == desc and len(desc) < 20 and not any(c.isdigit() for c in desc):
                    continue

                # Get unit (column F, index 5)
                unit = str(row.iloc[5]).strip() if df.shape[1] > 5 and pd.notna(row.iloc[5]) else None
                if unit in ('nan', 'None', ''):
                    unit = None
                # Truncate unit to fit database column (VARCHAR(10))
                if unit and len(unit) > 10:
                    unit = unit[:10]

                # Skip rows without unit (likely section headers)
                if not unit:
                    continue

                # Get quantity (column G, index 6)
                quantity = 0.0
                if df.shape[1] > 6 and pd.notna(row.iloc[6]):
                    try:
                        quantity = float(row.iloc[6])
                    except (ValueError, TypeError):
                        quantity = 0.0

                # Get unit price (column H, index 7 - material price)
                unit_price = 0.0
                if df.shape[1] > 7 and pd.notna(row.iloc[7]):
                    try:
                        unit_price = float(row.iloc[7])
                    except (ValueError, TypeError):
                        unit_price = 0.0

                # Normalize description
                try:
                    norm_result = orchestrator.normalize(desc)
                    normalized = norm_result.normalized
                    norm_confidence = norm_result.confidence
                    work_cat = norm_result.work_category.value if norm_result.work_category else 'general'
                except Exception:
                    normalized = desc.lower()
                    norm_confidence = 0.0
                    work_cat = 'general'

                row_counter += 1
                line_item = LineItem(
                    file_id=file_id,
                    project_id=project_id,
                    row_number=row_counter,
                    description=desc,
                    unit=unit,
                    quantity=quantity,
                    unit_price=unit_price,
                    amount=quantity * unit_price,
                    normalized_description=normalized,
                    normalization_confidence=norm_confidence,
                    work_category=work_cat,
                    classification_method='auto',
                    match_type='none',
                    needs_review=True,
                    original_sheet_name=sheet_name,
                )
                db.add(line_item)
                sheet_count += 1
                total_added += 1

            logger.info(f"  Sheet '{sheet_name}': {sheet_count} items")
            db.flush()

        except Exception as e:
            logger.warning(f"  Sheet '{sheet_name}': error - {e}")
            db.rollback()

    db.commit()
    logger.info(f"Multi-sheet extraction complete: {total_added} items from {len(detail_sheets)} sheets")
    return total_added + db.query(LineItem).filter(
        LineItem.file_id == file_id,
        LineItem.original_sheet_name.is_(None),
    ).count()


def export_master_review(db, output_path: Path):
    """Export all active master items to Excel for review."""
    items = db.query(MasterWorkItem).filter(
        MasterWorkItem.is_active == True
    ).order_by(MasterWorkItem.sec_code, MasterWorkItem.work_code).all()

    if not items:
        logger.warning("No master items to export")
        return

    wb = Workbook()

    # --- Styles ---
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    sec00_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    sec04_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # =====================
    # Sheet 1: All Items
    # =====================
    ws = wb.active
    ws.title = "All Master Items"

    headers = [
        "Work Code", "v4 Code", "Description", "Desc (Normalized)",
        "SEC Code", "Unit", "Occurrences",
        "Spec Category", "Spec Material", "Spec Grade", "Spec Dimension",
        "Matching Key", "Spec Status", "Spec Completeness",
        "Verified", "Synonyms Count",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for row_idx, item in enumerate(items, 2):
        syn_count = db.query(MasterSynonym).filter(
            MasterSynonym.master_id == item.master_id,
            MasterSynonym.is_active == True,
        ).count()

        data = [
            item.work_code,
            item.sec_code_v4 or "",
            item.description,
            item.description_normalized,
            item.sec_code,
            item.unit_standard,
            item.occurrence_count,
            item.spec_category or "",
            item.spec_material or "",
            item.spec_grade or "",
            item.spec_dimension or "",
            item.matching_key or "",
            item.spec_status if hasattr(item, 'spec_status') and item.spec_status else "",
            round(item.spec_completeness, 2) if hasattr(item, 'spec_completeness') and item.spec_completeness else 0,
            "Yes" if item.is_verified else "No",
            syn_count,
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if col in (1, 2, 5, 6, 13, 15):
                cell.alignment = center
            elif col in (7, 14, 16):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = left

        # Highlight SEC-00 rows
        if item.sec_code == "SEC-00":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = sec00_fill
        elif item.sec_code and item.sec_code.startswith("SEC-04"):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = sec04_fill

    # Column widths
    widths = [22, 22, 55, 55, 12, 8, 10, 15, 15, 12, 18, 30, 10, 12, 8, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(items) + 1}"

    # =====================
    # Sheet 2: SEC Summary
    # =====================
    ws2 = wb.create_sheet("SEC Summary")
    sec_counter = Counter(item.sec_code for item in items)

    ws2.cell(row=1, column=1, value="SEC Code").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Count").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=3, value="Percentage").font = header_font
    ws2.cell(row=1, column=3).fill = header_fill

    total = len(items)
    for row_idx, (sec, count) in enumerate(sorted(sec_counter.items()), 2):
        ws2.cell(row=row_idx, column=1, value=sec).border = border
        ws2.cell(row=row_idx, column=2, value=count).border = border
        pct = f"{count / total * 100:.1f}%" if total > 0 else "0%"
        ws2.cell(row=row_idx, column=3, value=pct).border = border

    ws2.cell(row=row_idx + 1, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=row_idx + 1, column=2, value=total).font = Font(bold=True)
    ws2.cell(row=row_idx + 1, column=3, value="100%").font = Font(bold=True)

    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 12

    # =====================
    # Sheet 3: Quality Checks
    # =====================
    ws3 = wb.create_sheet("Quality Checks")

    # Check 1: Degenerate descriptions
    degenerate_items = []
    for item in items:
        desc = item.description or ""
        words = desc.lower().split()
        has_repeat = False
        for i in range(len(words) - 1):
            if words[i] == words[i + 1] and len(words[i]) >= 2:
                has_repeat = True
                break
        if has_repeat or len(desc.strip()) < 5:
            degenerate_items.append(item)

    ws3.cell(row=1, column=1, value="Quality Check").font = Font(bold=True, size=13)
    ws3.cell(row=3, column=1, value="1. Degenerate Descriptions (repeated words, too short)").font = Font(bold=True)
    ws3.cell(row=3, column=1).fill = PatternFill(start_color="F8D7DA", fill_type="solid")
    ws3.cell(row=4, column=1, value=f"Found: {len(degenerate_items)} items").font = Font(bold=True)

    if degenerate_items:
        for col, h in enumerate(["Work Code", "Description", "SEC Code"], 1):
            ws3.cell(row=5, column=col, value=h).font = Font(bold=True)
        for i, item in enumerate(degenerate_items, 6):
            ws3.cell(row=i, column=1, value=item.work_code)
            ws3.cell(row=i, column=2, value=item.description)
            ws3.cell(row=i, column=3, value=item.sec_code)

    # Check 2: SEC-00 items
    sec00_items = [item for item in items if item.sec_code == "SEC-00"]
    start_row = 6 + len(degenerate_items) + 2
    ws3.cell(row=start_row, column=1, value="2. SEC-00 (Unclassified) Items").font = Font(bold=True)
    ws3.cell(row=start_row, column=1).fill = PatternFill(start_color="FFF3CD", fill_type="solid")
    pct_sec00 = f"{len(sec00_items) / total * 100:.1f}%" if total > 0 else "0%"
    ws3.cell(row=start_row + 1, column=1, value=f"Found: {len(sec00_items)} items ({pct_sec00})").font = Font(bold=True)

    if sec00_items:
        for col, h in enumerate(["Work Code", "Description", "Unit"], 1):
            ws3.cell(row=start_row + 2, column=col, value=h).font = Font(bold=True)
        for i, item in enumerate(sec00_items, start_row + 3):
            ws3.cell(row=i, column=1, value=item.work_code)
            ws3.cell(row=i, column=2, value=item.description)
            ws3.cell(row=i, column=3, value=item.unit_standard)

    # Check 3: Work code mismatches (pipe fittings in EARTH, etc.)
    mismatch_items = []
    mep_keywords = ['van ', 'côn ', 'cút ', 'bích ', 'ống ', 'contactor', 'mccb', 'mcb',
                     'aptomat', 'tủ điện', 'sprinkler', 'pccc']
    non_mep_cats = ['EARTH', 'CONC', 'PRELIM', 'WALL', 'BRICK']
    for item in items:
        desc_lower = (item.description or "").lower()
        code_upper = (item.work_code or "").upper()
        for kw in mep_keywords:
            if kw in desc_lower:
                for cat in non_mep_cats:
                    if cat in code_upper:
                        mismatch_items.append((item, kw, cat))
                break

    start_row2 = start_row + 3 + len(sec00_items) + 2
    ws3.cell(row=start_row2, column=1, value="3. Work Code Category Mismatches").font = Font(bold=True)
    ws3.cell(row=start_row2, column=1).fill = PatternFill(start_color="D4EDDA", fill_type="solid")
    ws3.cell(row=start_row2 + 1, column=1, value=f"Found: {len(mismatch_items)} items").font = Font(bold=True)

    if mismatch_items:
        for col, h in enumerate(["Work Code", "Description", "Matched Keyword", "Wrong Category"], 1):
            ws3.cell(row=start_row2 + 2, column=col, value=h).font = Font(bold=True)
        for i, (item, kw, cat) in enumerate(mismatch_items, start_row2 + 3):
            ws3.cell(row=i, column=1, value=item.work_code)
            ws3.cell(row=i, column=2, value=item.description)
            ws3.cell(row=i, column=3, value=kw)
            ws3.cell(row=i, column=4, value=cat)

    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 60
    ws3.column_dimensions["C"].width = 20
    ws3.column_dimensions["D"].width = 20

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info(f"Exported {len(items)} items to {output_path}")
    print(f"\n  Output: {output_path}")
    print(f"  Total items:  {total}")
    print(f"  SEC-00 count: {len(sec00_items)} ({pct_sec00})")
    print(f"  Degenerate:   {len(degenerate_items)}")
    print(f"  Mismatches:   {len(mismatch_items)}")


def main():
    if not BOQ_FILE.exists():
        logger.error(f"BOQ file not found: {BOQ_FILE}")
        sys.exit(1)

    logger.info(f"Processing: {BOQ_FILE.name}")

    db = SessionLocal()
    try:
        # Create project
        project = create_project(db)

        # Upload and process
        t0 = time.time()
        file_id = upload_and_process(db, project.project_id, BOQ_FILE)
        if not file_id:
            logger.error("Failed to process BOQ file")
            sys.exit(1)

        total_li = db.query(LineItem).filter(LineItem.file_id == file_id).count()
        logger.info(f"Total line items: {total_li}")

        # Build master database (clear and rebuild)
        logger.info("\n" + "=" * 70)
        logger.info("  Building Master Database...")
        logger.info("=" * 70)

        t1 = time.time()
        builder = get_master_database_builder(db)
        config = BuildConfig(
            pareto_threshold=0.80,
            clustering_threshold=0.85,
            min_frequency=1,
            auto_approve=True,
            clear_existing=True,
            include_only_pareto=False,
        )

        def progress(step, current, total):
            logger.info(f"  Step: {step} ({current}/{total})")

        result = builder.build(file_ids=[file_id], config=config, progress_callback=progress)
        t_build = time.time() - t1

        # Print report
        print("\n" + "=" * 70)
        print("  MASTER DATABASE BUILD REPORT")
        print("=" * 70)
        print(f"  File:                 {BOQ_FILE.name}")
        print(f"  Line items:           {total_li}")
        print(f"  Build time:           {t_build:.1f}s")
        print()
        print(f"  Step 1 - Aggregation:")
        print(f"    Unique descriptions: {result.step1_stats.output_count}")
        print()
        print(f"  Step 2 - Standardization:")
        print(f"    After clustering:    {result.step2_stats.output_count}")
        d2 = result.step2_stats.details
        print(f"    Clusters formed:     {d2.get('clusters_formed', 'N/A')}")
        print(f"    Pareto top:          {d2.get('pareto_top_count', 'N/A')}")
        print(f"    Total synonyms:      {d2.get('total_synonyms', 'N/A')}")
        print()
        print(f"  Step 3 - Coding & Tagging:")
        d3 = result.step3_stats.details
        print(f"    Approved (master):   {d3.get('approved', 0)}")
        print(f"    Pending review:      {d3.get('pending', 0)}")
        print(f"    Rejected:            {d3.get('rejected', 0)}")
        print(f"    Updated existing:    {d3.get('updated', 0)}")
        by_sec = d3.get("by_sec_code", {})
        if by_sec:
            print(f"    By SEC code:")
            for sec, count in sorted(by_sec.items()):
                print(f"      {sec}: {count}")
        print()
        print(f"  Totals:")
        print(f"    Master items added:  {result.total_master_added}")
        print(f"    Synonyms added:      {result.total_synonyms_added}")
        print("=" * 70)

        # Export to Excel
        output_file = OUTPUT_DIR / "master_items_review.xlsx"
        export_master_review(db, output_file)

        total_elapsed = time.time() - t0
        print(f"\n  Total elapsed: {total_elapsed:.1f}s")

    except KeyboardInterrupt:
        logger.info("\nInterrupted by user")
        db.rollback()
    except Exception as e:
        logger.error(f"\nFatal error: {e}")
        import traceback
        traceback.print_exc()
        db.rollback()
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    main()
