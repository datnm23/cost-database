#!/usr/bin/env python3
"""
Test BOQ Export Service with a real file.

Usage:
    cd backend && python scripts/test_boq_export.py "/path/to/file.xlsx"
    cd backend && python scripts/test_boq_export.py  # Uses default test file
"""
import argparse
import sys
import os
from datetime import datetime

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from sqlalchemy.orm import Session
from app.core.database import SessionLocal
from app.models.boq_file import BOQFile, FileStatus
from app.models.project import Project
from app.models.line_item import LineItem
from app.utils.excel_processor import ExcelProcessor
from app.services.boq_export_service import (
    BOQExportService,
    standardize_unit,
    get_default_unit_for_sec_code,
    get_unit_with_default,
    format_sec_code_display,
    inject_default_specs,
    format_defaults_explanation,
)


def test_unit_functions():
    """Test unit standardization functions."""
    print("\n" + "=" * 60)
    print("TESTING UNIT FUNCTIONS")
    print("=" * 60)

    # Test standardize_unit
    test_cases = [
        ('m3', 'm³'),
        ('m2', 'm²'),
        ('sqm', 'm²'),
        ('pcs', 'cái'),
        ('lump sum', 'trọn gói'),
        ('kg', 'kg'),
        ('tấn', 'tấn'),
        ('ton', 'tấn'),
        ('bộ', 'bộ'),
        ('set', 'bộ'),
    ]

    print("\n1. standardize_unit() tests:")
    all_passed = True
    for raw, expected in test_cases:
        result = standardize_unit(raw)
        status = "✓" if result == expected else "✗"
        if result != expected:
            all_passed = False
        print(f"   {status} standardize_unit('{raw}') = '{result}' (expected: '{expected}')")

    # Test get_default_unit_for_sec_code
    print("\n2. get_default_unit_for_sec_code() tests:")
    sec_tests = [
        ('SEC-02-01', 'm³'),  # Concrete
        ('SEC-03-02', 'm²'),  # Plastering
        ('SEC-04-03', 'bộ'),  # HVAC
        ('SEC-01-02', 'm'),   # Piling
        ('SEC-05-03', 'cây'), # Greenery
    ]

    for sec_code, expected in sec_tests:
        result = get_default_unit_for_sec_code(sec_code)
        status = "✓" if result == expected else "✗"
        if result != expected:
            all_passed = False
        print(f"   {status} get_default_unit_for_sec_code('{sec_code}') = '{result}' (expected: '{expected}')")

    # Test get_unit_with_default
    print("\n3. get_unit_with_default() tests:")

    # Has unit - standardize only
    unit, is_default = get_unit_with_default('m2', 'SEC-03-02')
    status = "✓" if unit == 'm²' and is_default == False else "✗"
    print(f"   {status} get_unit_with_default('m2', 'SEC-03-02') = ('{unit}', {is_default})")

    # No unit - use default
    unit, is_default = get_unit_with_default('', 'SEC-02-01')
    status = "✓" if unit == 'm³' and is_default == True else "✗"
    print(f"   {status} get_unit_with_default('', 'SEC-02-01') = ('{unit}', {is_default})")

    # Test format_sec_code_display
    print("\n4. format_sec_code_display() tests:")
    display_tests = [
        ('SEC-01-01', '01.01.EARTH'),
        ('SEC-02-01', '02.01.CONC'),
        ('SEC-03-03', '03.03.PAINT'),
        ('SEC-04-02', '04.02.PLUMB'),
    ]

    for sec_code, expected in display_tests:
        result = format_sec_code_display(sec_code)
        status = "✓" if result == expected else "✗"
        if result != expected:
            all_passed = False
        print(f"   {status} format_sec_code_display('{sec_code}') = '{result}' (expected: '{expected}')")

    # Test inject_default_specs
    print("\n5. inject_default_specs() tests:")

    # Concrete without grade
    desc, defaults = inject_default_specs("Bê tông móng", "bê tông")
    print(f"   inject_default_specs('Bê tông móng', 'bê tông'):")
    print(f"     Result: '{desc}'")
    print(f"     Defaults: {defaults}")

    # Earthworks without K grade
    desc, defaults = inject_default_specs("Đào đất hố móng", "đất")
    print(f"   inject_default_specs('Đào đất hố móng', 'đất'):")
    print(f"     Result: '{desc}'")
    print(f"     Defaults: {defaults}")

    # Test format_defaults_explanation
    print("\n6. format_defaults_explanation() tests:")
    details = ["Mác mặc định: M250", "Loại đá: Đá 1x2", "Đơn vị mặc định: m³"]
    result = format_defaults_explanation(details)
    print(f"   Input: {details}")
    print(f"   Output:\n{result}")

    print("\n" + "=" * 60)
    if all_passed:
        print("All unit function tests PASSED ✓")
    else:
        print("Some tests FAILED ✗")
    print("=" * 60)

    return all_passed


def test_export_with_file(file_path: str, db: Session):
    """Test export with a real BOQ file."""
    print("\n" + "=" * 60)
    print("TESTING BOQ EXPORT WITH REAL FILE")
    print("=" * 60)
    print(f"File: {file_path}")

    # Check if file exists
    if not os.path.exists(file_path):
        print(f"ERROR: File not found: {file_path}")
        return False

    # Create or get project
    project = db.query(Project).filter(Project.project_code == "TEST-EXPORT").first()
    if not project:
        project = Project(
            project_code="TEST-EXPORT",
            project_name="Test Export Project",
            project_type="infrastructure"  # Must be one of: residential, commercial, industrial, infrastructure
        )
        db.add(project)
        db.commit()
        db.refresh(project)
        print(f"Created test project: {project.project_name}")

    # Check if file already imported
    existing_file = db.query(BOQFile).filter(
        BOQFile.file_name == os.path.basename(file_path),
        BOQFile.project_id == project.project_id
    ).first()

    if existing_file:
        print(f"Using existing BOQ file record (ID: {existing_file.file_id})")
        boq_file = existing_file
    else:
        # Import file
        print("Importing BOQ file...")
        boq_file = BOQFile(
            project_id=project.project_id,
            file_name=os.path.basename(file_path),
            file_path=file_path,
            status=FileStatus.draft
        )
        db.add(boq_file)
        db.commit()
        db.refresh(boq_file)
        print(f"Created BOQ file record (ID: {boq_file.file_id})")

        # Process Excel file
        print("Processing Excel file...")
        processor = ExcelProcessor()
        rows = processor.process_file(file_path)
        print(f"Extracted {len(rows)} rows from Excel")

        # Create line items
        for i, row in enumerate(rows[:100]):  # Limit to 100 for testing
            line_item = LineItem(
                file_id=boq_file.file_id,
                row_number=i + 1,
                description=row.get('description', ''),
                unit=row.get('unit', ''),
                quantity=row.get('quantity'),
                unit_price=row.get('unit_price'),
                amount=row.get('amount'),
                original_sheet_name=row.get('sheet_name', 'Sheet1')
            )
            db.add(line_item)

        db.commit()
        print(f"Created {min(100, len(rows))} line items")

    # Count line items
    line_count = db.query(LineItem).filter(LineItem.file_id == boq_file.file_id).count()
    print(f"Total line items for file: {line_count}")

    # Export
    output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "test_output")
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"test_export_{timestamp}.xlsx")

    print(f"\nExporting to: {output_path}")

    try:
        export_service = BOQExportService(db)
        result_path = export_service.export_processing_result(
            file_id=boq_file.file_id,
            output_path=output_path,
            include_master_match=True
        )
        print(f"\n✓ Export successful!")
        print(f"Output file: {result_path}")
        print(f"File size: {os.path.getsize(result_path):,} bytes")

        # Verify the output
        from openpyxl import load_workbook
        wb = load_workbook(result_path)
        print(f"\nSheets in output file:")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  - {sheet_name}: {ws.max_row} rows x {ws.max_column} columns")

        # Check column headers in "All Items" sheet
        if "All Items" in wb.sheetnames:
            ws = wb["All Items"]
            headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
            print(f"\n'All Items' sheet headers:")
            for i, header in enumerate(headers, 1):
                print(f"  {i}. {header}")

            # Verify Đơn vị column exists at position 9
            if len(headers) >= 9 and headers[8] == "Đơn vị":
                print("\n✓ 'Đơn vị' column correctly at position 9")
            else:
                print(f"\n✗ 'Đơn vị' column issue - found at position 9: '{headers[8] if len(headers) >= 9 else 'N/A'}'")

        return True

    except Exception as e:
        print(f"\n✗ Export failed: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    parser = argparse.ArgumentParser(description="Test BOQ Export Service")
    parser.add_argument(
        "file_path",
        nargs="?",
        default=os.path.expanduser("~/Downloads/BOQ moi thau Gói Thi công đường và hạ tầng kỹ thuật phân khu 1 - DA Hoang Long.xlsx"),
        help="Path to BOQ Excel file"
    )
    parser.add_argument(
        "--skip-unit-tests",
        action="store_true",
        help="Skip unit function tests"
    )

    args = parser.parse_args()

    print("=" * 60)
    print("BOQ EXPORT SERVICE TEST")
    print("=" * 60)

    # Run unit tests
    if not args.skip_unit_tests:
        test_unit_functions()

    # Test with real file
    db = SessionLocal()
    try:
        test_export_with_file(args.file_path, db)
    finally:
        db.close()

    print("\n" + "=" * 60)
    print("TEST COMPLETE")
    print("=" * 60)


if __name__ == "__main__":
    main()
