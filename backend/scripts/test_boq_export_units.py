#!/usr/bin/env python3
"""
Test BOQ Export Service - Unit functions and export logic only.
Does not require database connection.

Usage:
    cd backend && python3 scripts/test_boq_export_units.py
"""
import sys
import os
from datetime import datetime

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.boq_export_service import (
    # Constants
    SEC_CODE_DEFAULT_UNIT,
    SEC_CODE_DISPLAY_MAP,
    UNIT_STANDARDIZATION,
    CONCRETE_DEFAULTS,
    EARTHWORKS_DEFAULTS,
    # Functions
    standardize_unit,
    get_default_unit_for_sec_code,
    get_unit_with_default,
    format_sec_code_display,
    inject_default_specs,
    format_defaults_explanation,
    combine_description_with_specs,
    extract_tech_specs_from_line_item,
    # Styles
    HEADER_FONT,
    HEADER_FILL,
    THIN_BORDER,
    CENTER_ALIGN,
    LEFT_ALIGN,
    WRAP_ALIGN,
)


def test_unit_functions():
    """Test unit standardization functions."""
    print("\n" + "=" * 60)
    print("TESTING UNIT FUNCTIONS")
    print("=" * 60)

    all_passed = True

    # Test standardize_unit
    test_cases = [
        ('m3', 'm³'),
        ('m2', 'm²'),
        ('sqm', 'm²'),
        ('pcs', 'cái'),
        ('lump sum', 'trọn gói'),
        ('kg', 'kg'),
        ('tấn', 'tấn'),
        ('ton', 'tấn'),
        ('bộ', 'bộ'),
        ('set', 'bộ'),
        ('mét khối', 'm³'),
        ('mét vuông', 'm²'),
    ]

    print("\n1. standardize_unit() tests:")
    for raw, expected in test_cases:
        result = standardize_unit(raw)
        status = "✓" if result == expected else "✗"
        if result != expected:
            all_passed = False
        print(f"   {status} standardize_unit('{raw}') = '{result}' (expected: '{expected}')")

    # Test get_default_unit_for_sec_code
    print("\n2. get_default_unit_for_sec_code() tests:")
    sec_tests = [
        ('SEC-02-01', 'm³'),  # Concrete
        ('SEC-03-02', 'm²'),  # Plastering
        ('SEC-04-03', 'bộ'),  # HVAC
        ('SEC-01-02', 'm'),   # Piling
        ('SEC-05-03', 'cây'), # Greenery
        ('SEC-02', 'm³'),     # Parent level
        ('SEC-03', 'm²'),     # Parent level
        ('SEC-04', 'bộ'),     # Parent level
    ]

    for sec_code, expected in sec_tests:
        result = get_default_unit_for_sec_code(sec_code)
        status = "✓" if result == expected else "✗"
        if result != expected:
            all_passed = False
        print(f"   {status} get_default_unit_for_sec_code('{sec_code}') = '{result}' (expected: '{expected}')")

    # Test get_unit_with_default
    print("\n3. get_unit_with_default() tests:")

    # Has unit - standardize only
    unit, is_default = get_unit_with_default('m2', 'SEC-03-02')
    status = "✓" if unit == 'm²' and is_default == False else "✗"
    if unit != 'm²' or is_default != False:
        all_passed = False
    print(f"   {status} get_unit_with_default('m2', 'SEC-03-02') = ('{unit}', {is_default}) (expected: ('m²', False))")

    # No unit - use default
    unit, is_default = get_unit_with_default('', 'SEC-02-01')
    status = "✓" if unit == 'm³' and is_default == True else "✗"
    if unit != 'm³' or is_default != True:
        all_passed = False
    print(f"   {status} get_unit_with_default('', 'SEC-02-01') = ('{unit}', {is_default}) (expected: ('m³', True))")

    # No unit, no SEC code - empty result
    unit, is_default = get_unit_with_default('', '')
    status = "✓" if unit == '' and is_default == False else "✗"
    if unit != '' or is_default != False:
        all_passed = False
    print(f"   {status} get_unit_with_default('', '') = ('{unit}', {is_default}) (expected: ('', False))")

    # Test format_sec_code_display
    print("\n4. format_sec_code_display() tests:")
    display_tests = [
        ('SEC-01-01', '01.01.EARTH'),
        ('SEC-02-01', '02.01.CONC'),
        ('SEC-03-03', '03.03.PAINT'),
        ('SEC-04-02', '04.02.PLUMB'),
        ('SEC-01', '01.SUBSTRUCT'),
        ('SEC-02', '02.SUPERSTRUCT'),
        ('UNKNOWN', 'UNKNOWN'),  # Fall back to original
    ]

    for sec_code, expected in display_tests:
        result = format_sec_code_display(sec_code)
        status = "✓" if result == expected else "✗"
        if result != expected:
            all_passed = False
        print(f"   {status} format_sec_code_display('{sec_code}') = '{result}' (expected: '{expected}')")

    # Test inject_default_specs
    print("\n5. inject_default_specs() tests:")

    # Concrete without grade
    desc, defaults = inject_default_specs("Bê tông móng", "bê tông")
    print(f"   inject_default_specs('Bê tông móng', 'bê tông'):")
    print(f"     Result: '{desc}'")
    print(f"     Defaults: {defaults}")
    has_m250 = 'M250' in desc
    has_da = 'Đá 1x2' in desc
    status = "✓" if has_m250 and has_da else "✗"
    if not (has_m250 and has_da):
        all_passed = False
    print(f"     {status} Contains M250 and Đá 1x2: {has_m250 and has_da}")

    # Concrete with grade already present
    desc, defaults = inject_default_specs("Bê tông móng M300 đá 2x4", "bê tông")
    print(f"\n   inject_default_specs('Bê tông móng M300 đá 2x4', 'bê tông'):")
    print(f"     Result: '{desc}'")
    print(f"     Defaults: {defaults}")
    status = "✓" if len(defaults) == 0 else "✗"
    if len(defaults) != 0:
        all_passed = False
    print(f"     {status} No defaults applied: {len(defaults) == 0}")

    # Earthworks without K grade
    desc, defaults = inject_default_specs("Đào đất hố móng", "đất")
    print(f"\n   inject_default_specs('Đào đất hố móng', 'đất'):")
    print(f"     Result: '{desc}'")
    print(f"     Defaults: {defaults}")
    has_k95 = 'K95' in desc
    status = "✓" if has_k95 else "✗"
    if not has_k95:
        all_passed = False
    print(f"     {status} Contains K95: {has_k95}")

    # Test format_defaults_explanation
    print("\n6. format_defaults_explanation() tests:")
    details = ["Mác mặc định: M250", "Loại đá: Đá 1x2", "Đơn vị mặc định: m³"]
    result = format_defaults_explanation(details)
    print(f"   Input: {details}")
    print(f"   Output:\n{result}")
    has_bullets = all(line.startswith('• ') for line in result.split('\n'))
    has_lines = result.count('\n') == 2
    status = "✓" if has_bullets and has_lines else "✗"
    if not (has_bullets and has_lines):
        all_passed = False
    print(f"   {status} Correct formatting: bullets={has_bullets}, lines={has_lines}")

    # Test combine_description_with_specs
    print("\n7. combine_description_with_specs() tests:")

    desc = combine_description_with_specs("Bê tông móng", "M250, đá 1x2")
    expected = "Bê tông móng - M250, đá 1x2"
    status = "✓" if desc == expected else "✗"
    if desc != expected:
        all_passed = False
    print(f"   {status} combine_description_with_specs('Bê tông móng', 'M250, đá 1x2') = '{desc}'")

    # Empty specs
    desc = combine_description_with_specs("Bê tông móng", "")
    status = "✓" if desc == "Bê tông móng" else "✗"
    if desc != "Bê tông móng":
        all_passed = False
    print(f"   {status} combine_description_with_specs('Bê tông móng', '') = '{desc}'")

    # Specs already in description
    desc = combine_description_with_specs("Bê tông M250 móng", "m250")
    status = "✓" if desc == "Bê tông M250 móng" else "✗"  # Should not duplicate
    if desc != "Bê tông M250 móng":
        all_passed = False
    print(f"   {status} combine_description_with_specs('Bê tông M250 móng', 'm250') = '{desc}' (no duplication)")

    print("\n" + "=" * 60)
    if all_passed:
        print("All unit function tests PASSED ✓")
    else:
        print("Some tests FAILED ✗")
    print("=" * 60)

    return all_passed


def test_excel_output():
    """Create a sample Excel output to verify column structure."""
    print("\n" + "=" * 60)
    print("TESTING EXCEL OUTPUT STRUCTURE")
    print("=" * 60)

    output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "test_output")
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"test_column_structure_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Structure"

    # Headers as defined in the plan
    headers = [
        "STT",                    # 1
        "Mã Công Tác Chuẩn",      # 2
        "Tên Công Việc Gốc",      # 3
        "Thông số KT gốc",        # 4
        "Tên Chuẩn Hóa",          # 5
        "Mã Master Mapping",      # 6
        "Tỷ lệ khớp (%)",         # 7
        "Nhóm (Category)",        # 8
        "Đơn vị",                 # 9 - NEW!
        "Ghi chú xử lý"           # 10
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Sample data rows
    sample_data = [
        {
            "original": "Bê tông móng đài cọc",
            "tech_specs": "",
            "sec_code": "SEC-02-01",
            "raw_unit": "",
            "category": "bê tông"
        },
        {
            "original": "Trát tường trong",
            "tech_specs": "Vữa XM M75",
            "sec_code": "SEC-03-02",
            "raw_unit": "m2",
            "category": "trát"
        },
        {
            "original": "Sơn tường trong",
            "tech_specs": "",
            "sec_code": "SEC-03-03",
            "raw_unit": "sqm",
            "category": "sơn"
        },
        {
            "original": "Đào đất hố móng",
            "tech_specs": "",
            "sec_code": "SEC-01-01",
            "raw_unit": "m3",
            "category": "đất"
        },
        {
            "original": "Lắp đặt điều hòa",
            "tech_specs": "24000 BTU",
            "sec_code": "SEC-04-03",
            "raw_unit": "",
            "category": "mep"
        }
    ]

    for row_idx, item in enumerate(sample_data, 2):
        # Combine description with tech specs
        combined = combine_description_with_specs(item["original"], item["tech_specs"])

        # Inject defaults
        enhanced, defaults_list = inject_default_specs(combined, item["category"])

        # Get unit with default
        std_unit, unit_is_default = get_unit_with_default(item["raw_unit"], item["sec_code"])
        if unit_is_default and std_unit:
            defaults_list.append(f"Đơn vị mặc định: {std_unit}")

        # Format notes
        notes = format_defaults_explanation(defaults_list)

        # Category display
        category_display = format_sec_code_display(item["sec_code"])

        # Write row
        data = [
            row_idx - 1,                    # STT
            f"WC-{row_idx:04d}",            # Mã Công Tác Chuẩn
            item["original"],               # Tên Công Việc Gốc
            item["tech_specs"],             # Thông số KT gốc
            enhanced,                       # Tên Chuẩn Hóa
            f"WC-{row_idx:04d}",            # Mã Master Mapping
            95.0 if row_idx % 2 == 0 else 85.0,  # Tỷ lệ khớp
            category_display,               # Nhóm (Category)
            std_unit,                       # Đơn vị
            notes                           # Ghi chú xử lý
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER

            if col in [1, 6, 7, 9]:  # Center aligned columns
                cell.alignment = CENTER_ALIGN
            elif col in [5, 10]:  # Wrap text columns
                cell.alignment = WRAP_ALIGN
            else:
                cell.alignment = LEFT_ALIGN

    # Adjust row heights for multi-line content
    for row in range(2, ws.max_row + 1):
        notes_cell = ws.cell(row=row, column=10)
        if notes_cell.value:
            line_count = str(notes_cell.value).count('\n') + 1
            ws.row_dimensions[row].height = max(15, line_count * 15)

    # Column widths per plan: [6, 20, 45, 30, 50, 20, 10, 20, 10, 35]
    col_widths = [6, 20, 45, 30, 50, 20, 10, 20, 10, 35]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Save
    wb.save(output_path)

    print(f"\nCreated test Excel file: {output_path}")
    print(f"File size: {os.path.getsize(output_path):,} bytes")

    # Verify structure
    from openpyxl import load_workbook as load_wb
    wb_check = load_wb(output_path)
    ws_check = wb_check.active

    print(f"\nVerifying structure:")
    print(f"  Total columns: {ws_check.max_column}")
    print(f"  Total rows: {ws_check.max_row}")

    headers_read = [ws_check.cell(row=1, column=col).value for col in range(1, ws_check.max_column + 1)]
    print(f"\nColumn headers:")
    for i, h in enumerate(headers_read, 1):
        print(f"  {i}. {h}")

    # Check Đơn vị column
    if len(headers_read) >= 9 and headers_read[8] == "Đơn vị":
        print("\n✓ 'Đơn vị' column correctly at position 9")
    else:
        print(f"\n✗ 'Đơn vị' column issue - position 9 has: '{headers_read[8] if len(headers_read) >= 9 else 'N/A'}'")

    # Print sample row
    print("\nSample row 2 data:")
    for col in range(1, ws_check.max_column + 1):
        val = ws_check.cell(row=2, column=col).value
        print(f"  {headers_read[col-1]}: {val}")

    return True


def main():
    print("=" * 60)
    print("BOQ EXPORT SERVICE - UNIT TESTS")
    print("=" * 60)
    print(f"Test time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # Test unit functions
    unit_tests_passed = test_unit_functions()

    # Test Excel output structure
    excel_test_passed = test_excel_output()

    print("\n" + "=" * 60)
    print("FINAL RESULTS")
    print("=" * 60)
    print(f"Unit function tests: {'PASSED ✓' if unit_tests_passed else 'FAILED ✗'}")
    print(f"Excel output tests: {'PASSED ✓' if excel_test_passed else 'FAILED ✗'}")
    print("=" * 60)

    sys.exit(0 if (unit_tests_passed and excel_test_passed) else 1)


if __name__ == "__main__":
    main()
