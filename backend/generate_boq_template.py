"""
Generate TEMPLATE_BOQ_Result_After_Processing.xlsx

This creates a sample template showing the expected output format
after BOQ processing with all sheets and formatting.
"""
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FILL_GREEN = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
HEADER_FILL_ORANGE = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
HEADER_FILL_RED = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
HEADER_FILL_PURPLE = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

EXACT_MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FUZZY_MATCH_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
NEW_ITEM_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')


def create_summary_sheet(wb):
    """Create summary sheet with statistics"""
    ws = wb.create_sheet("1. Summary", 0)

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "BOQ PROCESSING RESULT"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = CENTER_ALIGN

    # File Info
    row = 3
    info_data = [
        ("File Name:", "BOQ_Sample_Project.xlsx"),
        ("Project:", "Sample Construction Project"),
        ("Project Code:", "PRJ-2024-001"),
        ("Processed At:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    for label, value in info_data:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Statistics
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "PROCESSING STATISTICS"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    row += 1

    headers = ["Metric", "Count", "Percentage/Note"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row += 1

    stats = [
        ("Total Extracted", 1000, "100%"),
        ("After Raw Dedupe", 750, "75.0%"),
        ("After Normalization", 500, "50.0%"),
        ("", "", ""),
        ("Exact Matches (≥95%)", 300, "60.0%"),
        ("Fuzzy Matches (80-95%)", 100, "20.0%"),
        ("New Items (<80%)", 100, "20.0%"),
        ("", "", ""),
        ("New Items (Deduped)", 80, "Ready to add to Master"),
    ]

    for metric, count, pct in stats:
        ws.cell(row=row, column=1, value=metric)
        if count != "":
            ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=pct)
        row += 1

    # Match Type Legend
    row += 2
    ws[f'A{row}'] = "MATCH TYPE LEGEND"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    legends = [
        ("Exact Match", "≥95% similarity - Auto assign work code", EXACT_MATCH_FILL),
        ("Fuzzy Match", "80-95% similarity - Needs human review", FUZZY_MATCH_FILL),
        ("New Item", "<80% similarity - New work item to add", NEW_ITEM_FILL),
    ]

    for name, desc, fill in legends:
        ws.cell(row=row, column=1, value=name).fill = fill
        ws.cell(row=row, column=2, value=desc)
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 30


def create_all_items_sheet(wb):
    """Create sheet with all processed items"""
    ws = wb.create_sheet("2. All Items")

    headers = [
        "No.",
        "Original Description",
        "Normalized Description",
        "Match Type",
        "Similarity %",
        "Master Work Code",
        "Needs Review",
        "Suggested Matches"
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Sample data
    sample_data = [
        (1, "Đổ bê tông móng M300 thương phẩm", "Bê tông móng - M300 - thương phẩm", "EXACT", 98.5, "S02-CONC-M300-0001", "No", ""),
        (2, "Đổ bê tông cột M350", "Bê tông cột - M350", "EXACT", 96.2, "S02-CONC-M350-0015", "No", ""),
        (3, "Xây tường gạch đặc dày 220", "Xây tường - gạch đặc - dày 220", "FUZZY", 88.5, "S03-WALL-BRICK-0008", "Yes", "S03-WALL-BRICK-0008 (88.5%)"),
        (4, "Trát tường trong nhà dày 15mm", "Trát tường trong - dày 15mm - M75", "FUZZY", 82.3, "S04-PLAS-INT-0001", "Yes", "S04-PLAS-INT-0001 (82.3%)"),
        (5, "Lắp đặt ống HDPE D110 PN16", "Lắp đặt ống HDPE - D110 - PN16", "EXACT", 99.1, "S07-PIPE-HDPE-0022", "No", ""),
        (6, "Cáp Cu/XLPE/PVC 4x300mm2", "Lắp đặt Cáp Cu/XLPE/PVC - 4x300mm2", "NEW", 45.2, "", "No", "S08-ELEC-CABL-0001 (45.2%)"),
        (7, "Trồng cây Bàng Đài Loan H3-4m", "Trồng Cây Bàng Đài Loan - H3-4m", "NEW", 30.0, "", "No", ""),
        (8, "Đắp đất đầm chặt K95 đất mua mới", "Đắp đất - đất mua mới - K95", "EXACT", 97.8, "S01-EARTH-FILL-0003", "No", ""),
    ]

    for row_idx, (no, orig, norm, match_type, sim, code, review, suggested) in enumerate(sample_data, 2):
        if match_type == "EXACT":
            fill = EXACT_MATCH_FILL
        elif match_type == "FUZZY":
            fill = FUZZY_MATCH_FILL
        else:
            fill = NEW_ITEM_FILL

        data = [no, orig, norm, match_type, sim, code, review, suggested]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = fill
            cell.border = THIN_BORDER
            if col in [1, 4, 5, 7]:
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN

    # Column widths
    col_widths = [6, 45, 45, 12, 12, 25, 12, 35]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'


def create_exact_matches_sheet(wb):
    """Create sheet for exact matches"""
    ws = wb.create_sheet("3. Exact Matches")

    headers = [
        "No.",
        "Original Description",
        "Normalized Description",
        "Similarity %",
        "Matched Work Code",
        "Matched Description",
        "SEC Code",
        "Action"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_GREEN
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    sample_data = [
        (1, "Đổ bê tông móng M300 thương phẩm", "Bê tông móng - M300 - thương phẩm", 98.5, "S02-CONC-M300-0001", "Bê tông móng - M300 - thương phẩm", "SEC-02", "APPROVE"),
        (2, "Đổ bê tông cột M350", "Bê tông cột - M350", 96.2, "S02-CONC-M350-0015", "Bê tông cột - M350", "SEC-02", "APPROVE"),
        (3, "Lắp đặt ống HDPE D110 PN16", "Lắp đặt ống HDPE - D110 - PN16", 99.1, "S07-PIPE-HDPE-0022", "Lắp đặt ống HDPE - D110 - PN16", "SEC-07", "APPROVE"),
    ]

    for row_idx, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            if col == 1 or col == 4 or col == 8:
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN
            if col == 8:
                cell.font = Font(bold=True, color="228B22")

    col_widths = [6, 40, 40, 12, 25, 40, 12, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'


def create_fuzzy_matches_sheet(wb):
    """Create sheet for fuzzy matches (needs review)"""
    ws = wb.create_sheet("4. Needs Review")

    headers = [
        "No.",
        "Original Description",
        "Normalized Description",
        "Similarity %",
        "Suggested Work Code",
        "Suggested Description",
        "SEC Code",
        "Action"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_ORANGE
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    sample_data = [
        (1, "Xây tường gạch đặc dày 220", "Xây tường - gạch đặc - dày 220", 88.5, "S03-WALL-BRICK-0008", "Xây tường gạch đặc - dày 200", "SEC-03", "REVIEW"),
        (2, "Trát tường trong nhà dày 15mm", "Trát tường trong - dày 15mm - M75", 82.3, "S04-PLAS-INT-0001", "Trát tường trong - M75", "SEC-04", "REVIEW"),
    ]

    for row_idx, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            cell.fill = FUZZY_MATCH_FILL
            if col == 1 or col == 4 or col == 8:
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN
            if col == 8:
                cell.font = Font(bold=True, color="FF8C00")

    col_widths = [6, 40, 40, 12, 25, 40, 12, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'


def create_new_items_sheet(wb):
    """Create sheet for new items"""
    ws = wb.create_sheet("5. New Items")

    headers = [
        "No.",
        "Original Description",
        "Normalized Description",
        "Suggested SEC Code",
        "Suggested Work Code",
        "Unit",
        "Action"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_RED
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    sample_data = [
        (1, "Cáp Cu/XLPE/PVC 4x300mm2", "Lắp đặt Cáp Cu/XLPE/PVC - 4x300mm2", "SEC-08", "S08-ELEC-CABL-XXXX", "m", "ADD TO MASTER"),
        (2, "Trồng cây Bàng Đài Loan H3-4m", "Trồng Cây Bàng Đài Loan - H3-4m", "SEC-12", "S12-LAND-TREE-XXXX", "cây", "ADD TO MASTER"),
        (3, "Lắp đặt đèn LED panel 40W", "Lắp đặt đèn LED panel - 40W", "SEC-08", "S08-ELEC-LITE-XXXX", "bộ", "ADD TO MASTER"),
    ]

    for row_idx, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            cell.fill = NEW_ITEM_FILL
            if col == 1 or col == 6 or col == 7:
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN
            if col == 7:
                cell.font = Font(bold=True, color="C00000")

    col_widths = [6, 45, 45, 15, 25, 10, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'


def create_master_reference_sheet(wb):
    """Create sheet with Master database reference"""
    ws = wb.create_sheet("6. Master Reference")

    headers = [
        "Work Code",
        "Description",
        "SEC Code",
        "Unit",
        "Price (Min)",
        "Price (Avg)",
        "Price (Max)",
        "Occurrences",
        "Verified"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_PURPLE
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    sample_data = [
        ("S01-EARTH-EXCAV-0001", "Đào đất hố móng - máy đào", "SEC-01", "m3", 45000, 52000, 65000, 25, "Yes"),
        ("S01-EARTH-FILL-0003", "Đắp đất - đất mua mới - K95", "SEC-01", "m3", 85000, 95000, 110000, 18, "Yes"),
        ("S02-CONC-M200-0001", "Bê tông lót móng - M100", "SEC-02", "m3", 850000, 920000, 1050000, 42, "Yes"),
        ("S02-CONC-M300-0001", "Bê tông móng - M300 - thương phẩm", "SEC-02", "m3", 1250000, 1380000, 1520000, 38, "Yes"),
        ("S02-CONC-M350-0015", "Bê tông cột - M350", "SEC-02", "m3", 1350000, 1450000, 1600000, 35, "Yes"),
        ("S03-WALL-BRICK-0008", "Xây tường gạch đặc - dày 200", "SEC-03", "m3", 1850000, 2100000, 2350000, 22, "Yes"),
        ("S07-PIPE-HDPE-0022", "Lắp đặt ống HDPE - D110 - PN16", "SEC-07", "m", 125000, 145000, 175000, 15, "No"),
    ]

    for row_idx, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            if col in [5, 6, 7]:
                cell.alignment = RIGHT_ALIGN
                cell.number_format = '#,##0'
            elif col in [8, 9]:
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN

    col_widths = [25, 45, 12, 8, 15, 15, 15, 12, 10]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'


def create_line_items_sheet(wb):
    """Create sheet with line items and SEC classification"""
    ws = wb.create_sheet("7. Line Items Detail")

    headers = [
        "Row",
        "Original Description",
        "Normalized Description",
        "Work Category",
        "SEC Code",
        "Confidence %",
        "Method",
        "Unit",
        "Quantity",
        "Unit Price",
        "Amount",
        "Needs Review",
        "Validation Issues"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    sample_data = [
        (1, "Đổ bê tông móng M300 thương phẩm", "Bê tông móng - M300 - thương phẩm", "concrete_rebar", "SEC-02", 95.5, "auto", "m3", 125.5, 1380000, 173190000, "No", ""),
        (2, "Đổ bê tông cột M350", "Bê tông cột - M350", "concrete_rebar", "SEC-02", 92.3, "auto", "m3", 85.2, 1450000, 123540000, "No", ""),
        (3, "Xây tường gạch đặc dày 220", "Xây tường - gạch đặc - dày 220", "finishing", "SEC-03", 78.5, "auto", "m3", 45.8, 2100000, 96180000, "Yes", "Low confidence"),
        (4, "Lắp đặt ống HDPE D110", "Lắp đặt ống HDPE - D110 - PN16", "steel_mep", "SEC-07", 88.2, "auto", "m", 520, 145000, 75400000, "No", ""),
    ]

    for row_idx, data in enumerate(sample_data, 2):
        conf = data[5]
        if conf >= 90:
            fill = EXACT_MATCH_FILL
        elif conf >= 70:
            fill = FUZZY_MATCH_FILL
        else:
            fill = NEW_ITEM_FILL

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = fill
            cell.border = THIN_BORDER
            if col in [1, 6, 7, 12]:
                cell.alignment = CENTER_ALIGN
            elif col in [9, 10, 11]:
                cell.alignment = RIGHT_ALIGN
                if col == 10 or col == 11:
                    cell.number_format = '#,##0'
            else:
                cell.alignment = LEFT_ALIGN

    col_widths = [6, 40, 40, 18, 12, 12, 10, 8, 12, 15, 18, 12, 25]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'


def main():
    """Generate template Excel file"""
    wb = Workbook()

    # Create all sheets
    create_summary_sheet(wb)
    create_all_items_sheet(wb)
    create_exact_matches_sheet(wb)
    create_fuzzy_matches_sheet(wb)
    create_new_items_sheet(wb)
    create_master_reference_sheet(wb)
    create_line_items_sheet(wb)

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Save
    output_path = "TEMPLATE_BOQ_Result_After_Processing.xlsx"
    wb.save(output_path)
    print(f"✓ Template created: {output_path}")
    print(f"\nSheets included:")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {name}")


if __name__ == "__main__":
    main()
