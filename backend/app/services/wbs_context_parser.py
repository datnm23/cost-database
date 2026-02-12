"""
WBS Context Parser

Parses Excel BOQ files to extract Work Breakdown Structure (WBS) context:
- Detects indentation levels (leading spaces, cell indent, Roman numerals, bold/size)
- Builds a tree of WBS nodes
- Attaches contextual information (parent, section path, neighbors) to each work item

Output per item:
{
  "parent_title": "Cong tac be tong",
  "section_path": "PHAN THAN > Ket cau > Be tong",
  "neighbors": ["Cot thep dam san D16-D25", "Van khuon dam san"],
  "level": 3,
  "section_type": "concrete"
}
"""
import re
import json
import logging
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


# Header patterns for Vietnamese BOQ
HEADER_PATTERNS = [
    re.compile(r'^\s*PHẦN\s', re.IGNORECASE),
    re.compile(r'^\s*CHƯƠNG\s', re.IGNORECASE),
    re.compile(r'^\s*HẠNG\s*MỤC\s', re.IGNORECASE),
    re.compile(r'^\s*MỤC\s', re.IGNORECASE),
    re.compile(r'^\s*[IVX]+\.\s'),       # Roman numerals: I. II. III. IV.
    re.compile(r'^\s*[A-Z]\.\s'),         # Letter sections: A. B. C.
    re.compile(r'^\s*\d+\.\s+[A-ZĐ]'),   # Numbered section with capital start
]

# Section type detection
SECTION_TYPE_PATTERNS = {
    'earthworks': re.compile(r'(đào|đắp|san|lấp|đất|cọc|ép cọc|khoan nhồi)', re.IGNORECASE),
    'concrete': re.compile(r'(bê\s*tông|cốt\s*thép|ván\s*khuôn|coffa)', re.IGNORECASE),
    'finishing': re.compile(r'(hoàn\s*thiện|trát|lát|ốp|sơn|xây\s*tường|gạch)', re.IGNORECASE),
    'mep_electrical': re.compile(r'(điện|cáp|tủ\s*điện|chiếu\s*sáng|đèn)', re.IGNORECASE),
    'mep_plumbing': re.compile(r'(cấp\s*nước|thoát\s*nước|ống|van|bơm)', re.IGNORECASE),
    'mep_hvac': re.compile(r'(điều\s*hòa|thông\s*gió|HVAC|AHU|FCU)', re.IGNORECASE),
    'mep_fire': re.compile(r'(PCCC|chữa\s*cháy|báo\s*cháy|sprinkler)', re.IGNORECASE),
    'road': re.compile(r'(đường|mặt\s*đường|vỉa\s*hè|BTN|asphalt|rải\s*thảm)', re.IGNORECASE),
    'landscaping': re.compile(r'(cây\s*xanh|trồng\s*cây|cỏ|vườn|cảnh\s*quan)', re.IGNORECASE),
    'steel': re.compile(r'(kết\s*cấu\s*thép|thép\s*hình|nhà\s*thép)', re.IGNORECASE),
}


@dataclass
class WBSNode:
    """A node in the WBS tree."""
    row_index: int
    text: str
    level: int
    is_header: bool = False
    children: List['WBSNode'] = field(default_factory=list)
    parent: Optional['WBSNode'] = None

    def path(self) -> str:
        """Build section path from root to this node."""
        parts = []
        node = self
        while node is not None:
            if node.is_header:
                parts.append(node.text.strip())
            node = node.parent
        return ' > '.join(reversed(parts))


@dataclass
class WBSContext:
    """Contextual information for a work item."""
    parent_title: Optional[str] = None
    section_path: str = ''
    neighbors: List[str] = field(default_factory=list)
    level: int = 0
    section_type: Optional[str] = None

    def to_dict(self) -> Dict:
        return {
            'parent_title': self.parent_title,
            'section_path': self.section_path,
            'neighbors': self.neighbors,
            'level': self.level,
            'section_type': self.section_type,
        }

    def to_json(self) -> str:
        return json.dumps(self.to_dict(), ensure_ascii=False)


class WBSContextParser:
    """Parses Excel BOQ files to extract WBS context for each work item."""

    def parse_excel(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[int, WBSContext]:
        """
        Parse an Excel file and return WBS context for each row.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to parse (default: first sheet)

        Returns:
            Dict mapping row_index -> WBSContext
        """
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not available for WBS parsing")
            return {}

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

            rows = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=False)):
                # Find the description column (usually the longest text column)
                text, indent = self._extract_text_and_indent(row)
                if text:
                    rows.append((row_idx, text, indent))

            wb.close()

            if not rows:
                return {}

            # Detect levels
            levels = self.detect_indentation(rows)

            # Build tree
            nodes = self.build_tree(rows, levels)

            # Attach context
            return self.attach_context(nodes)

        except Exception as e:
            logger.error(f"Failed to parse WBS from Excel: {e}")
            return {}

    def parse_from_dataframe(self, texts: List[str], desc_col_values: Optional[List] = None) -> Dict[int, WBSContext]:
        """
        Parse WBS context from a list of description texts.

        Args:
            texts: List of description strings (in order)
            desc_col_values: Optional raw cell values with indent info

        Returns:
            Dict mapping index -> WBSContext
        """
        if not texts:
            return {}

        rows = [(i, text, self._count_leading_spaces(text)) for i, text in enumerate(texts) if text and text.strip()]

        if not rows:
            return {}

        levels = self.detect_indentation(rows)
        nodes = self.build_tree(rows, levels)
        return self.attach_context(nodes)

    def detect_indentation(self, rows: List[Tuple[int, str, int]]) -> List[int]:
        """
        Detect indentation levels from various sources.

        Sources:
        1. Leading spaces / openpyxl cell indent
        2. Roman numeral / number prefixes
        3. Bold/size indicators (from indent parameter)
        4. Header pattern detection

        Args:
            rows: List of (row_index, text, indent_value)

        Returns:
            List of detected levels (0-based) for each row
        """
        levels = []

        # Collect unique indentation values
        indent_values = sorted(set(indent for _, _, indent in rows))
        indent_to_level = {v: i for i, v in enumerate(indent_values)}

        for row_idx, text, indent in rows:
            stripped = text.strip()

            # Check if this is a top-level header
            if self.is_header_row(stripped):
                # Determine header level from pattern
                header_level = self._header_level(stripped)
                levels.append(header_level)
            else:
                # Use indentation-based level
                base_level = indent_to_level.get(indent, 0)

                # Boost level for numbered items (1.1, 1.2.3, etc.)
                number_depth = self._number_prefix_depth(stripped)
                if number_depth > 0:
                    base_level = max(base_level, number_depth)

                levels.append(base_level)

        return levels

    def build_tree(self, rows: List[Tuple[int, str, int]], levels: List[int]) -> List[WBSNode]:
        """
        Build a tree of WBSNode objects from rows and their levels.

        Args:
            rows: List of (row_index, text, indent)
            levels: List of levels corresponding to rows

        Returns:
            List of all WBSNode objects (flat list, with parent/children linked)
        """
        nodes = []
        stack: List[WBSNode] = []  # Stack of (node) maintaining current path

        for (row_idx, text, _indent), level in zip(rows, levels):
            stripped = text.strip()
            is_header = self.is_header_row(stripped)

            node = WBSNode(
                row_index=row_idx,
                text=stripped,
                level=level,
                is_header=is_header,
            )

            # Find parent: pop stack until we find a node with level < current
            while stack and stack[-1].level >= level:
                stack.pop()

            if stack:
                node.parent = stack[-1]
                stack[-1].children.append(node)

            stack.append(node)
            nodes.append(node)

        return nodes

    def attach_context(self, nodes: List[WBSNode]) -> Dict[int, WBSContext]:
        """
        Attach WBS context to each non-header node.

        Args:
            nodes: List of WBSNode objects

        Returns:
            Dict mapping row_index -> WBSContext
        """
        contexts = {}
        non_header_nodes = [n for n in nodes if not n.is_header]

        for i, node in enumerate(non_header_nodes):
            # Find parent header
            parent_title = None
            current = node.parent
            while current:
                if current.is_header:
                    parent_title = current.text
                    break
                current = current.parent

            # Build section path
            section_path = node.path() if node.parent else ''

            # Get neighbors (up to 2 before and 2 after)
            neighbors = []
            for j in range(max(0, i - 2), min(len(non_header_nodes), i + 3)):
                if j != i:
                    neighbors.append(non_header_nodes[j].text)

            # Detect section type
            section_type = self._detect_section_type(section_path or node.text)

            contexts[node.row_index] = WBSContext(
                parent_title=parent_title,
                section_path=section_path,
                neighbors=neighbors[:4],  # Limit to 4 neighbors
                level=node.level,
                section_type=section_type,
            )

        return contexts

    def is_header_row(self, text: str) -> bool:
        """
        Detect if a row is a header/section row.

        Patterns:
        - PHAN, CHUONG, HANG MUC, Roman numerals
        - All-caps text with no unit/quantity indicators
        - Bold formatting (detected by indent parameter)
        """
        if not text or not text.strip():
            return False

        stripped = text.strip()

        # Check against header patterns
        for pattern in HEADER_PATTERNS:
            if pattern.search(stripped):
                return True

        # All-caps Vietnamese text (at least 3 words, no digits mixed in)
        words = stripped.split()
        if len(words) >= 2:
            upper_words = sum(1 for w in words if w.isupper() and len(w) > 1)
            if upper_words >= len(words) * 0.6:
                # Check it doesn't have unit/quantity indicators
                if not re.search(r'\b(m2|m3|kg|tấn|bộ|cái|md)\b', stripped, re.IGNORECASE):
                    return True

        return False

    # ── Internal helpers ──

    def _extract_text_and_indent(self, row) -> Tuple[str, int]:
        """Extract description text and indent level from an Excel row."""
        best_text = ''
        best_indent = 0

        for cell in row:
            if cell.value and isinstance(cell.value, str) and len(str(cell.value).strip()) > len(best_text):
                best_text = str(cell.value).strip()
                # Try to get cell indent from alignment
                try:
                    if cell.alignment and cell.alignment.indent:
                        best_indent = int(cell.alignment.indent)
                    else:
                        best_indent = self._count_leading_spaces(str(cell.value))
                except (AttributeError, TypeError):
                    best_indent = self._count_leading_spaces(str(cell.value))

        return best_text, best_indent

    def _count_leading_spaces(self, text: str) -> int:
        """Count leading spaces in text."""
        if not text:
            return 0
        return len(text) - len(text.lstrip())

    def _header_level(self, text: str) -> int:
        """Determine header level from text pattern."""
        stripped = text.strip()

        if re.match(r'^\s*PHẦN\s', stripped, re.IGNORECASE):
            return 0
        if re.match(r'^\s*CHƯƠNG\s', stripped, re.IGNORECASE):
            return 1
        if re.match(r'^\s*HẠNG\s*MỤC\s', stripped, re.IGNORECASE):
            return 1
        if re.match(r'^\s*[IVX]+\.\s', stripped):
            # Roman numeral depth
            roman = re.match(r'^\s*([IVX]+)', stripped)
            if roman:
                return 1
        if re.match(r'^\s*[A-Z]\.\s', stripped):
            return 2
        if re.match(r'^\s*\d+\.\s', stripped):
            return 2

        return 1

    def _number_prefix_depth(self, text: str) -> int:
        """Detect numbering depth from prefix (1. = 1, 1.1 = 2, 1.1.1 = 3)."""
        match = re.match(r'^\s*(\d+(?:\.\d+)*)\.\s', text)
        if match:
            return match.group(1).count('.') + 1
        return 0

    def _detect_section_type(self, text: str) -> Optional[str]:
        """Detect section type from text content."""
        for section_type, pattern in SECTION_TYPE_PATTERNS.items():
            if pattern.search(text):
                return section_type
        return None


# Singleton
_wbs_parser = None


def get_wbs_context_parser() -> WBSContextParser:
    """Get or create WBS context parser singleton."""
    global _wbs_parser
    if _wbs_parser is None:
        _wbs_parser = WBSContextParser()
    return _wbs_parser
