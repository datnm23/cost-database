"""
Merged Cell Handler

Handles multi-level merged headers in Excel BOQ files.
Flattens hierarchical header structures into single-row headers.
"""
import logging
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


@dataclass
class MergedHeaderResult:
    """Result of merged header processing."""
    column_names: List[str]
    is_merged: bool
    header_depth: int  # Number of rows used for header
    raw_header_rows: List[List[str]]  # Original header rows


class MergedCellHandler:
    """
    Handler for multi-level merged headers.

    Example:
        Row 1: |     Vật liệu     |     Nhân công     |
        Row 2: | Đơn giá | Thành tiền | Đơn giá | Thành tiền |

        Flattens to:
        ['Vật liệu - Đơn giá', 'Vật liệu - Thành tiền',
         'Nhân công - Đơn giá', 'Nhân công - Thành tiền']
    """

    # Maximum header depth to consider
    MAX_HEADER_DEPTH = 5

    # Separator for flattened headers
    SEPARATOR = ' - '

    def __init__(self):
        """Initialize merged cell handler."""
        pass

    def process_headers(
        self,
        file_path: str,
        sheet_name: str,
        header_row: int,
        num_columns: int
    ) -> MergedHeaderResult:
        """
        Process headers with merged cell handling.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to process
            header_row: Detected header row (0-indexed)
            num_columns: Number of columns in the data

        Returns:
            MergedHeaderResult with flattened column names
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook[sheet_name]

            # Detect header depth by analyzing merged cells
            header_depth, merged_ranges = self._detect_header_depth(
                worksheet, header_row, num_columns
            )

            logger.info(f"Detected header depth: {header_depth} rows")

            if header_depth <= 1 and not merged_ranges:
                # Simple header, no merging
                column_names = self._read_simple_header(
                    worksheet, header_row, num_columns
                )
                return MergedHeaderResult(
                    column_names=column_names,
                    is_merged=False,
                    header_depth=1,
                    raw_header_rows=[column_names]
                )

            # Build header matrix
            header_matrix = self._build_header_matrix(
                worksheet, header_row, header_depth, num_columns, merged_ranges
            )

            # Flatten headers
            column_names = self._flatten_headers(header_matrix)

            workbook.close()

            return MergedHeaderResult(
                column_names=column_names,
                is_merged=True,
                header_depth=header_depth,
                raw_header_rows=header_matrix
            )

        except Exception as e:
            logger.error(f"Error processing merged headers: {e}")
            # Fall back to simple approach
            return MergedHeaderResult(
                column_names=[f"Column_{i+1}" for i in range(num_columns)],
                is_merged=False,
                header_depth=1,
                raw_header_rows=[]
            )

    def _detect_header_depth(
        self,
        worksheet: Worksheet,
        header_row: int,
        num_columns: int
    ) -> Tuple[int, Dict[Tuple[int, int], Tuple[int, int, int, int]]]:
        """
        Detect header depth by analyzing merged cell ranges.

        Returns:
            Tuple of (depth, merged_ranges dict)
            merged_ranges: {(row, col): (min_row, min_col, max_row, max_col)}
        """
        merged_ranges: Dict[Tuple[int, int], Tuple[int, int, int, int]] = {}
        max_depth = 1

        # Excel uses 1-based indexing
        excel_row = header_row + 1

        for merge_range in worksheet.merged_cells.ranges:
            min_row = merge_range.min_row
            max_row = merge_range.max_row
            min_col = merge_range.min_col
            max_col = merge_range.max_col

            # Check if this merge involves our header area
            if min_row <= excel_row + self.MAX_HEADER_DEPTH and min_col <= num_columns:
                # Store the merge range for each cell in it
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        merged_ranges[(row, col)] = (min_row, min_col, max_row, max_col)

                # Update max depth if this merge spans multiple rows
                if max_row > min_row:
                    depth = max_row - excel_row + 1
                    if min_row >= excel_row:
                        max_depth = max(max_depth, depth)

        # Check if there are column-spanning merges in header area (horizontal merges)
        has_col_spanning_merges = any(
            merge_range.max_col > merge_range.min_col
            for merge_range in worksheet.merged_cells.ranges
            if merge_range.min_row >= excel_row and merge_range.min_row <= excel_row + 2
        )

        # Check for row-spanning merges
        has_row_spanning_merges = any(
            merge_range.max_row > merge_range.min_row
            for merge_range in worksheet.merged_cells.ranges
            if merge_range.min_row >= excel_row and merge_range.min_row <= excel_row + 2
        )

        # If we have column-spanning merges, check if there's a second header row
        if has_col_spanning_merges and max_depth == 1:
            # Check the next row for text content that could be sub-headers
            next_row_idx = excel_row + 1
            if next_row_idx <= worksheet.max_row:
                text_cells = 0
                numeric_cells = 0
                for col in range(1, min(num_columns + 1, worksheet.max_column + 1)):
                    cell = worksheet.cell(row=next_row_idx, column=col)
                    if cell.value and not isinstance(cell, MergedCell):
                        val_str = str(cell.value).strip()
                        # Check if it contains letters (header-like)
                        if val_str and any(c.isalpha() for c in val_str):
                            text_cells += 1
                        elif val_str and any(c.isdigit() for c in val_str):
                            numeric_cells += 1

                # If second row has more text than numbers, it's likely a sub-header row
                if text_cells > 0 and text_cells >= numeric_cells:
                    max_depth = 2

        # If we have row-spanning merges, extend depth
        if has_row_spanning_merges and max_depth == 1:
            # Find the max row span
            for merge_range in worksheet.merged_cells.ranges:
                if merge_range.min_row >= excel_row and merge_range.max_row > merge_range.min_row:
                    depth = merge_range.max_row - excel_row + 1
                    max_depth = max(max_depth, depth)

        return min(max_depth, self.MAX_HEADER_DEPTH), merged_ranges

    def _read_simple_header(
        self,
        worksheet: Worksheet,
        header_row: int,
        num_columns: int
    ) -> List[str]:
        """Read simple (non-merged) header row."""
        excel_row = header_row + 1
        headers = []

        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=excel_row, column=col)
            value = cell.value

            if value is None:
                headers.append(f"Column_{col}")
            else:
                headers.append(str(value).strip())

        return headers

    def _build_header_matrix(
        self,
        worksheet: Worksheet,
        header_row: int,
        header_depth: int,
        num_columns: int,
        merged_ranges: Dict[Tuple[int, int], Tuple[int, int, int, int]]
    ) -> List[List[str]]:
        """
        Build header matrix with inherited values from merged parents.

        Returns list of rows, each row is list of header values.
        """
        excel_start_row = header_row + 1
        matrix: List[List[str]] = []

        for depth in range(header_depth):
            excel_row = excel_start_row + depth
            row_values: List[str] = []

            for col in range(1, num_columns + 1):
                cell = worksheet.cell(row=excel_row, column=col)

                # Check if this cell is part of a merge
                if (excel_row, col) in merged_ranges:
                    min_row, min_col, max_row, max_col = merged_ranges[(excel_row, col)]
                    # Get value from the top-left cell of merge
                    value = worksheet.cell(row=min_row, column=min_col).value
                else:
                    value = cell.value

                if value is None:
                    row_values.append('')
                else:
                    row_values.append(str(value).strip())

            matrix.append(row_values)

        return matrix

    def _flatten_headers(self, header_matrix: List[List[str]]) -> List[str]:
        """
        Flatten multi-row headers into single column names.

        Concatenates non-duplicate, non-empty values per column.
        """
        if not header_matrix:
            return []

        num_columns = len(header_matrix[0])
        flattened: List[str] = []

        for col_idx in range(num_columns):
            parts: List[str] = []
            seen: set = set()

            for row in header_matrix:
                if col_idx < len(row):
                    value = row[col_idx].strip()
                    # Skip empty or duplicate values
                    if value and value.lower() not in seen:
                        parts.append(value)
                        seen.add(value.lower())

            if parts:
                flattened.append(self.SEPARATOR.join(parts))
            else:
                flattened.append(f"Column_{col_idx + 1}")

        return flattened


# Module-level singleton
_merged_cell_handler: Optional[MergedCellHandler] = None


def get_merged_cell_handler() -> MergedCellHandler:
    """Get or create singleton MergedCellHandler instance."""
    global _merged_cell_handler
    if _merged_cell_handler is None:
        _merged_cell_handler = MergedCellHandler()
    return _merged_cell_handler
