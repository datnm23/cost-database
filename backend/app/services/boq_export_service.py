"""
BOQ Export Service - Export processed BOQ results to Excel

Creates structured Excel templates with:
1. Summary sheet with statistics
2. Processed items with normalization results
3. Matching results against Master database
4. Items needing review
5. New items to add to Master
6. Original format preservation for Excel files

Note: Unit standardization can be managed via database (unit_standards, sec_code_default_units tables)
or falls back to hardcoded values below. Use the /api/v1/units endpoints to manage units.
"""
import os
import re
from datetime import datetime
from typing import List, Dict, Optional, Any, Tuple
from dataclasses import dataclass
import logging
import shutil

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.hyperlink import Hyperlink

from sqlalchemy.orm import Session

from app.models.line_item import LineItem
from app.models.master_work_item import MasterWorkItem
from app.models.boq_file import BOQFile
from app.models.project import Project
from app.services.boq_processing_service import BOQProcessingService, ProcessingResult, MatchResult
from app.services.work_code_generator import WorkCodeGenerator

logger = logging.getLogger(__name__)

# =====================================================
# FALLBACK CONSTANTS (used when database tables are empty)
# For production, manage units via /api/v1/units endpoints
# =====================================================

# SEC Code Default Unit Mapping - per Vietnamese construction standards (TCVN) and FIDIC/QS
SEC_CODE_DEFAULT_UNIT = {
    # Level 1 defaults
    'SEC-00': 'trọn gói',  # Preliminaries - lump sum
    'SEC-01': 'm³',        # Substructure - volume
    'SEC-02': 'm³',        # Superstructure - volume
    'SEC-03': 'm²',        # Architecture - area
    'SEC-04': 'bộ',        # MEP - set/unit
    'SEC-05': 'm²',        # Landscape - area

    # Level 2 - Substructure
    'SEC-01-01': 'm³',     # Earthworks - volume
    'SEC-01-02': 'm',      # Piling - linear meter
    'SEC-01-03': 'm³',     # Foundation - volume

    # Level 2 - Superstructure
    'SEC-02-01': 'm³',     # Concrete - volume
    'SEC-02-02': 'm³',     # Slab - volume
    'SEC-02-03': 'm³',     # Beam - volume
    'SEC-02-04': 'm³',     # Column - volume
    'SEC-02-05': 'm³',     # Wall - volume
    'SEC-02-06': 'kg',     # Rebar/Cốt thép - weight

    # Level 2 - Architecture
    'SEC-03-01': 'm³',     # Masonry - volume (per TCVN)
    'SEC-03-02': 'm²',     # Plastering - area
    'SEC-03-03': 'm²',     # Painting - area
    'SEC-03-04': 'm²',     # Tiling - area
    'SEC-03-05': 'm²',     # Ceiling - area
    'SEC-03-06': 'bộ',     # Door & Window - set

    # Level 2 - MEP
    'SEC-04-01': 'điểm',   # Electrical - points
    'SEC-04-02': 'điểm',   # Plumbing - points
    'SEC-04-03': 'bộ',     # HVAC - set
    'SEC-04-04': 'bộ',     # Fire - set

    # Level 2 - Landscape
    'SEC-05-01': 'm²',     # Road - area
    'SEC-05-02': 'm²',     # Pavement - area
    'SEC-05-03': 'cây',    # Greenery - tree count
}

# Unit Standardization Mapping - Vietnamese convention preferred
UNIT_STANDARDIZATION = {
    # Volume - Vietnamese superscript preferred
    'm3': 'm³', 'm³': 'm³', 'mét khối': 'm³', 'm khối': 'm³',
    'khối': 'm³', 'cbm': 'm³', 'cubic meter': 'm³', 'cubic m': 'm³',

    # Area
    'm2': 'm²', 'm²': 'm²', 'mét vuông': 'm²', 'm vuông': 'm²',
    'sqm': 'm²', 'sq.m': 'm²', 'square meter': 'm²', 'sq m': 'm²',

    # Length
    'm': 'm', 'mét': 'm', 'met': 'm', 'meter': 'm',
    'mm': 'mm', 'milimet': 'mm', 'milimét': 'mm',
    'cm': 'cm', 'centimet': 'cm', 'centimét': 'cm',
    'km': 'km',

    # Weight
    'kg': 'kg', 'kilo': 'kg', 'kilogram': 'kg', 'kilôgam': 'kg', 'kí lô': 'kg',
    'tấn': 'tấn', 'tan': 'tấn', 'ton': 'tấn', 't': 'tấn', 'tonne': 'tấn',

    # Count/Piece
    'cái': 'cái', 'chiếc': 'cái', 'pc': 'cái', 'pcs': 'cái',
    'ea': 'cái', 'each': 'cái', 'piece': 'cái', 'no': 'cái', 'nos': 'cái',
    'bộ': 'bộ', 'set': 'bộ', 'combo': 'bộ',
    'điểm': 'điểm', 'point': 'điểm', 'pt': 'điểm',
    'cây': 'cây', 'tree': 'cây',

    # Other
    'lô': 'lô', 'lot': 'lô',
    'ls': 'trọn gói', 'lump sum': 'trọn gói', 'trọn gói': 'trọn gói', 'l.s': 'trọn gói',
    'công': 'công', 'man-day': 'công', 'ngày công': 'công', 'nc': 'công',
    'lít': 'lít', 'liter': 'lít', 'l': 'lít', 'litre': 'lít',
    'giờ': 'giờ', 'hour': 'giờ', 'hr': 'giờ', 'h': 'giờ',
    'ngày': 'ngày', 'day': 'ngày', 'd': 'ngày',
    'tháng': 'tháng', 'month': 'tháng', 'mo': 'tháng',
}

# SEC Code Display Mapping - converts SEC codes to readable acronym format
SEC_CODE_DISPLAY_MAP = {
    # Level 1
    'SEC-00': '00.PRELIM',      # Preliminaries
    'SEC-01': '01.SUBSTRUCT',   # Substructure
    'SEC-02': '02.SUPERSTRUCT', # Superstructure
    'SEC-03': '03.ARCH',        # Architecture & Finishes
    'SEC-04': '04.MEP',         # MEP Systems
    'SEC-05': '05.LANDSCAPE',   # Landscape & External

    # Level 2 - Substructure
    'SEC-01-01': '01.01.EARTH',   # Earthworks / Đào đất
    'SEC-01-02': '01.02.PILE',    # Piling / Cọc
    'SEC-01-03': '01.03.FOUND',   # Foundation / Móng

    # Level 2 - Superstructure
    'SEC-02-01': '02.01.CONC',    # Concrete Frame / BTCT
    'SEC-02-02': '02.02.SLAB',    # Floor Slab / Sàn
    'SEC-02-03': '02.03.BEAM',    # Beam / Dầm
    'SEC-02-04': '02.04.COL',     # Column / Cột
    'SEC-02-05': '02.05.WALL',    # Wall / Tường
    'SEC-02-06': '02.06.REBAR',   # Rebar / Cốt thép

    # Level 2 - Architecture
    'SEC-03-01': '03.01.MASON',   # Masonry / Tường xây
    'SEC-03-02': '03.02.PLAST',   # Plastering / Trát
    'SEC-03-03': '03.03.PAINT',   # Painting / Sơn
    'SEC-03-04': '03.04.TILE',    # Tiling / Lát gạch
    'SEC-03-05': '03.05.CEIL',    # Ceiling / Trần
    'SEC-03-06': '03.06.DOOR',    # Door & Window / Cửa

    # Level 2 - MEP
    'SEC-04-01': '04.01.ELEC',    # Electrical / Điện
    'SEC-04-02': '04.02.PLUMB',   # Plumbing / Nước
    'SEC-04-03': '04.03.HVAC',    # HVAC / Điều hòa
    'SEC-04-04': '04.04.FIRE',    # Fire Protection / PCCC

    # Level 2 - Landscape
    'SEC-05-01': '05.01.ROAD',    # Road / Đường
    'SEC-05-02': '05.02.PAVE',    # Pavement / Vỉa hè
    'SEC-05-03': '05.03.GREEN',   # Greenery / Cây xanh
}

# Default specifications for categories
CONCRETE_DEFAULTS = {
    'default_grade': 'M250',
    'default_stone': 'Đá 1x2',  # Replaces thương phẩm/thủ công
}

EXCAVATION_DEFAULTS = {
    'default_method': 'máy đào 0.8m3',  # Default excavator
}

FILL_DEFAULTS = {
    'default_grade': 'K90',  # Default compaction (not K95)
    # No default soil source - detect from description
}


def format_sec_code_display(sec_code: str) -> str:
    """Convert SEC code to acronym display format (e.g., 01.01.EARTH)."""
    return SEC_CODE_DISPLAY_MAP.get(sec_code, sec_code or "")


def get_default_unit_for_sec_code(sec_code: str) -> str:
    """
    Get default unit for a SEC code, falling back to parent level.

    Args:
        sec_code: SEC code like 'SEC-02-01' or 'SEC-02'

    Returns:
        Default unit string (e.g., 'm³', 'm²', 'bộ')
    """
    if not sec_code:
        return ""

    # Try exact match first
    if sec_code in SEC_CODE_DEFAULT_UNIT:
        return SEC_CODE_DEFAULT_UNIT[sec_code]

    # Try parent level (SEC-XX-YY -> SEC-XX)
    parts = sec_code.split('-')
    if len(parts) >= 2:
        parent = f"{parts[0]}-{parts[1]}"
        if parent in SEC_CODE_DEFAULT_UNIT:
            return SEC_CODE_DEFAULT_UNIT[parent]

    return ""


def standardize_unit(unit: str) -> str:
    """
    Standardize unit notation to Vietnamese convention.

    Args:
        unit: Raw unit string (e.g., 'm3', 'sqm', 'pcs')

    Returns:
        Standardized unit (e.g., 'm³', 'm²', 'cái')
    """
    if not unit:
        return ""

    unit_clean = unit.strip().lower()
    return UNIT_STANDARDIZATION.get(unit_clean, unit)


def get_unit_with_default(raw_unit: str, sec_code: str) -> Tuple[str, bool]:
    """
    Get standardized unit, falling back to SEC code default if empty.

    Args:
        raw_unit: Original unit from BOQ
        sec_code: SEC code for default fallback

    Returns:
        Tuple of (standardized_unit, is_default_applied)
    """
    if raw_unit and raw_unit.strip():
        return standardize_unit(raw_unit), False

    default = get_default_unit_for_sec_code(sec_code)
    return default, bool(default)


def inject_default_specs(description: str, category: str) -> Tuple[str, List[str]]:
    """
    Inject missing specs into generic descriptions.

    Rules:
    - Excavation (đào): Add default equipment "máy đào 0.8m3" if no method specified
    - Fill (đắp): Detect soil source from text + add default K90 if not specified
    - Concrete: Add M250 + Đá 1x2 if not specified

    Args:
        description: Normalized description
        category: Work category (concrete, earthworks, etc.)

    Returns:
        tuple: (enhanced_description, list_of_defaults_applied)
    """
    defaults_applied = []
    enhanced = description
    category_lower = (category or "").lower()
    desc_lower = description.lower()

    # Concrete defaults (M250, Đá 1x2)
    if category_lower in ('concrete', 'be tong', 'bê tông', 'btct') or 'bê tông' in desc_lower:
        # Check if grade missing
        if not re.search(r'\bM\d{2,3}\b', description, re.I):
            enhanced += f" - {CONCRETE_DEFAULTS['default_grade']}"
            defaults_applied.append(f"Mác mặc định: {CONCRETE_DEFAULTS['default_grade']}")

        # Check if stone type missing (thương phẩm/thủ công → Đá 1x2)
        if not re.search(r'đá\s*\d+[xX×]\d+', description, re.I):
            if 'thương phẩm' not in desc_lower and 'thủ công' not in desc_lower:
                enhanced += f" - {CONCRETE_DEFAULTS['default_stone']}"
                defaults_applied.append(f"Loại đá: {CONCRETE_DEFAULTS['default_stone']}")

    # Excavation defaults (máy đào 0.8m3) - only if no method specified
    # Must start with "đào" or have "đào đất/đào hố/đào móng" pattern to avoid false positives
    elif ('đào đất' in desc_lower or 'đào hố' in desc_lower or 'đào móng' in desc_lower
          or 'đào nền' in desc_lower or desc_lower.startswith('đào')) and 'đắp' not in desc_lower:
        # Check if excavation method already specified
        has_method = any(kw in desc_lower for kw in ['thủ công', 'máy đào', 'đào đá', 'máy xúc'])
        if not has_method:
            enhanced += f" - {EXCAVATION_DEFAULTS['default_method']}"
            defaults_applied.append(f"Phương pháp mặc định: {EXCAVATION_DEFAULTS['default_method']}")

    # Fill/Backfill defaults - detect soil source + add K90 if missing
    elif 'đắp' in desc_lower:
        # Detect soil source from description (don't add default if not specified)
        # "tận dụng" → keep as is, "mua mới" → keep as is
        # Note: Don't add soil source if not specified - we don't know which one

        # Add compaction grade if not specified
        # Match K90, K95, K98 or K=0,95 / K=0.95 formats
        if not re.search(r'\bK\d{2}\b|\bK\s*=\s*0[,\.]\d+', description, re.I):
            enhanced += f" - {FILL_DEFAULTS['default_grade']}"
            defaults_applied.append(f"Độ chặt mặc định: {FILL_DEFAULTS['default_grade']}")

    # Plastering defaults (thickness 15mm)
    elif 'trát' in desc_lower:
        if not re.search(r'dày\s*\d+', description, re.I) and not re.search(r'\d+\s*mm', description, re.I):
            enhanced += " - dày 15mm"
            defaults_applied.append("Độ dày mặc định: 15mm")

    return enhanced, defaults_applied


def format_defaults_explanation(details: List[str]) -> str:
    """
    Format list of applied defaults as multi-line text for Excel.

    Args:
        details: List like ['Category bonus: +25', 'Mac: M250']

    Returns:
        Multi-line string:
        • Category bonus: +25
        • Mác mặc định: M250
    """
    if not details:
        return ""
    return "\n".join(f"• {item}" for item in details)


def combine_description_with_specs(description: str, tech_specs: str) -> str:
    """
    Combine description with technical specifications for normalization.

    Args:
        description: Original work description (e.g., "Bê tông móng")
        tech_specs: Technical specifications column (e.g., "M250, đá 1x2, slump 10cm")

    Returns:
        Combined string for normalization: "Bê tông móng - M250, đá 1x2, slump 10cm"
    """
    if not tech_specs or not tech_specs.strip():
        return description

    # Clean up specs
    specs_clean = tech_specs.strip()

    # Avoid duplication if specs already in description
    if specs_clean.lower() in description.lower():
        return description

    return f"{description} - {specs_clean}"


def extract_tech_specs_from_line_item(line_item) -> str:
    """
    Extract technical specifications from line item.

    Checks common column names: 'technical_specs', 'specs', 'quy_cach', 'thong_so_ky_thuat'
    """
    # Check various possible attribute names
    for attr in ['technical_specs', 'specs', 'quy_cach', 'thong_so_ky_thuat', 'specifications']:
        if hasattr(line_item, attr):
            value = getattr(line_item, attr)
            if value and str(value).strip():
                return str(value).strip()
    return ""


def get_or_generate_work_code(
    item,
    db: Session,
    generator: Optional["WorkCodeGenerator"] = None
) -> Tuple[str, bool]:
    """
    Get work code from matched master or generate for new items.

    Args:
        item: MatchResult or line item object
        db: Database session
        generator: Optional WorkCodeGenerator instance (created if not provided)

    Returns:
        Tuple of (work_code, is_generated)
        - is_generated=False means code came from master database
        - is_generated=True means code was auto-generated for new item
    """
    # If already has master work code (matched), use it
    if hasattr(item, 'master_work_code') and item.master_work_code:
        return item.master_work_code, False

    # For new/unmatched items, generate work code
    if generator is None:
        generator = WorkCodeGenerator(db)

    # Get SEC code from item
    sec_code = ""
    if hasattr(item, 'suggested_matches') and item.suggested_matches:
        sec_code = item.suggested_matches[0].get('sec_code', '')
    elif hasattr(item, 'sec_code'):
        sec_code = item.sec_code or ""

    # Get description for code generation
    description = ""
    if hasattr(item, 'normalized_description') and item.normalized_description:
        description = item.normalized_description
    elif hasattr(item, 'original_description'):
        description = item.original_description

    if not description:
        return "", False

    # Generate work code
    work_code = generator.generate_work_code(
        description=description,
        sec_code=sec_code or "SEC-00",
        include_grade=True  # Include M250, M75 etc. in code
    )

    return work_code, True


# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FILL_GREEN = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
HEADER_FILL_ORANGE = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
HEADER_FILL_RED = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
HEADER_FILL_PURPLE = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

EXACT_MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FUZZY_MATCH_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
NEW_ITEM_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
WRAP_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Column indices that need wrap_text (for multi-line content)
WRAP_TEXT_COLUMNS = [5, 10]  # Tên Chuẩn Hóa, Ghi chú xử lý (position 10 after adding Đơn vị)


class BOQExportService:
    """Service to export processed BOQ results to Excel"""

    def __init__(self, db: Session):
        self.db = db
        self.processing_service = BOQProcessingService(db)

    def export_processing_result(
        self,
        file_id: int,
        output_path: str,
        include_master_match: bool = True
    ) -> str:
        """
        Export BOQ processing result to Excel file

        Args:
            file_id: BOQ file ID
            output_path: Path to save Excel file
            include_master_match: Include matching against Master database

        Returns:
            Path to created file
        """
        # Get file info
        boq_file = self.db.query(BOQFile).filter(BOQFile.file_id == file_id).first()
        if not boq_file:
            raise ValueError(f"BOQ file {file_id} not found")

        project = self.db.query(Project).filter(Project.project_id == boq_file.project_id).first()

        # Process BOQ
        result = self.processing_service.process_line_items(file_id)

        # Create workbook
        wb = Workbook()

        # Sheet 1: Summary
        self._create_summary_sheet(wb, boq_file, project, result)

        # Sheet 2: All Processed Items
        self._create_processed_items_sheet(wb, result)

        # Sheet 3: Exact Matches
        self._create_match_sheet(wb, result, 'exact', "Exact Matches", HEADER_FILL_GREEN)

        # Sheet 4: Fuzzy Matches (Needs Review)
        self._create_match_sheet(wb, result, 'fuzzy', "Needs Review", HEADER_FILL_ORANGE)

        # Sheet 5: New Items
        self._create_match_sheet(wb, result, 'new', "New Items", HEADER_FILL_RED)

        # Sheet 6: Master Database Reference
        if include_master_match:
            self._create_master_reference_sheet(wb)

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Save
        wb.save(output_path)
        logger.info(f"Exported BOQ processing result to {output_path}")

        return output_path

    def _create_summary_sheet(
        self,
        wb: Workbook,
        boq_file: BOQFile,
        project: Optional[Project],
        result: ProcessingResult
    ):
        """Create summary sheet with statistics"""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "BOQ PROCESSING RESULT"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = CENTER_ALIGN

        # File Info
        row = 3
        info_data = [
            ("File Name:", boq_file.file_name),
            ("Project:", project.project_name if project else "N/A"),
            ("Project Code:", project.project_code if project else "N/A"),
            ("Processed At:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

        for label, value in info_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        # Statistics
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "PROCESSING STATISTICS"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].fill = HEADER_FILL
        ws[f'A{row}'].font = HEADER_FONT
        row += 1

        stats = [
            ("Total Extracted", result.total_extracted, "100%"),
            ("After Raw Dedupe", result.unique_raw, f"{result.unique_raw/result.total_extracted*100:.1f}%"),
            ("After Normalization", result.unique_normalized, f"{result.unique_normalized/result.total_extracted*100:.1f}%"),
            ("", "", ""),
            ("Exact Matches (≥95%)", result.exact_matches, f"{result.exact_matches/result.unique_normalized*100:.1f}%" if result.unique_normalized else "0%"),
            ("Fuzzy Matches (80-95%)", result.fuzzy_matches, f"{result.fuzzy_matches/result.unique_normalized*100:.1f}%" if result.unique_normalized else "0%"),
            ("New Items (<80%)", result.new_items, f"{result.new_items/result.unique_normalized*100:.1f}%" if result.unique_normalized else "0%"),
            ("", "", ""),
            ("New Items (Deduped)", result.new_items_deduped, "Ready to add to Master"),
        ]

        headers = ["Metric", "Count", "Percentage/Note"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1

        for metric, count, pct in stats:
            ws.cell(row=row, column=1, value=metric)
            if count != "":
                ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=pct)
            row += 1

        # Match Type Legend
        row += 2
        ws[f'A{row}'] = "MATCH TYPE LEGEND"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        legends = [
            ("Exact Match", "≥95% similarity - Auto assign work code", EXACT_MATCH_FILL),
            ("Fuzzy Match", "80-95% similarity - Needs review", FUZZY_MATCH_FILL),
            ("New Item", "<80% similarity - New work item", NEW_ITEM_FILL),
        ]

        for name, desc, fill in legends:
            ws.cell(row=row, column=1, value=name).fill = fill
            ws.cell(row=row, column=2, value=desc)
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25

    def _create_processed_items_sheet(self, wb: Workbook, result: ProcessingResult):
        """Create sheet with all processed items"""
        ws = wb.create_sheet("All Items")

        # New column structure per plan (with Đơn vị at position 9)
        headers = [
            "STT",                    # Row number
            "Mã Công Tác Chuẩn",      # Master work code
            "Tên Công Việc Gốc",      # Original description
            "Thông số KT gốc",        # Original tech specs
            "Tên Chuẩn Hóa",          # Normalized with injected defaults
            "Mã Master Mapping",      # Master work code (same as #2)
            "Tỷ lệ khớp (%)",         # Similarity score
            "Nhóm (Category)",        # SEC code display
            "Đơn vị",                 # Standardized unit with default
            "Ghi chú xử lý"           # Processing notes
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # Initialize WorkCodeGenerator for new items
        generator = WorkCodeGenerator(self.db)

        # Write data
        for row_idx, item in enumerate(result.items, 2):
            # Determine fill based on match type
            if item.match_type == 'exact':
                fill = EXACT_MATCH_FILL
            elif item.match_type == 'fuzzy':
                fill = FUZZY_MATCH_FILL
            else:
                fill = NEW_ITEM_FILL

            # Extract tech specs from the item
            tech_specs = extract_tech_specs_from_line_item(item)

            # Combine description with tech specs
            combined_desc = combine_description_with_specs(
                item.normalized_description or "",
                tech_specs
            )

            # Get category from SEC code or work category
            category = ""
            sec_code = ""
            if item.suggested_matches:
                sec_code = item.suggested_matches[0].get('sec_code', '')

            # Try to get category from item if available
            if hasattr(item, 'work_category') and item.work_category:
                category = item.work_category

            # Inject default specs based on category
            enhanced_desc, defaults_applied = inject_default_specs(combined_desc, category)

            # Format SEC code for display
            category_display = format_sec_code_display(sec_code)

            # Get raw unit from item
            raw_unit = ""
            if hasattr(item, 'unit') and item.unit:
                raw_unit = item.unit

            # Standardize unit with default fallback
            standardized_unit, unit_is_default = get_unit_with_default(raw_unit, sec_code)

            # Add unit default note if applicable
            if unit_is_default and standardized_unit:
                defaults_applied.append(f"Đơn vị mặc định: {standardized_unit}")

            # Get or generate work code
            work_code, is_generated = get_or_generate_work_code(item, self.db, generator)
            if is_generated and work_code:
                defaults_applied.append(f"Mã tự sinh: {work_code}")

            # Format defaults explanation with bullet points
            notes = format_defaults_explanation(defaults_applied)

            # Prepare data row (10 columns now with Đơn vị)
            data = [
                row_idx - 1,  # STT
                work_code,  # Mã Công Tác Chuẩn (from master or generated)
                item.original_description[:200] if len(item.original_description) > 200 else item.original_description,  # Tên Công Việc Gốc
                tech_specs,  # Thông số KT gốc
                enhanced_desc[:200] if len(enhanced_desc) > 200 else enhanced_desc,  # Tên Chuẩn Hóa
                work_code,  # Mã Master Mapping (same as column 2)
                round(item.similarity_score * 100, 1),  # Tỷ lệ khớp (%)
                category_display,  # Nhóm (Category)
                standardized_unit,  # Đơn vị
                notes  # Ghi chú xử lý
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = fill
                cell.border = THIN_BORDER

                # Apply appropriate alignment
                if col in [1, 6, 7, 9]:  # STT, Mã Master Mapping, Tỷ lệ khớp, Đơn vị
                    cell.alignment = CENTER_ALIGN
                elif col in WRAP_TEXT_COLUMNS:  # Tên Chuẩn Hóa, Ghi chú xử lý
                    cell.alignment = WRAP_ALIGN
                else:
                    cell.alignment = LEFT_ALIGN

        # Adjust row heights for multi-line content
        for row in range(2, ws.max_row + 1):
            # Count line breaks in "Ghi chú xử lý" column (column 10 now)
            notes_cell = ws.cell(row=row, column=10)
            if notes_cell.value:
                line_count = str(notes_cell.value).count('\n') + 1
                ws.row_dimensions[row].height = max(15, line_count * 15)

        # Column widths per plan: [6, 20, 45, 30, 50, 20, 10, 20, 10, 35]
        col_widths = [6, 20, 45, 30, 50, 20, 10, 20, 10, 35]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

    def _create_match_sheet(
        self,
        wb: Workbook,
        result: ProcessingResult,
        match_type: str,
        sheet_name: str,
        header_fill: PatternFill
    ):
        """Create sheet for specific match type with new column structure"""
        ws = wb.create_sheet(sheet_name)

        # Filter items
        items = [i for i in result.items if i.match_type == match_type]

        if not items:
            ws['A1'] = f"No {match_type} matches found"
            return

        # New unified column structure for all match types (with Đơn vị at position 9)
        headers = [
            "STT",                    # Row number
            "Mã Công Tác Chuẩn",      # Master work code
            "Tên Công Việc Gốc",      # Original description
            "Thông số KT gốc",        # Original tech specs
            "Tên Chuẩn Hóa",          # Normalized with injected defaults
            "Mã Master Mapping",      # Master work code (same as #2)
            "Tỷ lệ khớp (%)",         # Similarity score
            "Nhóm (Category)",        # SEC code display
            "Đơn vị",                 # Standardized unit with default
            "Ghi chú xử lý"           # Processing notes
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = header_fill
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # Initialize WorkCodeGenerator for new items
        generator = WorkCodeGenerator(self.db)

        # Write data
        for row_idx, item in enumerate(items, 2):
            # Extract tech specs from the item
            tech_specs = extract_tech_specs_from_line_item(item)

            # Combine description with tech specs
            combined_desc = combine_description_with_specs(
                item.normalized_description or "",
                tech_specs
            )

            # Get category and SEC code
            category = ""
            sec_code = ""
            if item.suggested_matches:
                sec_code = item.suggested_matches[0].get('sec_code', '')

            # Try to get category from item if available
            if hasattr(item, 'work_category') and item.work_category:
                category = item.work_category

            # Inject default specs based on category
            enhanced_desc, defaults_applied = inject_default_specs(combined_desc, category)

            # Format SEC code for display
            category_display = format_sec_code_display(sec_code)

            # Get raw unit from item
            raw_unit = ""
            if hasattr(item, 'unit') and item.unit:
                raw_unit = item.unit

            # Standardize unit with default fallback
            standardized_unit, unit_is_default = get_unit_with_default(raw_unit, sec_code)

            # Add unit default note if applicable
            if unit_is_default and standardized_unit:
                defaults_applied.append(f"Đơn vị mặc định: {standardized_unit}")

            # Get or generate work code
            work_code, is_generated = get_or_generate_work_code(item, self.db, generator)
            if is_generated and work_code:
                defaults_applied.append(f"Mã tự sinh: {work_code}")

            # Add match-type specific notes
            if match_type == 'new':
                defaults_applied.append("Hành động: THÊM VÀO MASTER")
            elif match_type == 'exact':
                defaults_applied.append("Hành động: TỰ ĐỘNG PHÊ DUYỆT")
            else:  # fuzzy
                defaults_applied.append("Hành động: CẦN XEM XÉT")

            # Format defaults explanation with bullet points
            notes = format_defaults_explanation(defaults_applied)

            # Prepare data row (10 columns now with Đơn vị)
            data = [
                row_idx - 1,  # STT
                work_code,  # Mã Công Tác Chuẩn (from master or generated)
                item.original_description[:150] if item.original_description else "",  # Tên Công Việc Gốc
                tech_specs,  # Thông số KT gốc
                enhanced_desc[:150] if len(enhanced_desc) > 150 else enhanced_desc,  # Tên Chuẩn Hóa
                work_code,  # Mã Master Mapping (same as column 2)
                round(item.similarity_score * 100, 1),  # Tỷ lệ khớp (%)
                category_display,  # Nhóm (Category)
                standardized_unit,  # Đơn vị
                notes  # Ghi chú xử lý
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = THIN_BORDER

                # Apply appropriate alignment
                if col in [1, 6, 7, 9]:  # STT, Mã Master Mapping, Tỷ lệ khớp, Đơn vị
                    cell.alignment = CENTER_ALIGN
                elif col in WRAP_TEXT_COLUMNS:  # Tên Chuẩn Hóa, Ghi chú xử lý
                    cell.alignment = WRAP_ALIGN
                else:
                    cell.alignment = LEFT_ALIGN

        # Adjust row heights for multi-line content
        for row in range(2, ws.max_row + 1):
            # Count line breaks in "Ghi chú xử lý" column (column 10 now)
            notes_cell = ws.cell(row=row, column=10)
            if notes_cell.value:
                line_count = str(notes_cell.value).count('\n') + 1
                ws.row_dimensions[row].height = max(15, line_count * 15)

        # Column widths per plan: [6, 20, 45, 30, 50, 20, 10, 20, 10, 35]
        col_widths = [6, 20, 45, 30, 50, 20, 10, 20, 10, 35]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

    def _create_master_reference_sheet(self, wb: Workbook):
        """Create sheet with Master database reference"""
        ws = wb.create_sheet("Master Reference")

        # Get top 100 master items
        master_items = self.db.query(MasterWorkItem).filter(
            MasterWorkItem.is_active == True
        ).order_by(MasterWorkItem.occurrence_count.desc()).limit(100).all()

        if not master_items:
            ws['A1'] = "No master items found"
            return

        headers = [
            "Work Code",
            "Description",
            "SEC Code",
            "Unit",
            "Price (Min)",
            "Price (Avg)",
            "Price (Max)",
            "Occurrences",
            "Verified"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL_PURPLE
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # Write data
        for row_idx, item in enumerate(master_items, 2):
            data = [
                item.work_code,
                item.description[:100] if item.description and len(item.description) > 100 else item.description,
                item.sec_code,
                item.unit_standard,
                float(item.ref_unit_price_min) if item.ref_unit_price_min else "",
                float(item.ref_unit_price_avg) if item.ref_unit_price_avg else "",
                float(item.ref_unit_price_max) if item.ref_unit_price_max else "",
                item.occurrence_count,
                "Yes" if item.is_verified else "No"
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = THIN_BORDER
                if col in [5, 6, 7]:
                    cell.alignment = RIGHT_ALIGN
                    if value:
                        cell.number_format = '#,##0'
                elif col in [8, 9]:
                    cell.alignment = CENTER_ALIGN
                else:
                    cell.alignment = LEFT_ALIGN

        # Column widths
        col_widths = [25, 60, 12, 8, 15, 15, 15, 12, 10]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

    def export_line_items_with_classification(
        self,
        file_id: int,
        output_path: str
    ) -> str:
        """
        Export line items with SEC classification results

        Args:
            file_id: BOQ file ID
            output_path: Path to save Excel file

        Returns:
            Path to created file
        """
        # Get line items
        line_items = self.db.query(LineItem).filter(
            LineItem.file_id == file_id
        ).order_by(LineItem.row_number).all()

        if not line_items:
            raise ValueError(f"No line items found for file {file_id}")

        # Get file info
        boq_file = self.db.query(BOQFile).filter(BOQFile.file_id == file_id).first()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Line Items"

        headers = [
            "Row",
            "Original Description",
            "Normalized Description",
            "Work Category",
            "SEC Code",
            "Confidence %",
            "Method",
            "Unit",
            "Quantity",
            "Unit Price",
            "Amount",
            "Needs Review",
            "Validation Issues"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # Write data
        for row_idx, item in enumerate(line_items, 2):
            # Determine fill based on confidence
            if item.confidence_score and item.confidence_score >= 90:
                fill = EXACT_MATCH_FILL
            elif item.confidence_score and item.confidence_score >= 70:
                fill = FUZZY_MATCH_FILL
            elif item.needs_review:
                fill = NEW_ITEM_FILL
            else:
                fill = PatternFill()  # No fill

            data = [
                item.row_number,
                item.description[:150] if item.description else "",
                item.normalized_description[:150] if item.normalized_description else "",
                item.work_category or "",
                item.sec_code or "",
                float(item.confidence_score) if item.confidence_score else "",
                item.classification_method.value if item.classification_method else "",
                item.unit or "",
                float(item.quantity) if item.quantity else "",
                float(item.unit_price) if item.unit_price else "",
                float(item.amount) if item.amount else "",
                "Yes" if item.needs_review else "No",
                item.validation_issues or ""
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                if fill.fill_type:
                    cell.fill = fill
                cell.border = THIN_BORDER

                # Alignment and formatting
                if col in [1, 6, 7, 12]:
                    cell.alignment = CENTER_ALIGN
                elif col in [9, 10, 11]:
                    cell.alignment = RIGHT_ALIGN
                    if value:
                        cell.number_format = '#,##0.00' if col == 10 else '#,##0'
                else:
                    cell.alignment = LEFT_ALIGN

        # Column widths
        col_widths = [6, 50, 50, 20, 12, 12, 10, 8, 12, 15, 18, 12, 30]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save
        wb.save(output_path)
        logger.info(f"Exported line items to {output_path}")

        return output_path

    def export_with_original_format(
        self,
        file_id: int,
        output_path: str
    ) -> str:
        """
        Export with original Excel format preserved

        This method:
        1. Copies the original Excel file
        2. Adds a new "Processing Results" sheet
        3. Preserves all original formatting, comments, and notes

        Args:
            file_id: BOQ file ID
            output_path: Path to save the new Excel file

        Returns:
            Path to the created file
        """
        # Get file info
        boq_file = self.db.query(BOQFile).filter(BOQFile.file_id == file_id).first()
        if not boq_file:
            raise ValueError(f"BOQ file {file_id} not found")

        original_path = boq_file.file_path
        if not original_path or not os.path.exists(original_path):
            raise ValueError(f"Original file not found at: {original_path}")

        # Copy original file to preserve all formatting
        shutil.copy2(original_path, output_path)
        logger.info(f"Copied original file from {original_path} to {output_path}")

        # Load the copied workbook
        try:
            wb = load_workbook(output_path)
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            raise ValueError(f"Failed to load Excel file: {e}")

        # Get line items
        line_items = self.db.query(LineItem).filter(
            LineItem.file_id == file_id
        ).order_by(LineItem.row_number).all()

        # Create new sheet for processing results
        if "Processing Results" in wb.sheetnames:
            del wb["Processing Results"]

        ws = wb.create_sheet("Processing Results", 0)

        # Add header with styling
        headers = [
            "Row No.",
            "Original Description",
            "Normalized Description",
            "Work Category",
            "SEC Code",
            "Confidence %",
            "Match Type",
            "Master Work Code",
            "Unit",
            "Quantity",
            "Unit Price",
            "Amount",
            "Original Sheet",
            "Link to Original"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # Write data with links back to original rows
        for row_idx, item in enumerate(line_items, 2):
            # Determine fill based on confidence/match type
            if item.confidence_score and item.confidence_score >= 95:
                fill = EXACT_MATCH_FILL
            elif item.confidence_score and item.confidence_score >= 80:
                fill = FUZZY_MATCH_FILL
            elif item.needs_review:
                fill = NEW_ITEM_FILL
            else:
                fill = PatternFill()

            # Get match type value safely
            match_type_val = ""
            if hasattr(item, 'match_type') and item.match_type:
                match_type_val = item.match_type.value if hasattr(item.match_type, 'value') else str(item.match_type)

            # Get master work code if matched
            master_code = ""
            if hasattr(item, 'matched_master') and item.matched_master:
                master_code = item.matched_master.work_code

            # Original sheet name
            sheet_name = item.original_sheet_name if hasattr(item, 'original_sheet_name') else ""

            data = [
                item.row_number,
                item.description[:200] if item.description else "",
                item.normalized_description[:200] if item.normalized_description else "",
                item.work_category or "",
                item.sec_code or "",
                float(item.confidence_score) if item.confidence_score else "",
                match_type_val,
                master_code,
                item.unit or "",
                float(item.quantity) if item.quantity else "",
                float(item.unit_price) if item.unit_price else "",
                float(item.amount) if item.amount else "",
                sheet_name,
                ""  # Link will be added below
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                if fill.fill_type:
                    cell.fill = fill
                cell.border = THIN_BORDER

                if col in [1, 6, 7]:
                    cell.alignment = CENTER_ALIGN
                elif col in [10, 11, 12]:
                    cell.alignment = RIGHT_ALIGN
                    if value:
                        cell.number_format = '#,##0.00' if col == 11 else '#,##0'
                else:
                    cell.alignment = LEFT_ALIGN

            # Add hyperlink to original row if we have sheet name and row number
            if sheet_name and item.row_number and sheet_name in wb.sheetnames:
                link_cell = ws.cell(row=row_idx, column=14)
                link_cell.value = f"Go to Row {item.row_number}"
                link_cell.hyperlink = f"#'{sheet_name}'!A{item.row_number}"
                link_cell.font = Font(color="0563C1", underline="single")

        # Column widths
        col_widths = [8, 50, 50, 15, 12, 12, 12, 20, 8, 12, 15, 18, 20, 15]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add summary info at the top of sheet
        ws.insert_rows(1, 3)
        ws['A1'] = "BOQ PROCESSING RESULTS"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"File: {boq_file.file_name}"
        ws['A3'] = f"Processed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Save
        wb.save(output_path)
        logger.info(f"Exported with original format to {output_path}")

        return output_path


def get_boq_export_service(db: Session) -> BOQExportService:
    """Factory function to get BOQ export service"""
    return BOQExportService(db)
