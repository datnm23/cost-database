#!/usr/bin/env python3
"""
WBS Cost Coding Processor
Based on WBS_CostCoding_Manual.md v4.0

This script processes BOQ Excel files and generates SEC codes and Spec codes
according to the WBS Cost Coding system.
"""

import re
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# ============================================================================
# SEC CODE PATTERNS - Comprehensive list based on WBS Manual
# ============================================================================

SEC_CODE_PATTERNS = [
    # ========================================================================
    # GROUP 00: SOFT COSTS
    # ========================================================================
    (r'thi công tạm|lán trại|bảo vệ|hàng rào tạm', '00-01-00-00'),
    (r'thiết kế|tư vấn|quy hoạch', '00-02-00-00'),
    (r'khảo sát địa chất|quan trắc|đo đạc', '00-03-00-00'),
    (r'an toàn lao động|ppe|hse', '00-04-00-00'),
    (r'bảo hiểm|thuế', '00-05-00-00'),
    (r'giấy phép|thủ tục|nghiệm thu', '00-06-00-00'),

    # ========================================================================
    # GROUP 01: SUBSTRUCTURE - Foundation & Earthwork
    # ========================================================================
    # 01-01: Earthwork for building
    (r'đào\s*(đất|hố|móng|nền).*cấp\s*(i|1|ii|2)', '01-01-01-10'),
    (r'đào\s*(đất|hố|móng|nền).*cấp\s*(iii|3|iv|4)', '01-01-01-12'),
    (r'đào\s*(đất|hố|móng|nền).*đá|phá đá', '01-01-01-13'),
    (r'đào\s*(đất|hố|móng|nền)', '01-01-01-10'),
    (r'đắp\s*đất\s*(móng|công trình)', '01-01-01-20'),
    (r'lu\s*lèn|đầm\s*đất', '01-01-01-20'),

    # 01-02: Piling
    (r'cọc\s*ép.*d\s*[3-4]\d{2}', '01-02-01-10'),
    (r'cọc\s*khoan\s*nhồi|cọc\s*barrette', '01-02-01-20'),
    (r'cọc\s*xi\s*măng\s*đất', '01-02-01-30'),
    (r'cọc\s*(btct|bê\s*tông)', '01-02-01-10'),

    # 01-03: Foundation Structure
    # Concrete
    (r'bê\s*tông.*(móng|đài|giằng).*thủ\s*công', '01-03-01-10'),
    (r'bê\s*tông.*(móng|đài|giằng).*bơm', '01-03-01-20'),
    (r'bê\s*tông.*(móng|đài|giằng)', '01-03-01-20'),
    (r'bê\s*tông\s*lót|bê\s*tông\s*gạch\s*vỡ', '01-03-01-10'),
    # Formwork
    (r'ván\s*khuôn.*(móng|đài|giằng).*gỗ', '01-03-02-10'),
    (r'ván\s*khuôn.*(móng|đài|giằng).*thép', '01-03-02-20'),
    (r'ván\s*khuôn.*(móng|đài|giằng)', '01-03-02-10'),
    # Rebar
    (r'thép.*(móng|đài|giằng).*d\s*(6|8|10)\s*mm', '01-03-03-10'),
    (r'thép.*(móng|đài|giằng).*d\s*(12|14|16|18)\s*mm', '01-03-03-18'),
    (r'thép.*(móng|đài|giằng).*d\s*(20|22|25|28|32)\s*mm', '01-03-03-19'),
    (r'thép.*(móng|đài|giằng)', '01-03-03-18'),
    (r'cốt\s*thép.*(móng|đài)', '01-03-03-18'),

    # 01-04: Basement walls
    (r'tường\s*hầm|tường\s*vây', '01-04-01-20'),
    (r'chống\s*thấm.*ngầm|chống\s*thấm.*hầm', '01-04-02-20'),

    # Generic concrete/formwork/rebar (fallback patterns)
    (r'bê\s*tông\s*đá\s*1x2|bê\s*tông\s*đá', '01-03-01-20'),
    (r'bê\s*tông\s*m\s*(150|200|250|300)', '01-03-01-20'),
    (r'bê\s*tông\s*tấm\s*đan|tấm\s*đan.*bê\s*tông', '01-03-01-20'),
    (r'ván\s*khuôn\s*tấm\s*đan', '01-03-02-10'),
    (r'ván\s*khuôn\s*bê\s*tông', '01-03-02-10'),
    (r'thép\s*d\s*<=?\s*10|thép.*d\s*(6|8|10)\s*mm', '01-03-03-10'),
    (r'gạch\s*đặc|gạch\s*xây|vữa\s*xm', '03-01-01-10'),
    (r'đá\s*1x2|đá\s*chờ|đá\s*đặt\s*chờ', '01-03-01-90'),

    # ========================================================================
    # GROUP 02: SUPERSTRUCTURE
    # ========================================================================
    # 02-01: RC Frame
    # Concrete
    (r'bê\s*tông.*(cột|vách|lõi).*thủ\s*công', '02-01-01-10'),
    (r'bê\s*tông.*(cột|vách|lõi).*bơm', '02-01-01-20'),
    (r'bê\s*tông.*(cột|vách|lõi)', '02-01-01-20'),
    (r'bê\s*tông.*(dầm|sàn|lanh tô).*thủ\s*công', '02-01-01-10'),
    (r'bê\s*tông.*(dầm|sàn|lanh tô).*bơm', '02-01-01-20'),
    (r'bê\s*tông.*(dầm|sàn|lanh tô)', '02-01-01-20'),
    (r'bê\s*tông\s*cầu\s*thang', '02-01-01-20'),
    # Formwork
    (r'ván\s*khuôn.*(cột|vách|lõi)', '02-01-02-20'),
    (r'ván\s*khuôn.*(dầm|sàn)', '02-01-02-20'),
    (r'ván\s*khuôn.*cầu\s*thang', '02-01-02-20'),
    # Rebar
    (r'thép.*(cột|vách|dầm|sàn).*d\s*(6|8|10)\s*mm', '02-01-03-10'),
    (r'thép.*(cột|vách|dầm|sàn).*d\s*(12|14|16|18)\s*mm', '02-01-03-18'),
    (r'thép.*(cột|vách|dầm|sàn).*d\s*(20|22|25|28|32)\s*mm', '02-01-03-19'),
    (r'thép.*(cột|vách|dầm|sàn)', '02-01-03-18'),
    (r'cốt\s*thép.*(cột|vách|dầm|sàn)', '02-01-03-18'),

    # 02-02: Structural Steel
    (r'thép\s*h\s*beam|thép\s*hình\s*h', '02-02-04-10'),
    (r'thép\s*i\s*beam|thép\s*hình\s*i', '02-02-04-10'),
    (r'thép\s*hộp|thép\s*box', '02-02-04-10'),
    (r'thép\s*góc|thép\s*v|thép\s*l', '02-02-04-10'),
    (r'thép\s*c|thép\s*u', '02-02-04-10'),
    (r'kết\s*cấu\s*thép', '02-02-04-10'),

    # 02-04: Roof structure
    (r'giàn\s*thép|xà\s*gồ|cầu\s*phong', '02-04-01-10'),
    (r'dàn\s*mái|kết\s*cấu\s*mái', '02-04-01-10'),

    # ========================================================================
    # GROUP 03: INTERIOR ARCHITECTURE
    # ========================================================================
    # 03-01: Walls
    (r'xây\s*(gạch|tường).*10\s*cm', '03-01-01-10'),
    (r'xây\s*(gạch|tường).*20\s*cm', '03-01-01-20'),
    (r'xây\s*(gạch|tường)', '03-01-01-10'),
    (r'tường\s*gạch|xây\s*gạch', '03-01-01-10'),
    (r'vách\s*thạch\s*cao|vách\s*ngăn.*thạch\s*cao', '03-01-02-10'),
    (r'vách\s*ngăn.*kính', '03-01-02-20'),
    (r'trát\s*tường|trát\s*trong|trát\s*vữa', '03-01-04-10'),

    # 03-02: Ceiling
    (r'trần\s*thạch\s*cao.*phẳng', '03-02-01-10'),
    (r'trần\s*thạch\s*cao.*giật\s*cấp', '03-02-01-20'),
    (r'trần\s*thạch\s*cao', '03-02-01-10'),
    (r'trần\s*nhôm', '03-02-01-20'),

    # 03-03: Floor/Wall tiles
    (r'lát\s*(gạch|nền).*3[0-5]0\s*x', '03-03-01-10'),
    (r'lát\s*(gạch|nền).*[4-5]\d{2}\s*x', '03-03-01-10'),
    (r'lát\s*(gạch|nền).*6[0-5]0\s*x', '03-03-01-20'),
    (r'lát\s*(gạch|nền).*8\d{2}\s*x', '03-03-01-30'),
    (r'lát\s*(gạch|nền)', '03-03-01-10'),
    (r'ốp\s*gạch|ốp\s*tường', '03-03-01-10'),
    (r'lát\s*đá\s*granite', '03-03-02-20'),
    (r'lát\s*đá\s*(marble|cẩm\s*thạch)', '03-03-02-30'),
    (r'lát\s*đá', '03-03-02-20'),
    (r'sàn\s*gỗ|lát\s*gỗ', '03-03-03-20'),
    (r'sàn\s*epoxy', '03-03-04-20'),

    # 03-04: Paint
    (r'sơn\s*lót', '03-04-05-10'),
    (r'sơn\s*(nước|trong\s*nhà).*2\s*lớp', '03-04-05-10'),
    (r'sơn\s*(nước|trong\s*nhà).*3\s*lớp', '03-04-05-20'),
    (r'sơn\s*(nước|trong\s*nhà)', '03-04-05-10'),
    (r'sơn\s*epoxy', '03-04-05-30'),
    (r'sơn\s*pu', '03-04-05-30'),
    (r'bả\s*ma\s*tit|bả\s*matit', '03-04-01-10'),

    # 03-05: Interior doors
    (r'cửa\s*gỗ|cửa\s*hdf|cửa\s*mdf', '03-05-01-10'),
    (r'cửa\s*nhựa|cửa\s*composite', '03-05-01-10'),
    (r'cửa\s*chống\s*cháy', '03-05-01-30'),

    # 03-06: Railing
    (r'lan\s*can.*cầu\s*thang|tay\s*vịn.*cầu\s*thang', '03-06-01-10'),
    (r'lan\s*can.*kính', '03-06-01-20'),
    (r'lan\s*can.*inox', '03-06-01-10'),

    # 03-07: Sanitary
    (r'bồn\s*cầu|xí\s*bệt|toilet|wc', '03-07-01-10'),
    (r'lavabo|chậu\s*rửa\s*mặt', '03-07-02-10'),
    (r'sen\s*vòi|vòi\s*sen|shower', '03-07-03-10'),
    (r'bồn\s*tắm|bathtub', '03-07-04-20'),
    (r'gương\s*soi|gương\s*wc', '03-07-05-10'),

    # 03-08: Kitchen
    (r'bồn\s*rửa\s*bếp|chậu\s*rửa\s*bếp', '03-08-01-10'),
    (r'vòi\s*bếp', '03-08-02-10'),

    # ========================================================================
    # GROUP 04: MEP - ELECTRICAL
    # ========================================================================
    # 04-01: Electrical System
    (r'tủ\s*(điện|db|msb|mdb|acb)(?!.*pccc|.*chữa\s*cháy)', '04-01-01-01'),
    (r'tủ\s*tầng|tủ\s*phân\s*phối|tủ\s*điện\s*tầng', '04-01-01-02'),
    (r'tủ\s*chiếu\s*sáng|tủ\s*đèn', '04-01-01-03'),

    # Cables (based on size)
    (r'cáp.*(1c|1x|1\s*x)\s*[1-9]\.?[0-9]*\s*mm', '04-01-02-10'),
    (r'cáp.*(2c|2x|2\s*x)\s*[1-9]\.?[0-9]*\s*mm', '04-01-02-10'),
    (r'cáp.*(3c|3x|3\s*x)\s*[1-9]\.?[0-9]*\s*mm', '04-01-02-10'),
    (r'cáp.*(4c|4x|4\s*x)\s*([1-9]|1[0-6])\s*mm', '04-01-02-10'),
    (r'cáp.*(4c|4x|4\s*x)\s*(25|35|50|70|95|120|150|185|240|300)\s*mm', '04-01-02-20'),
    (r'cáp.*chống\s*cháy|cáp.*fr|cáp.*nhls', '04-01-02-30'),
    (r'cáp.*giáp\s*bảo\s*vệ|cáp.*swa', '04-01-02-40'),
    (r'cáp\s*điện|dây\s*điện', '04-01-02-10'),
    (r'dây\s*đơn|dây\s*cv', '04-01-02-10'),

    # Conduits
    (r'ống\s*luồn\s*dây|ống\s*pvc.*điện|ống\s*hdpe.*điện', '04-01-03-10'),
    (r'ống\s*thép\s*luồn|ống\s*imc|ống\s*emt', '04-01-03-20'),
    (r'máng\s*cáp|cable\s*tray|cable\s*ladder', '04-01-03-30'),

    # Outlets/Switches
    (r'ổ\s*cắm|socket|outlet', '04-01-04-10'),
    (r'công\s*tắc|switch|dimmer', '04-01-04-10'),
    (r'ổ\s*đơn|ổ\s*đôi|ổ\s*ba', '04-01-04-10'),

    # Circuit Breakers - MCCBs, MCBs, ACBs
    (r'mccb|cb\s*(đúc|khối)', '04-01-01-02'),
    (r'mcb|cb\s*tép|cb\s*mini', '04-01-01-03'),
    (r'acb|aptomat\s*(khí|chính)', '04-01-01-01'),
    (r'aptomat|áp\s*tô\s*mát', '04-01-01-02'),
    (r'cb\s*\d+\s*p|cb-\d+p', '04-01-01-02'),
    (r'rccb|rcbo|chống\s*giật', '04-01-01-03'),

    # Fuses & Meters
    (r'cầu\s*chì|fuse', '04-01-01-03'),
    (r'công\s*tơ|đồng\s*hồ\s*điện|meter\s*(điện)?', '04-01-01-03'),

    # Panel/Busbar accessories
    (r'thanh\s*cái|busbar', '04-01-01-02'),
    (r'đấu\s*tủ|đấu\s*nối\s*tủ', '04-01-01-02'),
    (r'vỏ\s*tủ|tủ\s*tôn|tủ\s*sắt', '04-01-01-02'),
    (r'bulong|bu\s*lông', '04-01-01-91'),
    (r'thép\s*hình\s*ldc|thép\s*đặt\s*sẵn', '04-01-01-91'),

    # Indicator lights
    (r'đèn\s*tín\s*hiệu|đèn\s*báo\s*pha', '04-01-04-10'),
    (r'đèn\s*đui\s*xoáy|đèn\s*sợi\s*đốt', '04-05-03-10'),

    # Electrical accessories
    (r'biến\s*dòng|ct\s*\d+/\d+', '04-01-01-03'),
    (r'ổn\s*áp|avr', '04-01-01-03'),
    (r'rơ\s*le\s*thời\s*gian|timer', '04-01-01-03'),
    (r'khóa\s*chuyển|switch\s*selector', '04-01-01-03'),
    (r'contactor', '04-01-01-02'),
    (r'dây\s*đồng\s*trần|dây\s*tiếp\s*địa', '04-06-01-02'),

    # 04-02: Medium voltage
    (r'máy\s*biến\s*áp|trạm\s*biến\s*áp|mba', '04-02-01-01'),
    (r'tủ\s*trung\s*thế', '04-02-01-02'),

    # 04-03: Generator
    (r'máy\s*phát\s*điện|generator|genset', '04-03-01-01'),
    (r'tủ\s*ats|bộ\s*chuyển\s*nguồn', '04-03-01-02'),

    # 04-04: UPS
    (r'ups|bộ\s*lưu\s*điện', '04-04-01-01'),
    (r'acquy|ắc\s*quy|battery', '04-04-01-02'),

    # 04-05: Lighting
    (r'đèn\s*led.*âm\s*trần', '04-05-03-10'),
    (r'đèn\s*panel|đèn\s*tấm', '04-05-03-10'),
    (r'đèn\s*tuýp|đèn\s*tube|đèn\s*huỳnh\s*quang', '04-05-03-10'),
    (r'đèn\s*downlight', '04-05-03-10'),
    (r'đèn\s*pha|đèn\s*rọi', '04-05-03-20'),
    (r'đèn\s*thoát\s*hiểm|đèn\s*exit', '04-05-03-30'),
    (r'đèn\s*sự\s*cố|đèn\s*khẩn\s*cấp', '04-05-03-30'),
    (r'đèn\s*led|đèn\s*chiếu\s*sáng|đèn\s*điện', '04-05-03-10'),

    # 04-06: Lightning protection
    (r'kim\s*thu\s*sét|cột\s*thu\s*sét', '04-06-01-01'),
    (r'dây\s*dẫn\s*sét|dây\s*tiếp\s*địa', '04-06-01-02'),
    (r'cọc\s*tiếp\s*địa|hệ\s*thống\s*tiếp\s*địa', '04-06-01-03'),

    # ========================================================================
    # GROUP 04-10: FIRE PROTECTION (PCCC)
    # ========================================================================
    # 04-10-01: Fire Alarm
    (r'đầu\s*báo\s*(cháy|khói|nhiệt)', '04-10-01-03'),
    (r'tủ\s*trung\s*tâm\s*báo\s*cháy|tủ\s*facp', '04-10-01-01'),
    (r'nút\s*nhấn\s*báo\s*cháy|nút\s*ấn\s*khẩn', '04-10-01-03'),
    (r'còi\s*báo\s*cháy|đèn\s*báo\s*cháy|còi\s*đèn', '04-10-01-03'),
    (r'module\s*báo\s*cháy', '04-10-01-02'),

    # 04-10-02: Water-based Fire Fighting
    (r'bơm\s*(chữa\s*cháy|pccc|cứu\s*hỏa)', '04-10-02-01'),
    (r'sprinkler|đầu\s*phun\s*nước|đầu\s*phun\s*pccc', '04-10-02-03'),
    (r'họng\s*vách\s*tường|họng\s*cứu\s*hỏa|hộp\s*chữa\s*cháy', '04-10-02-03'),
    (r'trụ\s*cứu\s*hỏa|trụ\s*nước\s*chữa\s*cháy', '04-10-02-03'),
    (r'bình\s*chữa\s*cháy|bình\s*cứu\s*hỏa', '04-10-02-03'),
    (r'van\s*pccc|van\s*chữa\s*cháy|alarm\s*valve', '04-10-02-02'),
    (r'ống\s*(pccc|chữa\s*cháy|cứu\s*hỏa)', '04-10-02-02'),

    # 04-10-03: Gas-based Fire Fighting
    (r'fm\s*200|fm200|hệ\s*thống\s*chữa\s*cháy\s*khí', '04-10-03-01'),
    (r'co2.*chữa\s*cháy|khí\s*co2', '04-10-03-01'),

    # 04-10-04: Smoke Exhaust / Pressurization
    (r'quạt\s*hút\s*khói|quạt\s*thông\s*gió\s*pccc', '04-10-04-01'),
    (r'quạt\s*tăng\s*áp|quạt\s*áp\s*lực', '04-10-04-01'),
    (r'cửa\s*gió\s*tăng\s*áp|van\s*gió\s*pccc', '04-10-04-02'),
    (r'ống\s*gió\s*pccc|ống\s*hút\s*khói', '04-10-04-02'),

    # 04-10-05: Passive Fire Protection
    (r'sơn\s*chống\s*cháy|bọc\s*chống\s*cháy', '04-10-05-10'),
    (r'chèn\s*kín.*xuyên\s*sàn|firestop', '04-10-05-10'),

    # ========================================================================
    # GROUP 04-20: PLUMBING
    # ========================================================================
    # 04-21: Water Supply
    (r'ống\s*(ppr|pp-r).*cấp\s*nước', '04-21-02-10'),
    (r'ống\s*(ppr|pp-r)', '04-21-02-10'),
    (r'ống\s*pvc.*cấp\s*nước', '04-21-02-10'),
    (r'ống\s*thép\s*mạ.*cấp|ống\s*gi.*cấp', '04-21-02-20'),
    (r'ống\s*inox.*cấp|ống\s*ss.*cấp', '04-21-02-30'),
    (r'ống\s*cấp\s*nước', '04-21-02-10'),
    (r'van\s*(cổng|bi|bướm).*cấp\s*nước', '04-21-01-10'),
    (r'đồng\s*hồ\s*nước|công\s*tơ\s*nước', '04-21-01-10'),

    # 04-22: Drainage
    (r'ống\s*pvc.*(thoát|thải)', '04-22-02-10'),
    (r'ống\s*hdpe.*(thoát|thải)', '04-22-02-10'),
    (r'ống\s*gang.*(thoát|thải)', '04-22-02-20'),
    (r'ống\s*(thoát|thải)', '04-22-02-10'),
    (r'xi\s*phông|siphon|phễu\s*thu', '04-22-03-10'),
    (r'ống\s*thông\s*hơi', '04-22-04-10'),

    # 04-23: Rainwater
    (r'ống\s*thoát\s*mưa.*trong\s*nhà', '04-23-02-10'),
    (r'phễu\s*thu\s*mưa', '04-23-03-10'),

    # 04-24: Pumps
    (r'bơm\s*cấp\s*nước|bơm\s*sinh\s*hoạt', '04-24-01-10'),
    (r'bơm\s*tăng\s*áp', '04-24-01-20'),
    (r'bơm\s*bể\s*phốt|bơm\s*nước\s*thải', '04-24-01-10'),
    (r'bơm\s*nước|bơm\s*ly\s*tâm', '04-24-01-10'),

    # 04-25: Water Treatment
    (r'bể\s*nước\s*ngầm|bể\s*chứa\s*nước', '04-25-01-10'),
    (r'bồn\s*nước\s*mái|bồn\s*inox', '04-25-01-10'),
    (r'bộ\s*lọc\s*nước|thiết\s*bị\s*lọc', '04-25-02-10'),

    # ========================================================================
    # GROUP 04-30: HVAC
    # ========================================================================
    # 04-31: Central Chiller
    (r'chiller|máy\s*lạnh\s*trung\s*tâm', '04-31-01-01'),
    (r'ahu|air\s*handling\s*unit|bộ\s*xử\s*lý\s*không\s*khí', '04-31-01-02'),
    (r'fcu|fan\s*coil\s*unit|quạt\s*dàn', '04-31-01-03'),
    (r'ống\s*nước\s*lạnh|đường\s*ống.*chiller', '04-31-02-10'),

    # 04-32: VRV/VRF
    (r'vrv|vrf|dàn\s*nóng.*multi', '04-32-01-01'),
    (r'dàn\s*lạnh.*vrv|dàn\s*lạnh.*vrf', '04-32-01-02'),
    (r'ống\s*gas.*vrv|ống\s*đồng.*vrv', '04-32-02-10'),

    # 04-33: Split AC
    (r'điều\s*hòa\s*split|máy\s*lạnh\s*split', '04-33-01-10'),
    (r'điều\s*hòa\s*cassette|máy\s*lạnh.*cassette', '04-33-01-20'),
    (r'điều\s*hòa\s*âm\s*trần|máy\s*lạnh.*âm\s*trần', '04-33-01-20'),
    (r'điều\s*hòa\s*tủ\s*đứng|máy\s*lạnh.*tủ\s*đứng', '04-33-01-20'),
    (r'điều\s*hòa|máy\s*lạnh', '04-33-01-10'),

    # 04-34: Ventilation
    (r'quạt\s*thông\s*gió|quạt\s*hút|exhaust\s*fan', '04-34-01-10'),
    (r'quạt\s*cấp\s*gió|supply\s*fan', '04-34-01-10'),
    (r'ống\s*gió.*tôn|ống\s*gió.*thép', '04-34-02-10'),
    (r'ống\s*gió\s*mềm|ống\s*gió.*flexible', '04-34-02-10'),
    (r'ống\s*gió', '04-34-02-10'),
    (r'miệng\s*gió|cửa\s*gió|louver', '04-34-03-10'),
    (r'van\s*gió|damper', '04-34-03-10'),

    # 04-35: Cooling Tower
    (r'tháp\s*giải\s*nhiệt|cooling\s*tower', '04-35-01-01'),

    # ========================================================================
    # GROUP 04-40: GAS
    # ========================================================================
    (r'bồn\s*gas|bồn\s*lpg', '04-41-01-01'),
    (r'ống\s*gas|đường\s*ống\s*gas', '04-42-02-10'),

    # ========================================================================
    # GROUP 04-50: ELEVATORS
    # ========================================================================
    (r'thang\s*máy\s*khách|thang\s*máy\s*chở\s*khách', '04-51-01-01'),
    (r'thang\s*máy.*hàng|thang\s*máy\s*tải|thang\s*máy\s*dịch\s*vụ', '04-52-01-01'),
    (r'thang\s*máy.*bệnh\s*viện|thang\s*máy.*cáng', '04-52-01-01'),
    (r'thang\s*máy\s*pccc|thang\s*máy.*chữa\s*cháy', '04-54-01-01'),
    (r'thang\s*máy', '04-51-01-01'),
    (r'thang\s*cuốn|escalator', '04-53-01-01'),

    # ========================================================================
    # GROUP 07: ENVELOPE
    # ========================================================================
    # 07-01: Curtain Wall
    (r'vách\s*kính|mặt\s*dựng\s*kính|curtain\s*wall', '07-01-01-20'),
    (r'cửa\s*sổ\s*nhôm|cửa\s*sổ.*kính', '07-01-02-20'),
    (r'kính\s*cường\s*lực|kính\s*an\s*toàn', '07-01-04-20'),
    (r'kính\s*hộp|kính\s*low-e', '07-01-04-30'),
    (r'kính.*6\s*mm|kính.*8\s*mm', '07-01-04-10'),
    (r'kính.*10\s*mm|kính.*12\s*mm', '07-01-04-20'),

    # 07-02: Cladding
    (r'ốp\s*aluminium|ốp\s*nhôm', '07-02-01-20'),
    (r'đá\s*treo|đá\s*ốp\s*ngoài', '07-02-02-20'),
    (r'lam\s*chắn\s*nắng|lam\s*nhôm', '07-02-03-20'),
    (r'sơn\s*ngoài\s*trời|sơn\s*ngoại\s*thất', '07-02-05-20'),

    # 07-03: Waterproofing
    (r'chống\s*thấm.*bitum|màng\s*bitum', '07-03-01-10'),
    (r'chống\s*thấm.*pvc|màng\s*pvc', '07-03-01-20'),
    (r'chống\s*thấm.*pu|polyurethane', '07-03-01-20'),
    (r'chống\s*thấm.*xi\s*măng|chống\s*thấm\s*gốc\s*xi\s*măng', '07-03-01-10'),
    (r'chống\s*thấm.*kết\s*tinh', '07-03-01-30'),
    (r'chống\s*thấm.*mái|chống\s*thấm.*sân\s*thượng', '07-03-01-20'),
    (r'chống\s*thấm', '07-03-01-10'),

    # 07-04: Roofing
    (r'lợp\s*ngói|ngói\s*lợp', '07-04-01-10'),
    (r'lợp\s*tôn|tôn\s*lợp|mái\s*tôn', '07-04-02-10'),
    (r'tấm\s*lấy\s*sáng', '07-04-03-10'),

    # 07-05: Insulation
    (r'cách\s*nhiệt\s*mái', '07-05-01-10'),
    (r'cách\s*nhiệt\s*tường', '07-05-02-10'),

    # 07-06: Exterior doors
    (r'cửa\s*đi.*nhôm|cửa\s*chính.*nhôm', '07-06-02-20'),
    (r'cửa\s*trượt.*nhôm', '07-06-02-20'),
    (r'cổng\s*sắt|cổng\s*nhôm', '07-06-03-10'),

    # 07-07: Exterior railing
    (r'lan\s*can.*ban\s*công|lan\s*can.*ngoài', '07-07-01-10'),

    # ========================================================================
    # GROUP 08: ELV & ICT
    # ========================================================================
    (r'lan|mạng\s*nội\s*bộ|network|cáp\s*mạng', '08-01-01-10'),
    (r'wifi|access\s*point', '08-01-02-10'),
    (r'server|máy\s*chủ', '08-01-03-01'),
    (r'camera|cctv', '08-02-01-02'),
    (r'nvr|dvr', '08-02-01-01'),
    (r'access\s*control|kiểm\s*soát\s*ra\s*vào', '08-03-01-01'),
    (r'thẻ\s*từ|vân\s*tay|face\s*id', '08-03-01-02'),
    (r'báo\s*trộm|cảm\s*biến.*an\s*ninh', '08-04-01-01'),
    (r'loa|ampli|âm\s*thanh', '08-05-01-01'),
    (r'bms|building\s*management', '08-07-01-01'),
    (r'barrier|barrie|thanh\s*chắn.*bãi\s*xe', '08-08-01-01'),
    (r'led\s*display|màn\s*hình\s*led', '08-09-01-01'),

    # ========================================================================
    # GROUP 10: ROADS & EXTERNAL WORKS
    # ========================================================================
    # 10-01: Road Earthwork
    (r'san\s*nền|san\s*lấp\s*mặt\s*bằng', '10-01-01-10'),
    (r'đắp\s*đất\s*k95|lu\s*lèn\s*k95', '10-01-01-20'),
    (r'đắp\s*đất\s*k98|lu\s*lèn\s*k98', '10-01-01-20'),
    (r'đắp\s*đất.*nền\s*đường', '10-01-01-20'),
    (r'đắp\s*đất.*đầm\s*chặt', '10-01-01-20'),
    (r'đắp\s*đất.*k\s*(90|95|98)', '10-01-01-20'),
    (r'đào\s*phá\s*dỡ|đào\s*phá.*nền\s*đường', '10-01-01-10'),
    (r'đào\s*khuôn\s*đường', '10-01-01-10'),
    (r'đào\s*nền\s*đường', '10-01-01-10'),
    (r'đào\s*đất\s*(không\s*thích\s*hợp)?', '10-01-01-10'),
    (r'vận\s*chuyển\s*đất|đổ\s*đất', '10-01-01-90'),
    (r'vải\s*địa\s*kỹ\s*thuật|geotextile', '10-01-02-10'),

    # 10-02: Road Works
    (r'cấp\s*phối\s*đá\s*dăm|cpdd', '10-02-02-10'),
    (r'bê\s*tông\s*nhựa|btn|asphalt', '10-02-03-20'),
    (r'bê\s*tông.*đường|bê\s*tông.*mặt\s*đường', '10-02-03-10'),
    (r'thảm\s*nhựa', '10-02-03-20'),
    (r'đá\s*dăm.*lớp|đá\s*base', '10-02-02-10'),
    (r'tưới\s*lớp\s*thấm|nhựa\s*pha\s*dầu|thấm\s*bám', '10-02-03-20'),
    (r'nilon\s*tái\s*sinh|màng\s*nilon', '10-02-03-90'),
    (r'lớp\s*đá\s*dăm|đá\s*đệm\s*móng', '10-02-02-10'),

    # 10-03: Sidewalk
    (r'vỉa\s*hè|lát\s*vỉa\s*hè', '10-03-01-10'),
    (r'bó\s*vỉa|bó\s*bê\s*tông', '10-03-02-10'),
    (r'xây\s*bó\s*hè|trát\s*bó\s*hè', '10-03-02-10'),
    (r'tấm\s*đan\s*rãnh', '10-03-03-10'),

    # 10-04: Parking
    (r'lát\s*sân.*bãi\s*đỗ|sân\s*đỗ\s*xe', '10-04-01-10'),
    (r'kẻ\s*vạch|sơn\s*vạch\s*kẻ', '10-04-02-10'),
    (r'gờ\s*giảm\s*tốc', '10-04-03-10'),

    # 10-06: Traffic Signs
    (r'biển\s*báo.*giao\s*thông|biển\s*chỉ\s*dẫn', '10-06-01-10'),
    (r'biển\s*báo\s*(tam\s*giác|tròn|vuông)', '10-06-01-10'),
    (r'biển\s*báo|biến\s*báo', '10-06-01-10'),
    (r'gương\s*cầu\s*lồi', '10-06-02-10'),
    (r'sơn\s*đường|vạch\s*sơn\s*đường', '10-06-03-10'),
    (r'đinh\s*phản\s*quang', '10-06-02-10'),

    # 10-07: Bridges/Culverts
    (r'cầu\s*nội\s*bộ|cầu.*dự\s*án', '10-07-01-10'),
    (r'cống\s*hộp|cống\s*qua\s*đường', '10-07-02-10'),
    (r'cống\s*tròn|cống\s*btct', '10-07-02-10'),
    (r'cửa\s*xả', '10-07-03-10'),
    # Culvert concrete
    (r'bê\s*tông.*(cống|tường\s*cánh|sân\s*cống|thân\s*cống|bản\s*chuyển)', '10-07-02-10'),
    (r'bê\s*tông.*(đế\s*cống|dàn\s*treo|trụ\s*đỡ)', '10-07-02-10'),
    (r'ván\s*khuôn.*(cống|tường\s*cánh)', '10-07-02-10'),
    (r'nối\s*cống|chét\s*khe|sợi\s*đay|bitum', '10-07-02-10'),
    # General infrastructure concrete
    (r'phần\s*bê\s*tông', '10-07-02-10'),
    (r'phần\s*ván\s*khuôn', '10-07-02-10'),
    (r'phần\s*cốt\s*thép', '10-07-02-10'),
    (r'phần\s*xây.*trát', '10-07-02-10'),
    (r'phần\s*cống', '10-07-02-10'),
    (r'phần\s*đế\s*cống', '10-07-02-10'),
    (r'công\s*tác\s*phụ', '00-01-00-00'),

    # ========================================================================
    # GROUP 11: UTILITIES
    # ========================================================================
    # 11-01: Utility Trenching
    (r'đào\s*rãnh.*ống|đào\s*rãnh.*cáp', '11-01-01-10'),
    (r'đào\s*hố\s*ga|đào\s*hầm\s*kỹ\s*thuật', '11-01-01-10'),
    (r'lấp\s*hoàn\s*trả', '11-01-01-20'),
    (r'đắp\s*(đất|cát).*hố\s*ga', '11-01-01-20'),
    (r'đắp\s*(đất|cát).*cống', '11-01-01-20'),

    # 11-02: External Water Supply
    (r'ống.*cấp\s*nước.*ngoài|tuyến\s*ống\s*cấp\s*nước', '11-02-02-10'),
    (r'ống\s*hdpe.*cấp\s*nước', '11-02-02-10'),

    # 11-03: Stormwater
    (r'cống\s*thoát\s*mưa|cống\s*btct.*mưa', '11-03-02-10'),
    (r'hố\s*ga\s*thu\s*nước|hố\s*ga.*mưa', '11-03-03-10'),
    (r'kênh\s*mương|mương\s*thoát', '11-03-04-10'),
    (r'rãnh\s*thoát\s*nước', '11-03-04-10'),
    # Manholes
    (r'hố\s*ga|nắp\s*hố\s*ga|thành\s*hố\s*ga', '11-03-03-10'),
    (r'xây\s*hố\s*ga|trát.*hố\s*ga|láng.*hố\s*ga', '11-03-03-10'),
    (r'bê\s*tông.*hố\s*ga|bê\s*tông.*đáy\s*ga|bê\s*tông.*nắp\s*ga', '11-03-03-10'),
    (r'ván\s*khuôn.*hố\s*ga|ván\s*khuôn.*ga', '11-03-03-10'),
    (r'thép.*ga|cốt\s*thép.*ga', '11-03-03-10'),
    (r'tấm\s*đan.*ga|nắp\s*ga.*gang', '11-03-03-10'),
    (r'chèn\s*vữa.*ga|chèn.*hố\s*ga', '11-03-03-10'),

    # 11-04: Sewerage
    (r'cống.*nước\s*thải|cống\s*btct.*thải', '11-04-02-10'),
    (r'trạm\s*xử\s*lý\s*nước\s*thải|xlnt', '11-04-03-01'),
    (r'bể\s*tự\s*hoại|bể\s*phốt', '11-04-04-10'),
    (r'hố\s*ga.*thải', '11-04-03-10'),
    # PVC pipes and fittings
    (r'ống\s*u\.?pvc|ống\s*upvc', '11-04-02-10'),
    (r'cút\s*\d+\s*độ|co\s*\d+\s*độ', '11-04-02-91'),
    (r'y\s*thu|tê\s*thu|y\s*rẽ', '11-04-02-91'),

    # 11-05: External Electrical
    (r'trạm\s*biến\s*áp\s*ngoài|tba\s*ngoài', '11-05-01-01'),
    (r'cáp\s*ngầm|cáp.*ngoài\s*nhà', '11-05-02-10'),
    (r'cột\s*điện|trụ\s*điện', '11-05-03-10'),

    # 11-06: Street Lighting
    (r'đèn\s*đường|đèn\s*chiếu\s*sáng.*đường', '11-06-03-20'),
    (r'cột\s*đèn|trụ\s*đèn', '11-06-03-10'),
    (r'đèn\s*cao\s*áp|đèn\s*hps', '11-06-03-20'),
    (r'đèn\s*sân\s*vườn|đèn\s*công\s*viên', '11-06-03-10'),

    # 11-07: Telecom Infrastructure
    (r'cống\s*cáp\s*viễn\s*thông|hầm\s*cáp', '11-07-02-10'),
    (r'tủ\s*cáp|manhole', '11-07-03-10'),
    # Cable conduit pipes (HDPE for electrical cables)
    (r'ống\s*hdpe.*\d+/\d+', '11-07-02-10'),
    (r'ống\s*hdpe.*luồn|ống\s*hdpe.*cáp', '11-07-02-10'),
    # Cable markers
    (r'băng\s*báo\s*hiệu|mốc\s*báo\s*hiệu', '11-07-02-10'),

    # 11-08: Fencing
    (r'hàng\s*rào.*khu\s*vực|hàng\s*rào\s*thép', '11-08-01-10'),
    (r'cổng\s*chính|cổng\s*khu', '11-08-02-10'),
    (r'nhà\s*bảo\s*vệ', '11-08-03-10'),

    # 11-09: Retaining
    (r'kè\s*bờ|kè\s*taluy', '11-09-01-10'),
    (r'tường\s*chắn\s*đất', '11-09-02-10'),

    # ========================================================================
    # GROUP 05: LANDSCAPE
    # ========================================================================
    (r'trồng\s*cây|cây\s*xanh|cây\s*bóng\s*mát', '05-01-01-10'),
    (r'thảm\s*cỏ|trồng\s*cỏ', '05-01-02-10'),
    (r'hố\s*trồng\s*cây|xây\s*hố\s*trồng\s*cây', '05-01-01-10'),
    (r'đất\s*màu|đất\s*trồng', '05-01-01-10'),
    (r'đường\s*dạo|lối\s*đi\s*dạo', '05-02-01-10'),
    (r'hệ\s*thống\s*tưới|tưới\s*cây', '05-01-03-10'),
    (r'hồ\s*bơi|bể\s*bơi', '05-07-01-10'),

    # ========================================================================
    # GROUP 01-05: GROUND IMPROVEMENT
    # ========================================================================
    (r'cọc\s*tre|gia\s*cố.*cọc\s*tre', '01-05-01-10'),
    (r'cọc\s*cát|gia\s*cố.*cọc\s*cát', '01-05-01-10'),
    (r'bấc\s*thấm', '01-05-02-10'),
    (r'gia\s*cố\s*nền\s*đất', '01-05-01-10'),

    # ========================================================================
    # MISCELLANEOUS / FALLBACK PATTERNS
    # ========================================================================
    # More concrete/structural
    (r'bê\s*tông\s*(cửa|máng|mương)', '10-07-02-10'),
    (r'ván\s*khuôn\s*(máng|mương|bản)', '10-07-02-10'),
    (r'cốt\s*thép\s*(máng|mương)', '10-07-02-10'),
    (r'xây\s*(chân\s*khay|đá\s*hộc|gia\s*cố)', '11-09-01-10'),
    (r'khe\s*phòng\s*lún|khe\s*co\s*giãn', '10-07-02-90'),
    (r'ổ\s*khóa|cơ\s*khí', '10-07-02-91'),
    # General earthwork
    (r'đắp\s*cát|đắp\s*hoàn\s*trả|đắp\s*trả', '10-01-01-20'),
    # Testing/commissioning
    (r'kiểm\s*định|thí\s*nghiệm|test', '12-01-01-10'),
    # Welding/metal works
    (r'hàn\s*hồ\s*quang|hàn\s*điện', '02-02-04-10'),
    (r'thang\s*leo|thang\s*thép', '02-02-04-10'),
    (r'tấm\s*đan|dàn\s*tải', '01-03-01-20'),
    # Cables with specific format
    (r'cáp\s*cu/pvc|cáp\s*đồng/pvc', '04-01-02-10'),
]

# ============================================================================
# SPEC CODE EXTRACTION PATTERNS
# ============================================================================

def extract_spec_code(description):
    """
    Extract Spec Code from description based on WBS Manual rules.
    Returns tuple of (spec_code, prefix) or (None, None)
    """
    desc_lower = description.lower()

    # ========================================================================
    # PIPE SPECS: D{size}P{PN} or D{size}
    # ========================================================================
    # PPR/PVC pipes with PN rating
    match = re.search(r'd\s*n?\s*(\d{2,3})\s*(pn\s*)?(\d{1,2})?', desc_lower)
    if match and ('ống' in desc_lower or 'pipe' in desc_lower or 'ppr' in desc_lower or 'pvc' in desc_lower):
        size = match.group(1).zfill(3)
        pn = match.group(3) if match.group(3) else None
        if pn:
            return f"D{size}P{pn}", "ONG"
        return f"D{size}", "ONG"

    # Pipe by diameter only (ống d25, ống φ32)
    match = re.search(r'[ốống]\s*.*[dφ]\s*(\d{2,3})', desc_lower)
    if match:
        size = match.group(1).zfill(3)
        return f"D{size}", "ONG"

    # ========================================================================
    # CONCRETE SPECS: M{grade} or BT-M{grade}
    # ========================================================================
    match = re.search(r'm[áac]*[ckc]?\s*(\d{3})', desc_lower)
    if match and ('bê tông' in desc_lower or 'bt ' in desc_lower or 'btông' in desc_lower):
        grade = match.group(1)
        return f"M{grade}", "BT"

    # ========================================================================
    # REBAR SPECS: D{diameter} or TH-D{diameter}
    # ========================================================================
    match = re.search(r'[φd]\s*(\d{1,2})\s*(mm)?', desc_lower)
    if match and ('thép' in desc_lower or 'cốt' in desc_lower or 'sắt' in desc_lower):
        dia = match.group(1)
        return f"D{dia.zfill(2)}", "TH"

    # ========================================================================
    # CABLE SPECS: {n}C{size}
    # ========================================================================
    # Pattern: 4x10mm², 3x6mm², 3x70+1x35
    match = re.search(r'(\d+)\s*[x×]\s*(\d+\.?\d*)\s*(mm)?', desc_lower)
    if match and ('cáp' in desc_lower or 'dây' in desc_lower or 'cable' in desc_lower):
        cores = match.group(1)
        size = match.group(2)
        # Check for additional conductor (e.g., 3x70+1x35)
        extra_match = re.search(r'\+\s*(\d+)\s*[x×]\s*(\d+)', desc_lower)
        if extra_match:
            return f"{cores}C{size}+{extra_match.group(1)}C{extra_match.group(2)}", "CAP"
        return f"{cores}C{size}", "CAP"

    # ========================================================================
    # TILE SPECS: {W}x{H}
    # ========================================================================
    match = re.search(r'(\d{3,4})\s*[x×]\s*(\d{3,4})', desc_lower)
    if match and ('gạch' in desc_lower or 'lát' in desc_lower or 'ốp' in desc_lower):
        w = match.group(1)
        h = match.group(2)
        return f"{w}x{h}", "GACH"

    # ========================================================================
    # DUCT SPECS: {W}x{H} or D{diameter}
    # ========================================================================
    match = re.search(r'(\d{2,4})\s*[x×]\s*(\d{2,4})', desc_lower)
    if match and ('ống gió' in desc_lower or 'duct' in desc_lower):
        w = match.group(1)
        h = match.group(2)
        return f"{w}x{h}", "GIO"

    match = re.search(r'[dφ]\s*(\d{3})', desc_lower)
    if match and ('ống gió' in desc_lower or 'duct' in desc_lower):
        return f"D{match.group(1)}", "GIO"

    # ========================================================================
    # HVAC SPECS: {capacity}K or {capacity}HP or {capacity}RT
    # ========================================================================
    # BTU (9K, 12K, 18K, 24K)
    match = re.search(r'(\d{1,2})[.,]?(\d{3})?\s*btu', desc_lower)
    if match:
        if match.group(2):
            btu = int(match.group(1)) * 1000 + int(match.group(2))
        else:
            btu = int(match.group(1)) * 1000
        k = btu // 1000
        return f"{k}K", "DH"

    # HP
    match = re.search(r'(\d+\.?\d*)\s*hp', desc_lower)
    if match and ('vrv' in desc_lower or 'vrf' in desc_lower or 'điều hòa' in desc_lower):
        return f"{match.group(1)}HP", "DH"

    # RT (Chiller)
    match = re.search(r'(\d+)\s*rt', desc_lower)
    if match:
        return f"{match.group(1)}RT", "DH"

    # ========================================================================
    # PUMP SPECS: {flow}M3H{head}M or {power}KW
    # ========================================================================
    match = re.search(r'(\d+)\s*m[³3]/h.*?(\d+)\s*m\b', desc_lower)
    if match and 'bơm' in desc_lower:
        return f"{match.group(1)}M3H{match.group(2)}M", "BOM"

    match = re.search(r'(\d+\.?\d*)\s*kw', desc_lower)
    if match and 'bơm' in desc_lower:
        return f"{match.group(1)}KW", "BOM"

    # ========================================================================
    # ELEVATOR SPECS: {capacity}KG{speed}
    # ========================================================================
    match = re.search(r'(\d{3,4})\s*kg', desc_lower)
    if match and 'thang' in desc_lower:
        cap = match.group(1)
        speed_match = re.search(r'(\d+\.?\d*)\s*m/s', desc_lower)
        if speed_match:
            return f"{cap}KG{speed_match.group(1)}", "TM"
        return f"{cap}KG", "TM"

    # ========================================================================
    # LIGHTING SPECS: {power}W{type}
    # ========================================================================
    match = re.search(r'(\d+)\s*w', desc_lower)
    if match and ('đèn' in desc_lower or 'led' in desc_lower):
        power = match.group(1)
        if 'led' in desc_lower:
            return f"{power}WLED", "DEN"
        elif 'hps' in desc_lower or 'cao áp' in desc_lower:
            return f"{power}WHPS", "DEN"
        return f"{power}W", "DEN"

    # ========================================================================
    # COMPACTION SPECS: K95, K98
    # ========================================================================
    match = re.search(r'k\s*(95|98|90)', desc_lower)
    if match:
        return f"K{match.group(1)}", "DAT"

    # ========================================================================
    # BTN SPECS: BTNC19, BTN12.5
    # ========================================================================
    match = re.search(r'btn\s*[c]?\s*(\d+\.?\d*)', desc_lower)
    if match:
        return f"BTNC{match.group(1)}", None

    # ========================================================================
    # DOOR/WINDOW SPECS: {W}x{H}
    # ========================================================================
    match = re.search(r'(\d{3,4})\s*[x×]\s*(\d{3,4})', desc_lower)
    if match and ('cửa' in desc_lower or 'door' in desc_lower or 'window' in desc_lower):
        return f"{match.group(1)}x{match.group(2)}", "CUA"

    # ========================================================================
    # GLASS SPECS: {thickness}MM
    # ========================================================================
    match = re.search(r'(\d+)\s*mm', desc_lower)
    if match and 'kính' in desc_lower:
        return f"{match.group(1)}MM", "KINH"

    # ========================================================================
    # WATERPROOFING SPECS: {thickness}MM or {type}
    # ========================================================================
    if 'chống thấm' in desc_lower:
        match = re.search(r'(\d+\.?\d*)\s*mm', desc_lower)
        if match:
            return f"{match.group(1)}MM", "CT"
        if 'bitum' in desc_lower:
            return "BIT", "CT"
        if 'pu' in desc_lower or 'polyurethane' in desc_lower:
            return "PU", "CT"
        if 'xi măng' in desc_lower:
            return "XI", "CT"

    return None, None


def normalize_description(description, sec_code=None):
    """
    Normalize description according to WBS Manual rules.

    Format: {PREFIX}-{PARAM1}-{PARAM2}[-{PARAM3}][-{PARAM4}]

    Examples:
    - BT-M300-BC (Bê tông M300 bơm cần)
    - TH-D16-CB400 (Thép D16 CB400)
    - ONG-PPR-D25-NH-PN16 (Ống PPR D25 nối nhiệt PN16)
    - CAP-CU-4C10-PVC (Cáp đồng 4x10mm² PVC)
    - GACH-GR-600x600-PRE (Gạch granite 600x600 Premium)
    """
    desc_lower = description.lower().strip()
    parts = []

    # ========================================================================
    # CONCRETE (BT): BT-{grade}-{method}[-{additive}]
    # ========================================================================
    if any(kw in desc_lower for kw in ['bê tông', 'bt ', 'btông', 'beton']):
        parts.append('BT')

        # Grade (M200, M250, M300, M350, M400)
        grade_match = re.search(r'm[áac]*[ckc]?\s*(\d{3})', desc_lower)
        if grade_match:
            parts.append(f"M{grade_match.group(1)}")

        # Method
        if 'thủ công' in desc_lower:
            parts.append('TC')
        elif 'bơm tĩnh' in desc_lower:
            parts.append('BT')
        elif 'bơm' in desc_lower or 'thương phẩm' in desc_lower:
            parts.append('BC')

        # Additives
        if 'chống thấm' in desc_lower:
            parts.append('CT')
        if 'phụ gia nở' in desc_lower:
            parts.append('PN')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # REBAR (TH): TH-D{diameter}-{grade}
    # ========================================================================
    if any(kw in desc_lower for kw in ['cốt thép', 'thép d', 'sản xuất, lắp dựng cốt', 'thép ø', 'thép φ', 'thép thanh vằn', 'thanh vằn']) or \
       (any(kw in desc_lower for kw in ['thép', 'sắt']) and any(kw in desc_lower for kw in ['móng', 'đài', 'cột', 'dầm', 'sàn', 'giằng', 'vách', 'tường', 'hố ga', 'cống', 'bản', 'nắp'])):
        parts.append('TH')

        # Diameter
        dia_match = re.search(r'[φødD]\s*(\d{1,2})\s*(mm)?', desc_lower)
        if not dia_match:
            dia_match = re.search(r'≤\s*(\d{1,2})\s*mm', desc_lower)
        if not dia_match:
            dia_match = re.search(r'>\s*(\d{1,2})\s*mm', desc_lower)
        if not dia_match:
            dia_match = re.search(r'đường kính\s*[≤>]?\s*(\d{1,2})', desc_lower)
        if dia_match:
            parts.append(f"D{dia_match.group(1)}")

        # Grade
        if 'cb500' in desc_lower or 'cb 500' in desc_lower:
            parts.append('CB500')
        elif 'cb400' in desc_lower or 'cb 400' in desc_lower:
            parts.append('CB400')
        elif 'cb300' in desc_lower or 'cb 300' in desc_lower:
            parts.append('CB300')
        elif 'cb240' in desc_lower or 'cb 240' in desc_lower:
            parts.append('CB240')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # PIPE (ONG): ONG-{material}-D{size}-{connection}[-{pressure}]
    # ========================================================================
    # Use word boundary to avoid matching "ống" inside "chống"
    if re.search(r'\bống\b', desc_lower) and any(kw in desc_lower for kw in ['ppr', 'pvc', 'hdpe', 'thép', 'inox', 'cấp', 'thoát']):
        parts.append('ONG')

        # Material
        if 'ppr' in desc_lower or 'pp-r' in desc_lower:
            parts.append('PPR')
        elif 'pvc' in desc_lower:
            parts.append('PVC')
        elif 'hdpe' in desc_lower:
            parts.append('HDPE')
        elif 'inox' in desc_lower or 'ss' in desc_lower:
            parts.append('SS')
        elif 'thép mạ' in desc_lower or 'mạ kẽm' in desc_lower:
            parts.append('GI')
        elif 'thép đen' in desc_lower:
            parts.append('CS')

        # Size
        size_match = re.search(r'd\s*n?\s*(\d{2,3})', desc_lower)
        if size_match:
            parts.append(f"D{size_match.group(1)}")

        # Connection
        if 'nhiệt' in desc_lower or 'hàn nhiệt' in desc_lower:
            parts.append('NH')
        elif 'dán' in desc_lower:
            parts.append('DAN')
        elif 'ren' in desc_lower:
            parts.append('REN')
        elif 'hàn' in desc_lower:
            parts.append('HAN')
        elif 'bích' in desc_lower or 'mặt bích' in desc_lower:
            parts.append('FL')

        # Pressure
        pn_match = re.search(r'pn\s*(\d{1,2})', desc_lower)
        if pn_match:
            parts.append(f"PN{pn_match.group(1)}")

        if len(parts) >= 3:
            return '-'.join(parts)

    # ========================================================================
    # CABLE (CAP): CAP-{conductor}-{size}-{insulation}[-{armour}]
    # ========================================================================
    if any(kw in desc_lower for kw in ['cáp', 'dây điện', 'cable']):
        parts.append('CAP')

        # Conductor
        if 'nhôm' in desc_lower or 'al' in desc_lower:
            parts.append('AL')
        else:
            parts.append('CU')

        # Size (cores x area)
        size_match = re.search(r'(\d+)\s*[x×]\s*(\d+\.?\d*)', desc_lower)
        if size_match:
            cores = size_match.group(1)
            area = size_match.group(2)
            # Check for additional conductor
            extra_match = re.search(r'\+\s*(\d+)\s*[x×]\s*(\d+)', desc_lower)
            if extra_match:
                parts.append(f"{cores}C{area}+{extra_match.group(1)}C{extra_match.group(2)}")
            else:
                parts.append(f"{cores}C{area}")

        # Insulation
        if 'chống cháy' in desc_lower or 'fr' in desc_lower:
            parts.append('FR')
        elif 'lszh' in desc_lower or 'nhls' in desc_lower or 'ít khói' in desc_lower:
            parts.append('LSZH')
        elif 'xlpe' in desc_lower:
            parts.append('XLPE')
        else:
            parts.append('PVC')

        # Armour
        if 'giáp' in desc_lower or 'swa' in desc_lower:
            parts.append('SWA')

        if len(parts) >= 3:
            return '-'.join(parts)

    # ========================================================================
    # TILE (GACH): GACH-{type}-{size}-{grade}
    # ========================================================================
    if any(kw in desc_lower for kw in ['gạch', 'lát nền', 'ốp tường']):
        parts.append('GACH')

        # Type
        if 'granite' in desc_lower or 'granit' in desc_lower:
            parts.append('GR')
        elif 'porcelain' in desc_lower:
            parts.append('POR')
        elif 'terracotta' in desc_lower:
            parts.append('TER')
        else:
            parts.append('MEN')

        # Size
        size_match = re.search(r'(\d{3,4})\s*[x×]\s*(\d{3,4})', desc_lower)
        if size_match:
            parts.append(f"{size_match.group(1)}x{size_match.group(2)}")

        # Grade
        if 'cao cấp' in desc_lower or 'premium' in desc_lower:
            parts.append('PRE')
        elif 'luxury' in desc_lower or 'sang trọng' in desc_lower:
            parts.append('LUX')
        else:
            parts.append('STD')

        if len(parts) >= 3:
            return '-'.join(parts)

    # ========================================================================
    # EARTHWORK (DAT): DAT-{action}-{grade}-{method}
    # ========================================================================
    if any(kw in desc_lower for kw in ['đào đất', 'đắp đất', 'san nền', 'đào hố', 'đào móng', 'đào nền', 'đào khuôn', 'đào phá']):
        parts.append('DAT')

        # Action
        if 'phá dỡ' in desc_lower:
            parts.append('PHADO')
        elif 'đào' in desc_lower:
            parts.append('DAO')
        elif 'đắp' in desc_lower:
            parts.append('DAP')
        elif 'san' in desc_lower:
            parts.append('SAN')

        # Grade/Compaction
        k_match = re.search(r'k\s*(90|95|98)', desc_lower)
        if k_match:
            parts.append(f"K{k_match.group(1)}")
        elif 'cấp i' in desc_lower or 'cấp 1' in desc_lower or 'cấp ii' in desc_lower or 'cấp 2' in desc_lower:
            parts.append('C1')
        elif 'cấp iii' in desc_lower or 'cấp 3' in desc_lower or 'cấp iv' in desc_lower or 'cấp 4' in desc_lower:
            parts.append('C3')
        elif 'đá' in desc_lower:
            parts.append('DA')

        # Method
        if 'thủ công' in desc_lower:
            parts.append('TC')
        elif 'máy' in desc_lower or 'cơ giới' in desc_lower:
            parts.append('MAY')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # HVAC (DH): DH-{type}-{capacity}[-{model}]
    # ========================================================================
    if any(kw in desc_lower for kw in ['điều hòa', 'điều hoà', 'máy lạnh', 'chiller', 'vrv', 'vrf']):
        parts.append('DH')

        # Type
        if 'chiller' in desc_lower:
            parts.append('CHILLER')
        elif 'vrv' in desc_lower or 'vrf' in desc_lower:
            parts.append('VRV')
        elif 'cassette' in desc_lower:
            parts.append('CASSETTE')
        elif 'âm trần' in desc_lower:
            parts.append('AM')
        elif 'tủ đứng' in desc_lower:
            parts.append('TU')
        else:
            parts.append('SPLIT')

        # Capacity - handle BTU with or without space/separator
        btu_match = re.search(r'(\d{1,2})[.,]?(\d{3})?\s*btu', desc_lower)
        if not btu_match:
            # Handle format like "36000BTU" (no separator)
            btu_match = re.search(r'(\d{4,5})\s*btu', desc_lower)
            if btu_match:
                btu = int(btu_match.group(1))
                parts.append(f"{btu // 1000}K")
        elif btu_match:
            if btu_match.group(2):
                btu = int(btu_match.group(1)) * 1000 + int(btu_match.group(2))
            else:
                btu = int(btu_match.group(1)) * 1000
            parts.append(f"{btu // 1000}K")

        hp_match = re.search(r'(\d+\.?\d*)\s*hp', desc_lower)
        if hp_match:
            parts.append(f"{hp_match.group(1)}HP")

        rt_match = re.search(r'(\d+)\s*rt', desc_lower)
        if rt_match:
            parts.append(f"{rt_match.group(1)}RT")

        # Model
        if 'inverter' in desc_lower:
            parts.append('INV')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # PUMP (BOM): BOM-{type}-{flow}M3-{head}M
    # ========================================================================
    if 'bơm' in desc_lower:
        parts.append('BOM')

        # Type
        if 'pccc' in desc_lower or 'chữa cháy' in desc_lower or 'cứu hỏa' in desc_lower:
            parts.append('PCCC')
        elif 'tăng áp' in desc_lower:
            parts.append('TA')
        elif 'thoát' in desc_lower or 'thải' in desc_lower:
            parts.append('TN')
        else:
            parts.append('CN')

        # Flow and Head
        flow_match = re.search(r'(\d+)\s*m[³3]/h', desc_lower)
        head_match = re.search(r'(\d+)\s*m\b', desc_lower)
        if flow_match:
            parts.append(f"{flow_match.group(1)}M3")
        if head_match and 'cột áp' in desc_lower:
            parts.append(f"{head_match.group(1)}M")

        # Power
        kw_match = re.search(r'(\d+\.?\d*)\s*kw', desc_lower)
        if kw_match:
            parts.append(f"{kw_match.group(1)}KW")

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # LIGHTING (DEN): DEN-{type}-{power}W[-{tech}]
    # ========================================================================
    if any(kw in desc_lower for kw in ['đèn', 'chiếu sáng']):
        parts.append('DEN')

        # Type
        if 'âm trần' in desc_lower or 'downlight' in desc_lower:
            parts.append('AM')
        elif 'panel' in desc_lower or 'tấm' in desc_lower:
            parts.append('PANEL')
        elif 'tuýp' in desc_lower or 'tube' in desc_lower:
            parts.append('TUBE')
        elif 'đường' in desc_lower or 'cao áp' in desc_lower:
            parts.append('DUONG')
        elif 'pha' in desc_lower or 'rọi' in desc_lower:
            parts.append('PHA')
        else:
            parts.append('NOI')

        # Power
        w_match = re.search(r'(\d+)\s*w', desc_lower)
        if w_match:
            parts.append(f"{w_match.group(1)}W")

        # Tech
        if 'led' in desc_lower:
            parts.append('LED')
        elif 'hps' in desc_lower or 'sodium' in desc_lower:
            parts.append('HPS')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # WATERPROOFING (CT): CT-{type}-{thickness}[-{layers}L]
    # ========================================================================
    if 'chống thấm' in desc_lower:
        parts.append('CT')

        # Type
        if 'bitum' in desc_lower:
            parts.append('BIT')
        elif 'pvc' in desc_lower or 'màng pvc' in desc_lower:
            parts.append('PVC')
        elif 'pu' in desc_lower or 'polyurethane' in desc_lower:
            parts.append('PU')
        elif 'xi măng' in desc_lower or 'gốc xi măng' in desc_lower:
            parts.append('XI')
        elif 'kết tinh' in desc_lower:
            parts.append('KT')

        # Thickness
        thick_match = re.search(r'(\d+\.?\d*)\s*mm', desc_lower)
        if thick_match:
            parts.append(f"{thick_match.group(1)}MM")

        # Layers
        layer_match = re.search(r'(\d+)\s*lớp', desc_lower)
        if layer_match:
            parts.append(f"{layer_match.group(1)}L")

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # FORMWORK (VK): VK-{material}-{height}
    # ========================================================================
    if 'ván khuôn' in desc_lower:
        parts.append('VK')

        # Material
        if 'gỗ' in desc_lower:
            parts.append('GO')
        elif 'thép' in desc_lower:
            parts.append('THEP')
        elif 'nhôm' in desc_lower:
            parts.append('NHOM')
        elif 'nhựa' in desc_lower:
            parts.append('NHUA')

        # Height (if mentioned) - parse from text like "h=8m", "h<=16", "cao 16m"
        height_match = re.search(r'[h=]\s*(\d{1,2})\s*m?\b', desc_lower)
        if not height_match:
            height_match = re.search(r'cao\s*(\d{1,2})\s*m?\b', desc_lower)

        if height_match:
            height = int(height_match.group(1))
            if height > 16:
                parts.append('H50')
            elif height > 8:
                parts.append('H16')
            elif height > 4:
                parts.append('H8')
            else:
                parts.append('H4')
        elif 'cao tầng' in desc_lower or 'h>16' in desc_lower:
            parts.append('H50')
        elif 'h16' in desc_lower or 'h<=16' in desc_lower:
            parts.append('H16')
        elif 'h8' in desc_lower or 'h<=8' in desc_lower:
            parts.append('H8')
        else:
            parts.append('H4')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # PAINT (SON): SON-{type}-{coats}L[-{location}]
    # ========================================================================
    if 'sơn' in desc_lower and not 'ống' in desc_lower:
        parts.append('SON')

        # Type
        if 'lót' in desc_lower:
            parts.append('LOT')
        elif 'epoxy' in desc_lower:
            parts.append('EP')
        elif 'pu' in desc_lower:
            parts.append('PU')
        elif 'chống thấm' in desc_lower:
            parts.append('CT')
        elif 'dầu' in desc_lower:
            parts.append('DU')
        else:
            parts.append('NC')

        # Coats
        coat_match = re.search(r'(\d+)\s*lớp', desc_lower)
        if coat_match:
            parts.append(f"{coat_match.group(1)}L")

        # Location
        if 'ngoài' in desc_lower or 'ngoại thất' in desc_lower:
            parts.append('NGOAI')
        elif 'trong' in desc_lower or 'nội thất' in desc_lower:
            parts.append('TRONG')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # DOOR (CUA): CUA-{type}-{material}-{size}[-{special}]
    # ========================================================================
    # Avoid matching pipe-related descriptions (ống as a word, not in chống)
    is_pipe_desc = bool(re.search(r'\bống\b', desc_lower)) or 'cửa xả' in desc_lower
    if 'cửa' in desc_lower and not is_pipe_desc:
        parts.append('CUA')

        # Type
        if 'sổ' in desc_lower or 'sổ' in desc_lower:
            parts.append('SO')
        elif 'trượt' in desc_lower:
            parts.append('TRUOT')
        else:
            parts.append('DI')

        # Material
        if 'gỗ' in desc_lower or 'hdf' in desc_lower or 'mdf' in desc_lower:
            parts.append('GO')
        elif 'nhôm' in desc_lower:
            parts.append('NHOM')
        elif 'nhựa' in desc_lower:
            parts.append('NHUA')
        elif 'thép' in desc_lower or 'sắt' in desc_lower:
            parts.append('THEP')

        # Size
        size_match = re.search(r'(\d{3,4})\s*[x×]\s*(\d{3,4})', desc_lower)
        if size_match:
            parts.append(f"{size_match.group(1)}x{size_match.group(2)}")

        # Special
        if 'chống cháy' in desc_lower:
            parts.append('CC')
        elif 'cách âm' in desc_lower:
            parts.append('AM')

        if len(parts) >= 3:
            return '-'.join(parts)

    # ========================================================================
    # ASPHALT (BTN): BTN-{type}-{thickness}CM
    # ========================================================================
    if any(kw in desc_lower for kw in ['bê tông nhựa', 'btn', 'nhựa đường', 'rải thảm']):
        parts.append('BTN')

        # Type
        if 'c19' in desc_lower:
            parts.append('C19')
        elif 'c12.5' in desc_lower or 'c12,5' in desc_lower:
            parts.append('C12')
        elif 'c9.5' in desc_lower or 'c9,5' in desc_lower:
            parts.append('C9')
        elif 'polymer' in desc_lower:
            parts.append('PMB')

        # Thickness
        thick_match = re.search(r'(\d+)\s*cm', desc_lower)
        if thick_match:
            parts.append(f"{thick_match.group(1)}CM")

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # AGGREGATE BASE (CPDD): CPDD-{type}[-{layer}]
    # ========================================================================
    if any(kw in desc_lower for kw in ['cấp phối đá dăm', 'cpđd', 'đá dăm', 'móng đường']):
        parts.append('CPDD')

        # Type
        if 'loại 1' in desc_lower or 'type 1' in desc_lower or 'lớp trên' in desc_lower:
            parts.append('L1')
        elif 'loại 2' in desc_lower or 'type 2' in desc_lower or 'lớp dưới' in desc_lower:
            parts.append('L2')
        else:
            parts.append('STD')

        # Thickness
        thick_match = re.search(r'(\d+)\s*cm', desc_lower)
        if thick_match:
            parts.append(f"{thick_match.group(1)}CM")

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # CURB (BOVIA): BOVIA-{material}-{size}
    # ========================================================================
    if any(kw in desc_lower for kw in ['bó vỉa', 'bó hè', 'bó đường', 'bo via']):
        parts.append('BOVIA')

        # Material
        if 'đá' in desc_lower:
            parts.append('DA')
        elif 'bê tông' in desc_lower:
            parts.append('BT')
        elif 'gạch' in desc_lower:
            parts.append('GACH')

        # Size
        size_match = re.search(r'(\d+)\s*[x×]\s*(\d+)\s*[x×]?\s*(\d+)?', desc_lower)
        if size_match:
            if size_match.group(3):
                parts.append(f"{size_match.group(1)}x{size_match.group(2)}x{size_match.group(3)}")
            else:
                parts.append(f"{size_match.group(1)}x{size_match.group(2)}")

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # GEOTEXTILE (VAI): VAI-{type}-{weight}GSM
    # ========================================================================
    if any(kw in desc_lower for kw in ['vải địa', 'geotextile', 'vải không dệt']):
        parts.append('VAI')

        # Type
        if 'dệt' in desc_lower and 'không dệt' not in desc_lower:
            parts.append('DET')
        else:
            parts.append('KD')  # Không dệt (non-woven)

        # Weight (gsm)
        gsm_match = re.search(r'(\d+)\s*g(?:/m|sm)?', desc_lower)
        if gsm_match:
            parts.append(f"{gsm_match.group(1)}GSM")

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # DRAIN COVER (TAMDAN): TAMDAN-{material}-{size}
    # ========================================================================
    if any(kw in desc_lower for kw in ['tấm đan', 'nắp hố ga', 'nắp đậy']):
        parts.append('TAMDAN')

        # Material
        if 'đá' in desc_lower:
            parts.append('DA')
        elif 'gang' in desc_lower:
            parts.append('GANG')
        elif 'bê tông' in desc_lower or 'btct' in desc_lower:
            parts.append('BT')
        elif 'thép' in desc_lower or 'inox' in desc_lower:
            parts.append('THEP')

        # Size
        size_match = re.search(r'(\d+)\s*[x×]\s*(\d+)', desc_lower)
        if size_match:
            parts.append(f"{size_match.group(1)}x{size_match.group(2)}")

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # CULVERT (CONG): CONG-{type}-D{size}
    # ========================================================================
    if any(kw in desc_lower for kw in ['cống', 'culvert']):
        parts.append('CONG')

        # Type
        if 'hộp' in desc_lower:
            parts.append('HOP')
        elif 'tròn' in desc_lower:
            parts.append('TRON')
        elif 'btct' in desc_lower or 'bê tông' in desc_lower:
            parts.append('BTCT')
        else:
            parts.append('STD')

        # Size
        dia_match = re.search(r'd\s*(\d{2,4})', desc_lower)
        if dia_match:
            parts.append(f"D{dia_match.group(1)}")

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # MANHOLE (HOGA): HOGA-{type}-{size}
    # ========================================================================
    if any(kw in desc_lower for kw in ['hố ga', 'hố thu', 'hố thăm']):
        parts.append('HOGA')

        # Type
        if 'thu' in desc_lower:
            parts.append('THU')
        elif 'thăm' in desc_lower:
            parts.append('THAM')
        else:
            parts.append('STD')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # TRAFFIC SIGN (BIENB): BIENB-{type}-{size}
    # ========================================================================
    if any(kw in desc_lower for kw in ['biển báo', 'biển chỉ', 'biến báo']):
        parts.append('BIENB')

        # Type
        if 'tam giác' in desc_lower:
            parts.append('TAMGIAC')
        elif 'tròn' in desc_lower:
            parts.append('TRON')
        elif 'vuông' in desc_lower:
            parts.append('VUONG')
        elif 'chỉ hướng' in desc_lower:
            parts.append('HUONG')

        # Size
        size_match = re.search(r'(\d+)\s*cm', desc_lower)
        if size_match:
            parts.append(f"{size_match.group(1)}CM")

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # LANDSCAPING (CAYXANH): CAYXANH-{type}
    # ========================================================================
    if any(kw in desc_lower for kw in ['trồng cây', 'cây xanh', 'cỏ', 'thảm cỏ', 'hố trồng cây']):
        parts.append('CAYXANH')

        # Type
        if 'cỏ' in desc_lower:
            parts.append('CO')
        elif 'bóng mát' in desc_lower:
            parts.append('BONGMAT')
        elif 'bụi' in desc_lower:
            parts.append('BUI')
        else:
            parts.append('CAY')

        if len(parts) >= 1:
            return '-'.join(parts)

    # ========================================================================
    # MASONRY (XAY): XAY-{material}-{size}
    # ========================================================================
    if any(kw in desc_lower for kw in ['xây gạch', 'xây tường', 'xây bó', 'xây hố']):
        parts.append('XAY')

        # Material
        if 'đặc' in desc_lower:
            parts.append('DAC')
        elif 'rỗng' in desc_lower or 'ống' in desc_lower:
            parts.append('RONG')
        elif 'bê tông' in desc_lower or 'block' in desc_lower:
            parts.append('BLOCK')
        else:
            parts.append('STD')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # PLASTERING (TRAT): TRAT-{type}[-{thickness}]
    # ========================================================================
    if any(kw in desc_lower for kw in ['trát', 'tô', 'vữa trát']):
        parts.append('TRAT')

        # Type
        if 'xi măng' in desc_lower:
            parts.append('XM')
        elif 'vôi' in desc_lower:
            parts.append('VOI')
        elif 'granitô' in desc_lower or 'granito' in desc_lower:
            parts.append('GR')
        else:
            parts.append('STD')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # TRANSPORTATION (VANCHUYEN): VANCHUYEN-{type}
    # ========================================================================
    if any(kw in desc_lower for kw in ['vận chuyển', 'van chuyen']):
        parts.append('VANCHUYEN')

        if 'đất' in desc_lower:
            parts.append('DAT')
        elif 'phế thải' in desc_lower:
            parts.append('PHE')
        elif 'vật liệu' in desc_lower:
            parts.append('VL')
        else:
            parts.append('STD')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # BAMBOO PILES (COCTRE): COCTRE-{type}
    # ========================================================================
    if any(kw in desc_lower for kw in ['cọc tre', 'thi cọc tre']):
        parts.append('COCTRE')

        if 'gia cố' in desc_lower:
            parts.append('GC')
        else:
            parts.append('STD')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # STONE MASONRY (DAHOC): DAHOC-{type}
    # ========================================================================
    if any(kw in desc_lower for kw in ['đá hộc', 'xây đá']):
        parts.append('DAHOC')

        if 'chân khay' in desc_lower:
            parts.append('CHANKHAY')
        elif 'gia cố mái' in desc_lower:
            parts.append('MAI')
        else:
            parts.append('STD')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # ASPHALT TACK COAT (TUOILOP): TUOILOP-{type}
    # ========================================================================
    if any(kw in desc_lower for kw in ['tưới lớp', 'lớp thấm bám', 'nhựa pha dầu']):
        parts.append('TUOILOP')

        if 'thấm bám' in desc_lower:
            parts.append('THAMBAM')
        elif 'dính bám' in desc_lower:
            parts.append('DINHBAM')
        else:
            parts.append('STD')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # NYLON/PLASTIC SHEET (NYLON): NYLON-{type}
    # ========================================================================
    if any(kw in desc_lower for kw in ['nilon', 'nylon', 'bạt', 'màng chống thấm']):
        parts.append('NYLON')

        if 'tái sinh' in desc_lower:
            parts.append('TS')
        elif 'hdpe' in desc_lower:
            parts.append('HDPE')
        else:
            parts.append('STD')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # STEEL LADDER/STRUCTURE (THEPKC): THEPKC-{type}
    # ========================================================================
    if any(kw in desc_lower for kw in ['thang thép', 'lan can thép', 'cầu thang thép', 'kết cấu thép']):
        parts.append('THEPKC')

        if 'thang' in desc_lower:
            parts.append('THANG')
        elif 'lan can' in desc_lower:
            parts.append('LANCAN')
        else:
            parts.append('KC')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # REFLECTIVE MARKER (PHANQUANG): PHANQUANG-{type}
    # ========================================================================
    if any(kw in desc_lower for kw in ['phản quang', 'đinh phản quang', 'tiêu phản quang']):
        parts.append('PHANQUANG')

        if 'đinh' in desc_lower:
            parts.append('DINH')
        elif 'tiêu' in desc_lower:
            parts.append('TIEU')
        else:
            parts.append('STD')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # EXPANSION JOINT (KHECO): KHECO-{type}
    # ========================================================================
    if any(kw in desc_lower for kw in ['khe co', 'khe giãn', 'khe phòng lún', 'khe nối']):
        parts.append('KHECO')

        if 'phòng lún' in desc_lower:
            parts.append('LUN')
        elif 'giãn' in desc_lower or 'giản' in desc_lower:
            parts.append('GIAN')
        elif 'nối' in desc_lower:
            parts.append('NOI')
        else:
            parts.append('STD')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # OUTLET/DISCHARGE (CUAXA): CUAXA-{type}
    # ========================================================================
    if any(kw in desc_lower for kw in ['cửa xả', 'cửa điều tiết']):
        parts.append('CUAXA')

        if 'điều tiết' in desc_lower:
            parts.append('DT')
        else:
            parts.append('XA')

        if len(parts) >= 2:
            return '-'.join(parts)

    # ========================================================================
    # BOLT/ANCHOR (BULONG): BULONG-{type}-D{size}
    # ========================================================================
    if any(kw in desc_lower for kw in ['bu lông', 'bulong', 'bu-lông', 'đai ốc']):
        parts.append('BULONG')

        # Size
        dia_match = re.search(r'd\s*(\d{1,2})', desc_lower)
        if dia_match:
            parts.append(f"D{dia_match.group(1)}")

        if len(parts) >= 2:
            return '-'.join(parts)

    # If no specific pattern matched, return None (will use original)
    return None


def get_sec_code(description):
    """Get SEC code for a description by matching against patterns."""
    desc_lower = description.lower().strip()

    for pattern, code in SEC_CODE_PATTERNS:
        if re.search(pattern, desc_lower, re.IGNORECASE):
            return code

    return None


def process_excel_file(input_path, output_path=None):
    """
    Process an Excel BOQ file and generate WBS codes.

    Args:
        input_path: Path to input Excel file
        output_path: Path to output Excel file (optional)

    Returns:
        dict with processing statistics
    """
    if output_path is None:
        input_p = Path(input_path)
        output_path = input_p.parent / f"{input_p.stem}_wbs.xlsx"

    print(f"Loading workbook: {input_path}")
    wb = load_workbook(input_path)

    # Create results sheet
    if "Processing Results" in wb.sheetnames:
        del wb["Processing Results"]

    results_ws = wb.create_sheet("Processing Results", 0)

    # Headers
    headers = [
        "Row", "Sheet", "Original Description", "Normalized",
        "SEC Code", "Spec Code", "Full Code", "Unit", "Qty", "Link"
    ]

    # Style headers
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = results_ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Column widths
    widths = [8, 15, 60, 60, 15, 20, 30, 8, 12, 15]
    for col, width in enumerate(widths, 1):
        results_ws.column_dimensions[get_column_letter(col)].width = width

    # Process each sheet
    stats = {
        'total_items': 0,
        'with_full_code': 0,
        'with_sec_only': 0,
        'no_code': 0,
        'skipped_no_unit': 0
    }

    result_row = 2

    for sheet_name in wb.sheetnames:
        if sheet_name == "Processing Results":
            continue

        ws = wb[sheet_name]
        print(f"  Processing sheet: {sheet_name}")

        # Find header row and columns
        desc_col = None
        spec_col = None  # Technical specifications column
        unit_col = None
        qty_col = None
        header_row = 1

        for row in range(1, min(20, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                if val:
                    val_str = str(val).lower().strip()
                    # Remove newlines and extra spaces for matching
                    val_str_clean = ' '.join(val_str.split())
                    # Technical specs column (thông số kỹ thuật, mô tả, quy cách)
                    if any(kw in val_str_clean for kw in ['thông số kỹ thuật', 'quy cách', 'specifications',
                                                         'mô tả công việc', 'diễn giải công việc']):
                        spec_col = col
                        header_row = row
                    # Description/Item name columns
                    elif any(kw in val_str_clean for kw in ['mô tả', 'diễn giải', 'description',
                                                     'nội dung công việc', 'tên hạng mục',
                                                     'hạng mục', 'công việc']):
                        desc_col = col
                        header_row = row
                    # Unit columns - use partial match
                    elif any(kw in val_str_clean for kw in ['đơn vị', 'đvt', 'unit', 'đ.vị']):
                        unit_col = col
                    # Quantity columns
                    elif any(kw in val_str_clean for kw in ['khối lượng', 'k.lượng', 'qty',
                                                       'quantity', 'sl', 'số lượng']):
                        qty_col = col

        if not desc_col and not spec_col:
            continue

        # Process rows
        for row in range(header_row + 1, ws.max_row + 1):
            # Get item name/description
            desc = None
            if desc_col:
                desc_val = ws.cell(row=row, column=desc_col).value
                if desc_val:
                    desc = str(desc_val).strip()

            # Get technical specifications
            specs = None
            if spec_col:
                spec_val = ws.cell(row=row, column=spec_col).value
                if spec_val:
                    specs = str(spec_val).strip()

            # Combine description and specs for processing
            full_desc = ""
            if desc:
                full_desc = desc
            if specs and specs not in ['\n', '', '-']:
                full_desc = f"{full_desc} {specs}".strip() if full_desc else specs

            if not full_desc or len(full_desc) < 5:
                continue

            # Get unit
            unit = None
            if unit_col:
                unit_val = ws.cell(row=row, column=unit_col).value
                if unit_val:
                    unit = str(unit_val).strip()

            # Skip items without unit (likely headers/sections)
            if not unit:
                stats['skipped_no_unit'] += 1
                continue

            # Get quantity
            qty = None
            if qty_col:
                qty_val = ws.cell(row=row, column=qty_col).value
                if qty_val:
                    try:
                        qty = float(qty_val)
                    except:
                        qty = qty_val

            stats['total_items'] += 1

            # Get SEC code from combined description + specs
            sec_code = get_sec_code(full_desc)

            # Get Spec code from combined description + specs
            spec_code, _ = extract_spec_code(full_desc)

            # Generate Full code
            full_code = None
            if sec_code and spec_code:
                full_code = f"{sec_code}.{spec_code}"
                stats['with_full_code'] += 1
            elif sec_code:
                stats['with_sec_only'] += 1
            else:
                stats['no_code'] += 1

            # Use original description for display, keep specs separate
            display_desc = desc if desc else specs
            normalized = display_desc  # Keep original description for line items

            # Write result row
            results_ws.cell(row=result_row, column=1, value=row)
            results_ws.cell(row=result_row, column=2, value=sheet_name)
            results_ws.cell(row=result_row, column=3, value=full_desc)  # Combined desc + specs
            results_ws.cell(row=result_row, column=4, value=normalized)
            results_ws.cell(row=result_row, column=5, value=sec_code or "")
            results_ws.cell(row=result_row, column=6, value=spec_code or "")
            results_ws.cell(row=result_row, column=7, value=full_code or "")
            results_ws.cell(row=result_row, column=8, value=unit or "")
            results_ws.cell(row=result_row, column=9, value=qty or "")

            # Add hyperlink to original row
            link_cell = results_ws.cell(row=result_row, column=10, value="Go to source")
            link_cell.hyperlink = f"#'{sheet_name}'!A{row}"
            link_cell.font = Font(color="0563C1", underline="single")

            # Color coding based on match status
            if full_code:
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
            elif sec_code:
                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
            else:
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red

            for col in range(1, 10):
                results_ws.cell(row=result_row, column=col).fill = fill
                results_ws.cell(row=result_row, column=col).border = thin_border

            result_row += 1

    # Save workbook
    print(f"Saving to: {output_path}")
    wb.save(output_path)

    return stats


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("Usage: python wbs_processor.py <input_excel> [output_excel]")
        print("\nThis script processes BOQ Excel files and generates WBS codes.")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    if not Path(input_path).exists():
        print(f"Error: File not found: {input_path}")
        sys.exit(1)

    stats = process_excel_file(input_path, output_path)

    print("\n" + "=" * 60)
    print("PROCESSING RESULTS")
    print("=" * 60)
    print(f"Total items processed: {stats['total_items']}")
    print(f"  With Full Code (SEC + Spec): {stats['with_full_code']} ({stats['with_full_code']*100/max(stats['total_items'],1):.1f}%)")
    print(f"  With SEC Code only: {stats['with_sec_only']} ({stats['with_sec_only']*100/max(stats['total_items'],1):.1f}%)")
    print(f"  No Code: {stats['no_code']} ({stats['no_code']*100/max(stats['total_items'],1):.1f}%)")
    print(f"  Skipped (no unit): {stats['skipped_no_unit']}")
    print("=" * 60)


if __name__ == "__main__":
    main()
